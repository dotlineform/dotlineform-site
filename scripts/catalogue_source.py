from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping

try:
    from series_ids import normalize_series_id, parse_series_ids
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.series_ids import normalize_series_id, parse_series_ids


DEFAULT_WORKBOOK_PATH = Path("data/works.xlsx")
DEFAULT_SOURCE_DIR = Path("assets/studio/data/catalogue")

ACTIONABLE_STATUSES = {"draft", "published"}

SCHEMAS = {
    "works": "catalogue_source_works_v1",
    "work_details": "catalogue_source_work_details_v1",
    "series": "catalogue_source_series_v1",
    "work_files": "catalogue_source_work_files_v1",
    "work_links": "catalogue_source_work_links_v1",
    "meta": "catalogue_source_meta_v1",
}

SOURCE_FILES = {
    "works": "works.json",
    "work_details": "work_details.json",
    "series": "series.json",
    "work_files": "work_files.json",
    "work_links": "work_links.json",
    "meta": "meta.json",
}

WORK_FIELDS = [
    "work_id",
    "status",
    "published_date",
    "series_ids",
    "series_title",
    "project_folder",
    "project_filename",
    "title",
    "width_cm",
    "height_cm",
    "year",
    "year_display",
    "medium_type",
    "medium_caption",
    "storage_location",
    "work_prose_file",
    "notes",
    "duration",
    "depth_cm",
    "width_px",
    "height_px",
    "provenance",
    "artist",
    "downloads",
    "links",
]

SERIES_FIELDS = [
    "series_id",
    "title",
    "series_type",
    "status",
    "published_date",
    "year",
    "year_display",
    "primary_work_id",
    "series_prose_file",
    "notes",
    "sort_fields",
]

DETAIL_FIELDS = [
    "detail_uid",
    "work_id",
    "detail_id",
    "project_subfolder",
    "project_filename",
    "title",
    "status",
    "published_date",
    "width_px",
    "height_px",
]

FILE_FIELDS = [
    "file_uid",
    "work_id",
    "filename",
    "label",
    "status",
    "published_date",
]

LINK_FIELDS = [
    "link_uid",
    "work_id",
    "url",
    "label",
    "status",
    "published_date",
]

DOWNLOAD_FIELDS = ["filename", "label"]
WORK_LINK_ENTRY_FIELDS = ["url", "label"]

WORKBOOK_HEADERS = {
    "Works": [field for field in WORK_FIELDS if field not in {"downloads", "links"}],
    "Series": [
        "series_id",
        "title",
        "series_type",
        "status",
        "published_date",
        "year",
        "year_display",
        "primary_work_id",
        "series_prose_file",
        "notes",
    ],
    "SeriesSort": ["series_id", "sort_fields"],
    "WorkDetails": [
        "work_id",
        "detail_id",
        "project_subfolder",
        "project_filename",
        "title",
        "status",
        "published_date",
        "width_px",
        "height_px",
    ],
    "WorkFiles": ["work_id", "filename", "label", "status", "published_date"],
    "WorkLinks": ["work_id", "url", "label", "status", "published_date"],
}

WORK_TEXT_FIELDS = set(WORK_FIELDS) - {
    "series_ids",
    "downloads",
    "links",
    "width_cm",
    "height_cm",
    "year",
    "depth_cm",
    "width_px",
    "height_px",
}
SERIES_TEXT_FIELDS = set(SERIES_FIELDS) - {"year"}
DETAIL_TEXT_FIELDS = set(DETAIL_FIELDS) - {"width_px", "height_px"}
FILE_TEXT_FIELDS = set(FILE_FIELDS)
LINK_TEXT_FIELDS = set(LINK_FIELDS)


@dataclass(frozen=True)
class CatalogueSourceRecords:
    works: Dict[str, Dict[str, Any]]
    work_details: Dict[str, Dict[str, Any]]
    series: Dict[str, Dict[str, Any]]
    work_files: Dict[str, Dict[str, Any]]
    work_links: Dict[str, Dict[str, Any]]

    def as_maps(self) -> Dict[str, Dict[str, Dict[str, Any]]]:
        return {
            "works": self.works,
            "work_details": self.work_details,
            "series": self.series,
            "work_files": self.work_files,
            "work_links": self.work_links,
        }


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.startswith("'") and len(text) > 1:
        text = text[1:]
    return text


def normalize_status(value: Any) -> str:
    return normalize_text(value).lower()


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def slug_id(raw: Any, width: int = 5) -> str:
    if raw is None:
        raise ValueError("Missing id")
    text = normalize_text(raw)
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"\D", "", text)
    if not text:
        raise ValueError(f"Invalid id value: {raw!r}")
    return text.zfill(width)


def normalize_json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    return text if text else None


def normalize_embedded_entry(value: Any, fields: Iterable[str]) -> Dict[str, Any] | None:
    if not isinstance(value, Mapping):
        return None
    out: Dict[str, Any] = {}
    for field in fields:
        normalized = normalize_scalar_text(value.get(field))
        if normalized is not None:
            out[field] = normalized
    return out if out else None


def normalize_downloads(value: Any) -> list[Dict[str, Any]]:
    if not isinstance(value, list):
        return []
    out: list[Dict[str, Any]] = []
    for item in value:
        entry = normalize_embedded_entry(item, DOWNLOAD_FIELDS)
        if entry is not None:
            out.append(entry)
    return out


def normalize_links(value: Any) -> list[Dict[str, Any]]:
    if not isinstance(value, list):
        return []
    out: list[Dict[str, Any]] = []
    for item in value:
        entry = normalize_embedded_entry(item, WORK_LINK_ENTRY_FIELDS)
        if entry is not None:
            out.append(entry)
    return out


def normalize_scalar_text(value: Any) -> str | None:
    text = normalize_text(value)
    return text if text else None


def normalize_source_record(
    record: Mapping[str, Any],
    field_order: Iterable[str],
    *,
    text_fields: set[str] | None = None,
) -> Dict[str, Any]:
    text_fields = text_fields or set()
    out: Dict[str, Any] = {}
    for field in field_order:
        value = record.get(field)
        if field == "downloads":
            entries = normalize_downloads(value)
            if entries:
                out[field] = entries
        elif field == "links":
            entries = normalize_links(value)
            if entries:
                out[field] = entries
        elif isinstance(value, list):
            out[field] = [normalize_json_value(item) for item in value]
        elif field in text_fields:
            out[field] = normalize_scalar_text(value)
        else:
            out[field] = normalize_json_value(value)
    return out


def header_map(row: Iterable[Any]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for idx, raw in enumerate(row):
        key = normalize_text(raw)
        if not key:
            continue
        out[key] = idx
        out[key.lower()] = idx
    return out


def cell(row: tuple[Any, ...], headers: Mapping[str, int], name: str) -> Any:
    idx = headers.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def safe_uid_fragment(value: Any, fallback: str) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text or fallback


def unique_uid(base: str, used: set[str]) -> str:
    candidate = base
    suffix = 2
    while candidate in used:
        candidate = f"{base}-{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def require_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found in workbook: {sheet_name}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet is empty: {sheet_name}")
    return rows, header_map(rows[0])


def optional_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return [], {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}
    return rows, header_map(rows[0])


def build_work_record(row: tuple[Any, ...], headers: Mapping[str, int]) -> Dict[str, Any]:
    work_id = slug_id(cell(row, headers, "work_id"))
    record: Dict[str, Any] = {
        "work_id": work_id,
        "status": normalize_status(cell(row, headers, "status")) or None,
        "published_date": normalize_json_value(cell(row, headers, "published_date")),
        "series_ids": parse_series_ids(cell(row, headers, "series_ids")),
    }
    for field in WORK_FIELDS:
        if field in record:
            continue
        record[field] = normalize_json_value(cell(row, headers, field))
    return normalize_source_record(record, WORK_FIELDS, text_fields=WORK_TEXT_FIELDS)


def build_series_sort_map(rows: list[tuple[Any, ...]], headers: Mapping[str, int]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    if not rows:
        return out
    for row in rows[1:]:
        raw_series_id = cell(row, headers, "series_id")
        if is_empty(raw_series_id):
            continue
        series_id = normalize_series_id(raw_series_id)
        out[series_id] = normalize_json_value(cell(row, headers, "sort_fields"))
    return out


def build_series_record(
    row: tuple[Any, ...],
    headers: Mapping[str, int],
    sort_fields_by_series_id: Mapping[str, Any],
) -> Dict[str, Any]:
    series_id = normalize_series_id(cell(row, headers, "series_id"))
    raw_primary_work_id = cell(row, headers, "primary_work_id")
    record: Dict[str, Any] = {
        "series_id": series_id,
        "title": normalize_json_value(cell(row, headers, "title")),
        "series_type": normalize_json_value(cell(row, headers, "series_type")),
        "status": normalize_status(cell(row, headers, "status")) or None,
        "published_date": normalize_json_value(cell(row, headers, "published_date")),
        "year": normalize_json_value(cell(row, headers, "year")),
        "year_display": normalize_json_value(cell(row, headers, "year_display")),
        "primary_work_id": slug_id(raw_primary_work_id) if not is_empty(raw_primary_work_id) else None,
        "series_prose_file": normalize_json_value(cell(row, headers, "series_prose_file")),
        "notes": normalize_json_value(cell(row, headers, "notes")),
        "sort_fields": normalize_json_value(sort_fields_by_series_id.get(series_id)) or "work_id",
    }
    return normalize_source_record(record, SERIES_FIELDS, text_fields=SERIES_TEXT_FIELDS)


def build_detail_record(row: tuple[Any, ...], headers: Mapping[str, int]) -> Dict[str, Any]:
    work_id = slug_id(cell(row, headers, "work_id"))
    detail_id = slug_id(cell(row, headers, "detail_id"), width=3)
    record = {
        "detail_uid": f"{work_id}-{detail_id}",
        "work_id": work_id,
        "detail_id": detail_id,
        "project_subfolder": normalize_json_value(cell(row, headers, "project_subfolder")),
        "project_filename": normalize_json_value(cell(row, headers, "project_filename")),
        "title": normalize_json_value(cell(row, headers, "title")),
        "status": normalize_status(cell(row, headers, "status")) or None,
        "published_date": normalize_json_value(cell(row, headers, "published_date")),
        "width_px": normalize_json_value(cell(row, headers, "width_px")),
        "height_px": normalize_json_value(cell(row, headers, "height_px")),
    }
    return normalize_source_record(record, DETAIL_FIELDS, text_fields=DETAIL_TEXT_FIELDS)


def build_file_record(
    row: tuple[Any, ...],
    headers: Mapping[str, int],
    used_uids: set[str],
) -> Dict[str, Any]:
    work_id = slug_id(cell(row, headers, "work_id"))
    filename = normalize_scalar_text(cell(row, headers, "filename"))
    label = normalize_scalar_text(cell(row, headers, "label"))
    fragment = safe_uid_fragment(Path(filename or "").stem or label, "file")
    file_uid = unique_uid(f"{work_id}:{fragment}", used_uids)
    record = {
        "file_uid": file_uid,
        "work_id": work_id,
        "filename": filename,
        "label": label,
        "status": normalize_status(cell(row, headers, "status")) or None,
        "published_date": normalize_json_value(cell(row, headers, "published_date")),
    }
    return normalize_source_record(record, FILE_FIELDS, text_fields=FILE_TEXT_FIELDS)


def build_link_record(
    row: tuple[Any, ...],
    headers: Mapping[str, int],
    used_uids: set[str],
) -> Dict[str, Any]:
    work_id = slug_id(cell(row, headers, "work_id"))
    url = normalize_scalar_text(cell(row, headers, "url")) or normalize_scalar_text(cell(row, headers, "URL"))
    label = normalize_scalar_text(cell(row, headers, "label"))
    fragment = safe_uid_fragment(label or url, "link")
    link_uid = unique_uid(f"{work_id}:{fragment}", used_uids)
    record = {
        "link_uid": link_uid,
        "work_id": work_id,
        "url": url,
        "label": label,
        "status": normalize_status(cell(row, headers, "status")) or None,
        "published_date": normalize_json_value(cell(row, headers, "published_date")),
    }
    return normalize_source_record(record, LINK_FIELDS, text_fields=LINK_TEXT_FIELDS)


def build_work_file_records_from_downloads(
    works: Mapping[str, Mapping[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    used_uids: set[str] = set()
    out: Dict[str, Dict[str, Any]] = {}
    for work_id in sorted(works):
        work_record = works[work_id]
        for download in normalize_downloads(work_record.get("downloads")):
            filename = normalize_scalar_text(download.get("filename"))
            label = normalize_scalar_text(download.get("label"))
            if not filename and not label:
                continue
            fragment = safe_uid_fragment(Path(filename or "").stem or label, "file")
            file_uid = unique_uid(f"{work_id}:{fragment}", used_uids)
            out[file_uid] = normalize_source_record(
                {
                    "file_uid": file_uid,
                    "work_id": work_id,
                    "filename": filename,
                    "label": label,
                    "status": "published",
                    "published_date": None,
                },
                FILE_FIELDS,
                text_fields=FILE_TEXT_FIELDS,
            )
    return sort_record_map(out)


def build_work_link_records_from_links(
    works: Mapping[str, Mapping[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    used_uids: set[str] = set()
    out: Dict[str, Dict[str, Any]] = {}
    for work_id in sorted(works):
        work_record = works[work_id]
        for link in normalize_links(work_record.get("links")):
            url = normalize_scalar_text(link.get("url"))
            label = normalize_scalar_text(link.get("label"))
            if not url and not label:
                continue
            fragment = safe_uid_fragment(label or url, "link")
            link_uid = unique_uid(f"{work_id}:{fragment}", used_uids)
            out[link_uid] = normalize_source_record(
                {
                    "link_uid": link_uid,
                    "work_id": work_id,
                    "url": url,
                    "label": label,
                    "status": "published",
                    "published_date": None,
                },
                LINK_FIELDS,
                text_fields=LINK_TEXT_FIELDS,
            )
    return sort_record_map(out)


def attach_work_owned_files_and_links(
    works: Mapping[str, Dict[str, Any]],
    work_files: Mapping[str, Mapping[str, Any]],
    work_links: Mapping[str, Mapping[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    out = {work_id: dict(record) for work_id, record in works.items()}
    files_by_work_id: Dict[str, list[Dict[str, Any]]] = {}
    for file_uid, file_record in sorted(work_files.items()):
        work_id = normalize_text(file_record.get("work_id"))
        if not work_id or work_id not in out:
            continue
        entry = normalize_embedded_entry(file_record, DOWNLOAD_FIELDS)
        if entry is None:
            continue
        files_by_work_id.setdefault(work_id, []).append(entry)
    links_by_work_id: Dict[str, list[Dict[str, Any]]] = {}
    for link_uid, link_record in sorted(work_links.items()):
        work_id = normalize_text(link_record.get("work_id"))
        if not work_id or work_id not in out:
            continue
        entry = normalize_embedded_entry(link_record, WORK_LINK_ENTRY_FIELDS)
        if entry is None:
            continue
        links_by_work_id.setdefault(work_id, []).append(entry)
    for work_id, entries in files_by_work_id.items():
        if not normalize_downloads(out[work_id].get("downloads")):
            out[work_id]["downloads"] = entries
    for work_id, entries in links_by_work_id.items():
        if not normalize_links(out[work_id].get("links")):
            out[work_id]["links"] = entries
    return sort_record_map({
        work_id: normalize_source_record(record, WORK_FIELDS, text_fields=WORK_TEXT_FIELDS)
        for work_id, record in out.items()
    })


def records_from_workbook(
    workbook_path: Path,
    *,
    works_sheet: str = "Works",
    series_sheet: str = "Series",
    series_sort_sheet: str = "SeriesSort",
    work_details_sheet: str = "WorkDetails",
    work_files_sheet: str = "WorkFiles",
    work_links_sheet: str = "WorkLinks",
) -> CatalogueSourceRecords:
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    works_rows, works_headers = require_sheet(wb, works_sheet)
    series_rows, series_headers = require_sheet(wb, series_sheet)
    details_rows, details_headers = optional_sheet(wb, work_details_sheet)
    files_rows, files_headers = optional_sheet(wb, work_files_sheet)
    links_rows, links_headers = optional_sheet(wb, work_links_sheet)
    sort_rows, sort_headers = optional_sheet(wb, series_sort_sheet)
    sort_fields_by_series_id = build_series_sort_map(sort_rows, sort_headers)

    works: Dict[str, Dict[str, Any]] = {}
    for row in works_rows[1:]:
        if is_empty(cell(row, works_headers, "work_id")):
            continue
        record = build_work_record(row, works_headers)
        works[str(record["work_id"])] = record

    series: Dict[str, Dict[str, Any]] = {}
    for row in series_rows[1:]:
        if is_empty(cell(row, series_headers, "series_id")):
            continue
        record = build_series_record(row, series_headers, sort_fields_by_series_id)
        series[str(record["series_id"])] = record

    work_details: Dict[str, Dict[str, Any]] = {}
    for row in details_rows[1:] if details_rows else []:
        if is_empty(cell(row, details_headers, "work_id")) or is_empty(cell(row, details_headers, "detail_id")):
            continue
        record = build_detail_record(row, details_headers)
        work_details[str(record["detail_uid"])] = record

    used_file_uids: set[str] = set()
    work_files: Dict[str, Dict[str, Any]] = {}
    for row in files_rows[1:] if files_rows else []:
        if is_empty(cell(row, files_headers, "work_id")):
            continue
        record = build_file_record(row, files_headers, used_file_uids)
        work_files[str(record["file_uid"])] = record

    used_link_uids: set[str] = set()
    work_links: Dict[str, Dict[str, Any]] = {}
    for row in links_rows[1:] if links_rows else []:
        if is_empty(cell(row, links_headers, "work_id")):
            continue
        record = build_link_record(row, links_headers, used_link_uids)
        work_links[str(record["link_uid"])] = record

    works = attach_work_owned_files_and_links(works, work_files, work_links)
    work_files = build_work_file_records_from_downloads(works)
    work_links = build_work_link_records_from_links(works)

    return CatalogueSourceRecords(
        works=works,
        work_details=sort_record_map(work_details),
        series=sort_record_map(series),
        work_files=work_files,
        work_links=work_links,
    )


def sort_record_map(records: Mapping[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    return {key: records[key] for key in sorted(records)}


def payload_for_map(kind: str, records: Mapping[str, Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "header": {
            "schema": SCHEMAS[kind],
            "count": len(records),
        },
        kind: sort_record_map(records),
    }


def payloads_from_records(records: CatalogueSourceRecords, *, workbook_path: Path | None = None) -> Dict[str, Dict[str, Any]]:
    payloads = {
        kind: payload_for_map(kind, record_map)
        for kind, record_map in records.as_maps().items()
    }
    payloads["meta"] = {
        "header": {
            "schema": SCHEMAS["meta"],
        },
        "source": {
            "canonical": "json",
            "created_from": workbook_path.as_posix() if workbook_path is not None else None,
        },
        "id_policy": {
            "work_id_width": 5,
            "detail_id_width": 3,
        },
    }
    return payloads


def write_payloads(source_dir: Path, payloads: Mapping[str, Mapping[str, Any]]) -> list[Path]:
    source_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for kind in ["works", "work_details", "series", "work_files", "work_links", "meta"]:
        path = source_dir / SOURCE_FILES[kind]
        path.write_text(json.dumps(payloads[kind], ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        written.append(path)
    return written


def write_source_record_payloads(
    source_dir: Path,
    records: CatalogueSourceRecords,
    *,
    kinds: Iterable[str] = ("works", "work_details", "series"),
) -> list[Path]:
    source_dir.mkdir(parents=True, exist_ok=True)
    record_maps = records.as_maps()
    written: list[Path] = []
    for kind in kinds:
        if kind not in record_maps:
            raise ValueError(f"Unknown source record kind: {kind}")
        path = source_dir / SOURCE_FILES[kind]
        payload = payload_for_map(kind, record_maps[kind])
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        written.append(path)
    return written


def load_json_file(path: Path) -> Dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ValueError(f"Missing source file: {path}") from exc
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON in {path}: {exc}") from exc
    if not isinstance(payload, dict):
        raise ValueError(f"Invalid source file shape in {path}: expected object")
    return payload


def records_from_json_source(source_dir: Path) -> CatalogueSourceRecords:
    maps: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for kind in ["works", "work_details", "series"]:
        path = source_dir / SOURCE_FILES[kind]
        payload = load_json_file(path)
        record_map = payload.get(kind)
        if not isinstance(record_map, dict):
            raise ValueError(f"Invalid source file shape in {path}: missing object key {kind!r}")
        maps[kind] = {
            str(record_id): dict(record)
            for record_id, record in record_map.items()
            if isinstance(record, dict)
        }
    works = sort_record_map({
        work_id: normalize_source_record(record, WORK_FIELDS, text_fields=WORK_TEXT_FIELDS)
        for work_id, record in maps["works"].items()
    })
    return CatalogueSourceRecords(
        works=works,
        work_details=sort_record_map(maps["work_details"]),
        series=sort_record_map(maps["series"]),
        work_files=build_work_file_records_from_downloads(works),
        work_links=build_work_link_records_from_links(works),
    )


def workbook_cell_value(value: Any) -> Any:
    if isinstance(value, list):
        return ",".join(str(item) for item in value if item is not None)
    return value


def append_sheet_rows(ws, headers: list[str], records: Iterable[Mapping[str, Any]]) -> None:
    ws.append(headers)
    for record in records:
        ws.append([workbook_cell_value(record.get(header)) for header in headers])


def write_records_to_workbook(records: CatalogueSourceRecords, workbook_path: Path) -> None:
    """Materialize JSON source records as a workbook compatible with generate_work_pages.py."""
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    append_sheet_rows(
        wb.create_sheet("Works"),
        WORKBOOK_HEADERS["Works"],
        records.works.values(),
    )
    append_sheet_rows(
        wb.create_sheet("Series"),
        WORKBOOK_HEADERS["Series"],
        records.series.values(),
    )
    append_sheet_rows(
        wb.create_sheet("SeriesSort"),
        WORKBOOK_HEADERS["SeriesSort"],
        (
            {
                "series_id": record.get("series_id"),
                "sort_fields": record.get("sort_fields") or "work_id",
            }
            for record in records.series.values()
        ),
    )
    append_sheet_rows(
        wb.create_sheet("WorkDetails"),
        WORKBOOK_HEADERS["WorkDetails"],
        records.work_details.values(),
    )
    append_sheet_rows(
        wb.create_sheet("WorkFiles"),
        WORKBOOK_HEADERS["WorkFiles"],
        records.work_files.values(),
    )
    append_sheet_rows(
        wb.create_sheet("WorkLinks"),
        WORKBOOK_HEADERS["WorkLinks"],
        records.work_links.values(),
    )

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)


def sync_mutable_fields_from_workbook(source_dir: Path, workbook_path: Path) -> list[Path]:
    """Copy generator-updated mutable fields from a materialized workbook back to JSON source."""
    source_records = records_from_json_source(source_dir)
    workbook_records = records_from_workbook(workbook_path)

    for work_id, workbook_record in workbook_records.works.items():
        source_record = source_records.works.get(work_id)
        if source_record is None:
            continue
        for field in ["status", "published_date", "width_px", "height_px"]:
            source_record[field] = workbook_record.get(field)

    for detail_uid, workbook_record in workbook_records.work_details.items():
        source_record = source_records.work_details.get(detail_uid)
        if source_record is None:
            continue
        for field in ["status", "published_date", "width_px", "height_px"]:
            source_record[field] = workbook_record.get(field)

    for series_id, workbook_record in workbook_records.series.items():
        source_record = source_records.series.get(series_id)
        if source_record is None:
            continue
        for field in ["status", "published_date"]:
            source_record[field] = workbook_record.get(field)

    for file_uid, workbook_record in workbook_records.work_files.items():
        source_record = source_records.work_files.get(file_uid)
        if source_record is None:
            continue
        for field in ["status", "published_date"]:
            source_record[field] = workbook_record.get(field)

    for link_uid, workbook_record in workbook_records.work_links.items():
        source_record = source_records.work_links.get(link_uid)
        if source_record is None:
            continue
        for field in ["status", "published_date"]:
            source_record[field] = workbook_record.get(field)

    return write_source_record_payloads(source_dir, source_records)


def validate_source_records(records: CatalogueSourceRecords) -> list[str]:
    errors: list[str] = []
    all_work_ids: set[str] = set()
    all_series_ids: set[str] = set()
    work_series_ids_by_work_id: Dict[str, list[str]] = {}

    for key, record in records.works.items():
        try:
            work_id = slug_id(record.get("work_id") or key)
        except ValueError as exc:
            errors.append(f"works {key}: invalid work_id ({exc})")
            continue
        if key != work_id:
            errors.append(f"works {key}: key does not match normalized work_id {work_id}")
        all_work_ids.add(work_id)
        series_ids = record.get("series_ids")
        if not isinstance(series_ids, list):
            errors.append(f"works {key}: series_ids must be an array")
            series_ids = []
        parsed_series_ids: list[str] = []
        for raw_series_id in series_ids:
            try:
                parsed_series_ids.append(normalize_series_id(raw_series_id))
            except ValueError as exc:
                errors.append(f"works {key}: {exc}")
        work_series_ids_by_work_id[work_id] = parsed_series_ids
        downloads = record.get("downloads")
        if downloads is not None:
            if not isinstance(downloads, list):
                errors.append(f"works {key}: downloads must be an array")
            else:
                for idx, download in enumerate(downloads, start=1):
                    if not isinstance(download, Mapping):
                        errors.append(f"works {key}: downloads item {idx} must be an object")
                        continue
                    if is_empty(download.get("filename")):
                        errors.append(f"works {key}: downloads item {idx} missing filename")
                    if is_empty(download.get("label")):
                        errors.append(f"works {key}: downloads item {idx} missing label")
        links = record.get("links")
        if links is not None:
            if not isinstance(links, list):
                errors.append(f"works {key}: links must be an array")
            else:
                for idx, link in enumerate(links, start=1):
                    if not isinstance(link, Mapping):
                        errors.append(f"works {key}: links item {idx} must be an object")
                        continue
                    if is_empty(link.get("url")):
                        errors.append(f"works {key}: links item {idx} missing url")
                    if is_empty(link.get("label")):
                        errors.append(f"works {key}: links item {idx} missing label")

    for key, record in records.series.items():
        try:
            series_id = normalize_series_id(record.get("series_id") or key)
        except ValueError as exc:
            errors.append(f"series {key}: {exc}")
            continue
        if key != series_id:
            errors.append(f"series {key}: key does not match normalized series_id {series_id}")
        all_series_ids.add(series_id)

        if normalize_status(record.get("status")) not in ACTIONABLE_STATUSES:
            continue
        status = normalize_status(record.get("status"))
        raw_primary_work_id = record.get("primary_work_id")
        if is_empty(raw_primary_work_id):
            if status == "published":
                errors.append(f"series {series_id}: missing primary_work_id")
            continue
        try:
            primary_work_id = slug_id(raw_primary_work_id)
        except ValueError as exc:
            errors.append(f"series {series_id}: invalid primary_work_id {raw_primary_work_id!r} ({exc})")
            continue
        if primary_work_id not in all_work_ids:
            errors.append(f"series {series_id}: primary_work_id {primary_work_id!r} not found in works")
            continue
        if series_id not in work_series_ids_by_work_id.get(primary_work_id, []):
            errors.append(
                f"series {series_id}: primary_work_id {primary_work_id!r} is not in that work's series_ids"
            )

    for work_id, series_ids in work_series_ids_by_work_id.items():
        if normalize_status(records.works.get(work_id, {}).get("status")) not in ACTIONABLE_STATUSES:
            continue
        for series_id in series_ids:
            if series_id not in all_series_ids:
                errors.append(f"works {work_id}: references unknown series_id {series_id!r}")

    for key, record in records.work_details.items():
        raw_work_id = record.get("work_id")
        raw_detail_id = record.get("detail_id")
        if is_empty(raw_work_id) or is_empty(raw_detail_id):
            errors.append(f"work_details {key}: missing work_id or detail_id")
            continue
        try:
            work_id = slug_id(raw_work_id)
            detail_id = slug_id(raw_detail_id, width=3)
        except ValueError as exc:
            errors.append(f"work_details {key}: invalid id ({exc})")
            continue
        detail_uid = f"{work_id}-{detail_id}"
        if key != detail_uid or record.get("detail_uid") != detail_uid:
            errors.append(f"work_details {key}: key/detail_uid does not match normalized detail_uid {detail_uid}")
        if normalize_status(record.get("status")) in ACTIONABLE_STATUSES and work_id not in all_work_ids:
            errors.append(f"work_details {key}: work_id {work_id!r} not found in works")

    for key, record in records.work_files.items():
        work_id = record.get("work_id")
        if is_empty(work_id):
            errors.append(f"work_files {key}: missing work_id")
            continue
        try:
            normalized_work_id = slug_id(work_id)
        except ValueError as exc:
            errors.append(f"work_files {key}: invalid work_id ({exc})")
            continue
        if normalize_status(record.get("status")) in ACTIONABLE_STATUSES and normalized_work_id not in all_work_ids:
            errors.append(f"work_files {key}: work_id {normalized_work_id!r} not found in works")

    for key, record in records.work_links.items():
        work_id = record.get("work_id")
        if is_empty(work_id):
            errors.append(f"work_links {key}: missing work_id")
            continue
        try:
            normalized_work_id = slug_id(work_id)
        except ValueError as exc:
            errors.append(f"work_links {key}: invalid work_id ({exc})")
            continue
        if normalize_status(record.get("status")) in ACTIONABLE_STATUSES and normalized_work_id not in all_work_ids:
            errors.append(f"work_links {key}: work_id {normalized_work_id!r} not found in works")

    return sorted(dict.fromkeys(errors))


def compare_record_maps(
    left: Mapping[str, Dict[str, Any]],
    right: Mapping[str, Dict[str, Any]],
    *,
    label: str,
) -> list[str]:
    diffs: list[str] = []
    left_keys = set(left.keys())
    right_keys = set(right.keys())
    for key in sorted(left_keys - right_keys):
        diffs.append(f"{label} {key}: missing from JSON source")
    for key in sorted(right_keys - left_keys):
        diffs.append(f"{label} {key}: extra in JSON source")
    for key in sorted(left_keys & right_keys):
        if left[key] != right[key]:
            diffs.append(f"{label} {key}: record differs")
    return diffs


def compare_sources(workbook_records: CatalogueSourceRecords, json_records: CatalogueSourceRecords) -> list[str]:
    diffs: list[str] = []
    for kind in ["works", "work_details", "series", "work_files", "work_links"]:
        diffs.extend(
            compare_record_maps(
                workbook_records.as_maps()[kind],
                json_records.as_maps()[kind],
                label=kind,
            )
        )
    return diffs


def add_common_source_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--source-dir", default=str(DEFAULT_SOURCE_DIR), help="Catalogue source JSON directory")
