#!/usr/bin/env python3
"""
Build the catalogue pipeline end-to-end (fail-fast):

1) copy_draft_media_files.py --mode work --write
2) make_srcset_images.sh
3) copy_draft_media_files.py --mode work_details --write
4) make_srcset_images.sh (for work details)
5) copy_draft_media_files.py --mode moment --write
6) make_srcset_images.sh (for moments)
7) generate_work_pages.py data/works.xlsx --write (for affected work IDs only)
8) build_search.rb --scope catalogue --write

Dry-run mode:
- copy + srcset + generate + catalogue-search build run in preview mode
- no workbook writes/deletes are performed

Flag usage summary:
- --xlsx: workbook path (repo-relative by default)
- --dry-run: preview-only mode for all steps, including catalogue search rebuild
- --force-generate: pass --force to generate_work_pages.py and the catalogue search rebuild
- --jobs: parallelism for make_srcset_images.sh
- --*-input-dir / --*-output-dir: source/derivative directories for each media type
- --mode: select one or more flows: work, work_details, moment (repeatable)
- --work-ids / --work-ids-file: limit work + work_details processing scope
- --series-ids / --series-ids-file: pass series filter into generation
- --moment-ids / --moment-ids-file: limit moment processing scope

Required local environment (when path flags are not passed):
- env vars are resolved from `_data/pipeline.json`
- defaults remain `DOTLINEFORM_MEDIA_BASE_DIR` and `DOTLINEFORM_PROJECTS_BASE_DIR`

Behavior note:
- When mode includes `work` and no explicit `--series-ids*` are provided,
  draft series IDs are auto-included for generation.

Manifest files (created in a temporary directory):
- draft_*: candidates from workbook (status=draft)
- copied_*: successfully copied source IDs (input to srcset step)
- *_success_ids: srcset-success IDs returned by make_srcset_images.sh
- generate_ids.txt: work IDs to regenerate (works + any work_details parents)
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Set

import openpyxl

try:
    from build_activity import append_build_activity
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.build_activity import append_build_activity

try:
    from display_paths import format_display_command, format_display_path
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.display_paths import format_display_command, format_display_path

try:
    from script_logging import append_script_log
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.script_logging import append_script_log

try:
    from pipeline_config import (
        env_var_name,
        env_var_value,
        join_base_and_subdir,
        load_pipeline_config,
        media_mode_input_subdir,
        media_mode_output_subdir,
        media_work_files_subdir,
        source_moments_images_subdir,
        source_moments_root_subdir,
        source_works_prose_subdir,
        source_works_root_subdir,
    )
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.pipeline_config import (
        env_var_name,
        env_var_value,
        join_base_and_subdir,
        load_pipeline_config,
        media_mode_input_subdir,
        media_mode_output_subdir,
        media_work_files_subdir,
        source_moments_images_subdir,
        source_moments_root_subdir,
        source_works_prose_subdir,
        source_works_root_subdir,
    )


try:
    from catalogue_preflight import raise_if_invalid_catalogue_workbook
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.catalogue_preflight import raise_if_invalid_catalogue_workbook


PIPELINE_CONFIG = load_pipeline_config(Path(__file__))
MEDIA_BASE_DIR_ENV_NAME = env_var_name(PIPELINE_CONFIG, "media_base_dir")
SRCSET_JOBS_ENV_NAME = env_var_name(PIPELINE_CONFIG, "srcset_jobs")
SRCSET_SELECTED_IDS_ENV_NAME = env_var_name(PIPELINE_CONFIG, "srcset_selected_ids_file")
SRCSET_SUCCESS_IDS_ENV_NAME = env_var_name(PIPELINE_CONFIG, "srcset_success_ids_file")
BUILD_STATE_SCHEMA = "build_catalogue_state"
LEGACY_BUILD_STATE_SCHEMAS = {"build_catalogue_state_v1"}
BUILD_STATE_PLANNER_VERSION = 4
BUILD_STATE_MIGRATION_NOTE = (
    "planner_version 4 adds moment prose tracking. Legacy state files are "
    "accepted and normalized in memory, then rewritten on the next successful write run."
)
DEFAULT_BUILD_STATE_PATH = Path("var/build_catalogue_state.json")
REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PROJECTS_BASE_DIR = Path(projects_base).expanduser() if (projects_base := env_var_value(PIPELINE_CONFIG, "projects_base_dir")) else None
DEFAULT_MEDIA_BASE_DIR = Path(media_base).expanduser() if (media_base := env_var_value(PIPELINE_CONFIG, "media_base_dir")) else None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("'") and len(s) > 1:
        s = s[1:]
    return s


def normalize_status(value: Any) -> str:
    return normalize_text(value).lower()


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def slug_id(raw: Any, width: int = 5) -> str:
    if raw is None:
        raise ValueError("Missing id")
    s = normalize_text(raw)
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"\D", "", s)
    if not s:
        raise ValueError(f"Invalid id value: {raw!r}")
    return s.zfill(width)


def header_map(ws) -> Dict[str, int]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    hi: Dict[str, int] = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).strip()
        hi[key] = i
        hi[key.lower()] = i
    return hi


def first_present_col(header_index: Dict[str, int], names: list[str]) -> str | None:
    for name in names:
        if name in header_index:
            return name
    return None


def parse_series_ids_from_works_row(row: tuple[Any, ...], hi: Dict[str, int]) -> list[str]:
    parsed: list[str] = []
    series_ids_col = first_present_col(hi, ["series_ids"])
    if series_ids_col is not None:
        raw_value = row[hi[series_ids_col]]
        parsed = [normalize_text(part) for part in normalize_text(raw_value).split(",") if normalize_text(part)]

    out: list[str] = []
    seen: set[str] = set()
    for sid in parsed:
        if not sid or sid in seen:
            continue
        seen.add(sid)
        out.append(sid)
    return out


def run_step(label: str, cmd: list[str], cwd: Path, env: Dict[str, str] | None = None) -> None:
    """Run one pipeline step and fail fast on non-zero exit."""
    print(f"\n==> {label}")
    print(
        "+",
        format_display_command(
            cmd,
            repo_root=REPO_ROOT,
            projects_base_dir=DEFAULT_PROJECTS_BASE_DIR,
            media_base_dir=DEFAULT_MEDIA_BASE_DIR,
        ),
    )
    subprocess.run(cmd, cwd=str(cwd), env=env, check=True)


def log_event(repo_root: Path, event: str, details: Dict[str, Any] | None = None) -> None:
    try:
        append_script_log(Path(__file__), event=event, details=details or {}, repo_root=repo_root)
    except Exception:
        # Logging failures must not stop pipeline execution.
        pass


def read_ids(path: Path) -> Set[str]:
    """Read newline-delimited IDs from a manifest file."""
    if not path.exists():
        return set()
    out: Set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s:
            out.add(s)
    return out


def read_optional_ids_file(path: str) -> Set[str]:
    if not path:
        return set()
    ids_path = Path(path).expanduser()
    if not ids_path.exists():
        raise SystemExit(f"IDs file not found: {ids_path}")
    return read_ids(ids_path)


def parse_work_id_selection(raw: str) -> Set[str]:
    """
    Parse comma-separated work-id selectors supporting individual IDs and ranges.
    Examples:
      "66,74" -> {"00066", "00074"}
      "66-74,38-40,12" -> {"00012", "00038", ..., "00074"}
    """
    selected: Set[str] = set()
    for token in (part.strip() for part in str(raw).split(",") if part.strip()):
        m = re.match(r"^(\d+)\s*-\s*(\d+)$", token)
        if m:
            start = int(m.group(1))
            end = int(m.group(2))
            if start > end:
                start, end = end, start
            for n in range(start, end + 1):
                selected.add(slug_id(n))
        else:
            selected.add(slug_id(token))
    return selected


def collect_draft_ids(
    xlsx_path: Path,
    allowed_ids: Set[str] | None = None,
    include_published: bool = False,
) -> Set[str]:
    """Collect draft work_ids from Works.status."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Works" not in wb.sheetnames:
        raise SystemExit("Worksheet not found: 'Works'")
    ws = wb["Works"]
    hi = header_map(ws)

    required = ["work_id", "status"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"Works sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_work_id = row[hi["work_id"]]
        if is_empty(raw_work_id):
            continue
        wid = slug_id(raw_work_id)
        if allowed_ids is not None and wid not in allowed_ids:
            continue
        status = normalize_status(row[hi["status"]])
        if status != "draft" and not (include_published and status == "published"):
            continue
        out.add(wid)
    return out


def collect_draft_detail_uids(
    xlsx_path: Path,
    allowed_uids: Set[str] | None = None,
    allowed_work_ids: Set[str] | None = None,
    include_published: bool = False,
) -> Set[str]:
    """Collect draft detail_uids from WorkDetails.status."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "WorkDetails" not in wb.sheetnames:
        return set()
    ws = wb["WorkDetails"]
    hi = header_map(ws)

    required = ["work_id", "detail_id", "status"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"WorkDetails sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_work_id = row[hi["work_id"]]
        raw_detail_id = row[hi["detail_id"]]
        if is_empty(raw_work_id) or is_empty(raw_detail_id):
            continue
        wid = slug_id(raw_work_id)
        did = slug_id(raw_detail_id, width=3)
        uid = f"{wid}-{did}"
        if allowed_uids is not None and uid not in allowed_uids:
            continue
        if allowed_work_ids is not None and wid not in allowed_work_ids:
            continue
        status = normalize_status(row[hi["status"]])
        if status != "draft" and not (include_published and status == "published"):
            continue
        out.add(uid)
    return out


def collect_draft_moment_ids(
    xlsx_path: Path,
    allowed_ids: Set[str] | None = None,
    include_published: bool = False,
) -> Set[str]:
    """Collect draft moment_ids from Moments.status."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Moments" not in wb.sheetnames:
        return set()
    ws = wb["Moments"]
    hi = header_map(ws)

    required = ["moment_id", "status"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"Moments sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_moment_id = row[hi["moment_id"]]
        if is_empty(raw_moment_id):
            continue
        mid = normalize_text(raw_moment_id).lower()
        if not mid:
            continue
        if allowed_ids is not None and mid not in allowed_ids:
            continue
        status = normalize_status(row[hi["status"]])
        if status != "draft" and not (include_published and status == "published"):
            continue
        out.add(mid)
    return out


def collect_draft_series_ids(xlsx_path: Path, allowed_ids: Set[str] | None = None) -> Set[str]:
    """Collect draft series_ids from Series.status."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Series" not in wb.sheetnames:
        return set()
    ws = wb["Series"]
    hi = header_map(ws)

    required = ["series_id", "status"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"Series sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_series_id = row[hi["series_id"]]
        if is_empty(raw_series_id):
            continue
        sid = normalize_text(raw_series_id)
        if not sid:
            continue
        if allowed_ids is not None and sid not in allowed_ids:
            continue
        status = normalize_status(row[hi["status"]])
        if status != "draft":
            continue
        out.add(sid)
    return out


def collect_series_ids_for_work_ids(xlsx_path: Path, work_ids: Set[str]) -> Set[str]:
    """Map work_ids to series_ids using Works sheet rows."""
    if not work_ids:
        return set()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Works" not in wb.sheetnames:
        raise SystemExit("Worksheet not found: 'Works'")
    ws = wb["Works"]
    hi = header_map(ws)

    required = ["work_id", "series_ids"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"Works sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_work_id = row[hi["work_id"]]
        if is_empty(raw_work_id):
            continue
        wid = slug_id(raw_work_id)
        if wid not in work_ids:
            continue
        for sid in parse_series_ids_from_works_row(row, hi):
            out.add(sid)
    return out


def work_ids_from_detail_uids(detail_uids: Iterable[str]) -> Set[str]:
    """Map detail_uids (work_id-detail_id) to parent work_ids."""
    out: Set[str] = set()
    for uid in detail_uids:
        s = str(uid).strip()
        if not s or "-" not in s:
            continue
        out.add(s.split("-", 1)[0])
    return out


def utc_timestamp_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def normalize_row_payload(headers: list[str], row: tuple[Any, ...]) -> Dict[str, Any]:
    payload: Dict[str, Any] = {}
    for idx, header in enumerate(headers):
        key = normalize_text(header)
        if not key:
            continue
        payload[key] = normalize_cell_value(row[idx] if idx < len(row) else None)
    return payload


def stable_payload_hash(payload: Any) -> str:
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    digest = hashlib.blake2b(encoded.encode("utf-8"), digest_size=16).hexdigest()
    return f"blake2b-{digest}"


def stable_file_hash(path: Path) -> str:
    digest = hashlib.blake2b(digest_size=16)
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return f"blake2b-{digest.hexdigest()}"


def relative_to_base(path: Path, base: Path) -> str:
    try:
        return path.resolve().relative_to(base.resolve()).as_posix()
    except ValueError:
        return path.resolve().as_posix()


def fingerprint_media_path(
    resolved_path: Path | None,
    *,
    base_dir: Path,
    previous_entry: Dict[str, Any] | None = None,
    missing_reason: str | None = None,
) -> Dict[str, Any]:
    if resolved_path is None:
        return {
            "path": None,
            "exists": False,
            "missing_reason": missing_reason or "unmapped",
        }

    path_info = relative_to_base(resolved_path, base_dir)
    if not resolved_path.exists() or not resolved_path.is_file():
        return {
            "path": path_info,
            "exists": False,
            "missing_reason": missing_reason or "missing",
        }

    stat = resolved_path.stat()
    size = int(stat.st_size)
    mtime_ns = int(stat.st_mtime_ns)
    previous_path = normalize_text(previous_entry.get("path")) if isinstance(previous_entry, dict) else ""
    previous_exists = bool(previous_entry.get("exists")) if isinstance(previous_entry, dict) else False
    previous_size = previous_entry.get("size") if isinstance(previous_entry, dict) else None
    previous_mtime_ns = previous_entry.get("mtime_ns") if isinstance(previous_entry, dict) else None
    previous_hash = normalize_text(previous_entry.get("hash")) if isinstance(previous_entry, dict) else ""

    file_hash = ""
    if previous_exists and previous_path == path_info and previous_size == size and previous_mtime_ns == mtime_ns and previous_hash:
        file_hash = previous_hash
    else:
        file_hash = stable_file_hash(resolved_path)

    return {
        "path": path_info,
        "exists": True,
        "size": size,
        "mtime_ns": mtime_ns,
        "hash": file_hash,
    }


def sorted_unique_series_ids(raw_value: Any) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for part in normalize_text(raw_value).split(","):
        sid = normalize_text(part)
        if not sid or sid in seen:
            continue
        seen.add(sid)
        out.append(sid)
    return out


def workbook_headers(ws) -> list[str]:
    return [normalize_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def build_workbook_state(
    xlsx_path: Path,
    *,
    projects_base_dir: Path,
    previous_media: Dict[str, Any] | None = None,
    previous_prose: Dict[str, Any] | None = None,
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    works: Dict[str, Dict[str, Any]] = {}
    series: Dict[str, Dict[str, Any]] = {}
    work_details: Dict[str, Dict[str, Any]] = {}
    work_files_rows: Dict[str, list[Dict[str, Any]]] = {}
    work_links_rows: Dict[str, list[Dict[str, Any]]] = {}
    moments: Dict[str, Dict[str, Any]] = {}
    media_work: Dict[str, Dict[str, Any]] = {}
    media_work_details: Dict[str, Dict[str, Any]] = {}
    media_moments: Dict[str, Dict[str, Any]] = {}
    prose_work: Dict[str, Dict[str, Any]] = {}
    prose_series: Dict[str, Dict[str, Any]] = {}
    prose_moments: Dict[str, Dict[str, Any]] = {}
    previous_media = previous_media or {}
    previous_prose = previous_prose or {}
    previous_media_work = previous_media.get("work", {}) if isinstance(previous_media.get("work"), dict) else {}
    previous_media_work_details = previous_media.get("work_details", {}) if isinstance(previous_media.get("work_details"), dict) else {}
    previous_media_moments = previous_media.get("moment", {}) if isinstance(previous_media.get("moment"), dict) else {}
    previous_prose_work = previous_prose.get("work", {}) if isinstance(previous_prose.get("work"), dict) else {}
    previous_prose_series = previous_prose.get("series", {}) if isinstance(previous_prose.get("series"), dict) else {}
    previous_prose_moments = previous_prose.get("moment", {}) if isinstance(previous_prose.get("moment"), dict) else {}
    works_root = projects_base_dir / source_works_root_subdir(PIPELINE_CONFIG)
    moments_root = projects_base_dir / source_moments_root_subdir(PIPELINE_CONFIG)
    moments_images_root = projects_base_dir / source_moments_images_subdir(PIPELINE_CONFIG)
    works_prose_root = source_works_prose_subdir(PIPELINE_CONFIG)
    work_project_folder_by_id: Dict[str, str] = {}

    if "Works" in wb.sheetnames:
        ws = wb["Works"]
        headers = workbook_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            payload = normalize_row_payload(headers, row)
            raw_work_id = payload.get("work_id")
            if is_empty(raw_work_id):
                continue
            work_id = slug_id(raw_work_id)
            works[work_id] = {
                "hash": stable_payload_hash(payload),
                "status": normalize_status(payload.get("status")),
                "series_ids": sorted_unique_series_ids(payload.get("series_ids")),
            }
            project_folder = normalize_text(payload.get("project_folder"))
            project_filename = normalize_text(payload.get("project_filename"))
            if project_folder:
                work_project_folder_by_id[work_id] = project_folder
            resolved_path: Path | None = None
            missing_reason: str | None = None
            if project_folder and project_filename:
                resolved_path = works_root / project_folder / project_filename
            elif project_filename:
                missing_reason = "missing_project_folder"
            else:
                missing_reason = "missing_project_filename"
            media_work[work_id] = fingerprint_media_path(
                resolved_path,
                base_dir=projects_base_dir,
                previous_entry=previous_media_work.get(work_id),
                missing_reason=missing_reason,
            )
            prose_filename = Path(normalize_text(payload.get("work_prose_file"))).name if not is_empty(payload.get("work_prose_file")) else ""
            prose_path: Path | None = None
            prose_missing_reason: str | None = None
            if project_folder and prose_filename:
                prose_path = works_root / project_folder / works_prose_root / prose_filename
            elif prose_filename:
                prose_missing_reason = "missing_project_folder"
            else:
                prose_missing_reason = "missing_work_prose_file"
            prose_work[work_id] = fingerprint_media_path(
                prose_path,
                base_dir=projects_base_dir,
                previous_entry=previous_prose_work.get(work_id),
                missing_reason=prose_missing_reason,
            )

    if "Series" in wb.sheetnames:
        ws = wb["Series"]
        headers = workbook_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            payload = normalize_row_payload(headers, row)
            raw_series_id = payload.get("series_id")
            if is_empty(raw_series_id):
                continue
            series_id = normalize_text(raw_series_id)
            if not series_id:
                continue
            series[series_id] = {
                "hash": stable_payload_hash(payload),
                "status": normalize_status(payload.get("status")),
            }
            primary_work_id = slug_id(payload.get("primary_work_id")) if not is_empty(payload.get("primary_work_id")) else ""
            project_folder = work_project_folder_by_id.get(primary_work_id, "")
            prose_filename = Path(normalize_text(payload.get("series_prose_file"))).name if not is_empty(payload.get("series_prose_file")) else ""
            prose_path = None
            prose_missing_reason = None
            if project_folder and prose_filename:
                prose_path = works_root / project_folder / works_prose_root / prose_filename
            elif prose_filename:
                prose_missing_reason = "missing_project_folder"
            else:
                prose_missing_reason = "missing_series_prose_file"
            prose_series[series_id] = fingerprint_media_path(
                prose_path,
                base_dir=projects_base_dir,
                previous_entry=previous_prose_series.get(series_id),
                missing_reason=prose_missing_reason,
            )

    if "WorkDetails" in wb.sheetnames:
        ws = wb["WorkDetails"]
        headers = workbook_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            payload = normalize_row_payload(headers, row)
            raw_work_id = payload.get("work_id")
            raw_detail_id = payload.get("detail_id")
            if is_empty(raw_work_id) or is_empty(raw_detail_id):
                continue
            work_id = slug_id(raw_work_id)
            detail_id = slug_id(raw_detail_id, width=3)
            uid = f"{work_id}-{detail_id}"
            work_details[uid] = {
                "hash": stable_payload_hash(payload),
                "status": normalize_status(payload.get("status")),
                "work_id": work_id,
            }
            project_subfolder = normalize_text(payload.get("project_subfolder"))
            project_filename = normalize_text(payload.get("project_filename"))
            project_folder = work_project_folder_by_id.get(work_id, "")
            resolved_path = None
            missing_reason = None
            if project_folder and project_filename:
                resolved_path = works_root / project_folder
                if project_subfolder:
                    resolved_path = resolved_path / project_subfolder
                resolved_path = resolved_path / project_filename
            elif project_filename:
                missing_reason = "missing_project_folder"
            else:
                missing_reason = "missing_project_filename"
            media_work_details[uid] = fingerprint_media_path(
                resolved_path,
                base_dir=projects_base_dir,
                previous_entry=previous_media_work_details.get(uid),
                missing_reason=missing_reason,
            )

    if "WorkFiles" in wb.sheetnames:
        ws = wb["WorkFiles"]
        headers = workbook_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            payload = normalize_row_payload(headers, row)
            raw_work_id = payload.get("work_id")
            if is_empty(raw_work_id):
                continue
            work_id = slug_id(raw_work_id)
            work_files_rows.setdefault(work_id, []).append(payload)

    if "WorkLinks" in wb.sheetnames:
        ws = wb["WorkLinks"]
        headers = workbook_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            payload = normalize_row_payload(headers, row)
            raw_work_id = payload.get("work_id")
            if is_empty(raw_work_id):
                continue
            work_id = slug_id(raw_work_id)
            work_links_rows.setdefault(work_id, []).append(payload)

    if "Moments" in wb.sheetnames:
        ws = wb["Moments"]
        headers = workbook_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            payload = normalize_row_payload(headers, row)
            raw_moment_id = payload.get("moment_id")
            if is_empty(raw_moment_id):
                continue
            moment_id = normalize_text(raw_moment_id).lower()
            if not moment_id:
                continue
            moments[moment_id] = {
                "hash": stable_payload_hash(payload),
                "status": normalize_status(payload.get("status")),
            }
            project_filename = normalize_text(payload.get("project_filename")) or f"{moment_id}.jpg"
            media_moments[moment_id] = fingerprint_media_path(
                moments_images_root / project_filename,
                base_dir=projects_base_dir,
                previous_entry=previous_media_moments.get(moment_id),
                missing_reason="missing_project_filename",
            )
            prose_moments[moment_id] = fingerprint_media_path(
                moments_root / f"{moment_id}.md",
                base_dir=projects_base_dir,
                previous_entry=previous_prose_moments.get(moment_id),
            )

    work_files = {
        work_id: {"hash": stable_payload_hash(rows), "count": len(rows)}
        for work_id, rows in work_files_rows.items()
    }
    work_links = {
        work_id: {"hash": stable_payload_hash(rows), "count": len(rows)}
        for work_id, rows in work_links_rows.items()
    }

    return {
        "schema": BUILD_STATE_SCHEMA,
        "planner_version": BUILD_STATE_PLANNER_VERSION,
        "migration_note": BUILD_STATE_MIGRATION_NOTE,
        "updated_at_utc": utc_timestamp_now(),
        "inputs": {
            "works": works,
            "series": series,
            "work_details": work_details,
            "work_files": work_files,
            "work_links": work_links,
            "moments": moments,
            "media": {
                "work": media_work,
                "work_details": media_work_details,
                "moment": media_moments,
            },
            "prose": {
                "work": prose_work,
                "series": prose_series,
                "moment": prose_moments,
            },
        },
    }


def load_build_state(path: Path) -> Dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    if not isinstance(payload, dict):
        return None
    schema = payload.get("schema")
    if schema != BUILD_STATE_SCHEMA and schema not in LEGACY_BUILD_STATE_SCHEMAS:
        return None
    inputs = payload.get("inputs")
    if not isinstance(inputs, dict):
        return None
    planner_version = payload.get("planner_version")
    legacy_state_loaded = schema != BUILD_STATE_SCHEMA or planner_version != BUILD_STATE_PLANNER_VERSION
    normalized = {
        "schema": BUILD_STATE_SCHEMA,
        "planner_version": BUILD_STATE_PLANNER_VERSION,
        "migration_note": BUILD_STATE_MIGRATION_NOTE,
        "updated_at_utc": payload.get("updated_at_utc") if isinstance(payload.get("updated_at_utc"), str) else utc_timestamp_now(),
        "inputs": inputs,
        "_migration": {
            "legacy_state_loaded": legacy_state_loaded,
            "source_schema": schema,
            "source_planner_version": planner_version,
        },
    }
    return normalized


def write_build_state(path: Path, state: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def diff_state_entries(previous: Dict[str, Any], current: Dict[str, Any]) -> Dict[str, Set[str]]:
    previous_ids = set(previous.keys())
    current_ids = set(current.keys())
    added = current_ids - previous_ids
    removed = previous_ids - current_ids
    changed = {entry_id for entry_id in (previous_ids & current_ids) if previous[entry_id].get("hash") != current[entry_id].get("hash")}
    return {
        "added": added,
        "changed": changed,
        "removed": removed,
    }


def diff_media_entries(previous: Dict[str, Any], current: Dict[str, Any]) -> Dict[str, Set[str]]:
    previous_ids = set(previous.keys())
    current_ids = set(current.keys())
    added = current_ids - previous_ids
    removed = previous_ids - current_ids
    changed: Set[str] = set()
    for entry_id in previous_ids & current_ids:
        previous_entry = previous[entry_id]
        current_entry = current[entry_id]
        if previous_entry.get("path") != current_entry.get("path"):
            changed.add(entry_id)
            continue
        if bool(previous_entry.get("exists")) != bool(current_entry.get("exists")):
            changed.add(entry_id)
            continue
        if previous_entry.get("missing_reason") != current_entry.get("missing_reason"):
            changed.add(entry_id)
            continue
        if bool(current_entry.get("exists")):
            if previous_entry.get("hash") != current_entry.get("hash"):
                changed.add(entry_id)
    return {
        "added": added,
        "changed": changed,
        "removed": removed,
    }


def ids_with_status(entries: Dict[str, Dict[str, Any]], statuses: Set[str]) -> Set[str]:
    return {
        entry_id
        for entry_id, payload in entries.items()
        if normalize_status(payload.get("status")) in statuses
    }


def series_ids_for_work_ids(entries: Dict[str, Dict[str, Any]], work_ids: Iterable[str]) -> Set[str]:
    out: Set[str] = set()
    for work_id in work_ids:
        for series_id in entries.get(work_id, {}).get("series_ids", []):
            out.add(series_id)
    return out


def summarize_ids(ids: Set[str], limit: int = 8) -> str:
    ordered = sorted(ids)
    if len(ordered) <= limit:
        return ", ".join(ordered)
    head = ", ".join(ordered[:limit])
    return f"{head}, +{len(ordered) - limit} more"


def confirm_continue(*, no_confirm: bool) -> None:
    if no_confirm:
        print("Continue? [Y|N] auto-accepted via --no-confirm")
        return

    while True:
        try:
            response = input("Continue? [Y|N] ").strip().lower()
        except EOFError as exc:
            raise SystemExit("Confirmation required. Re-run with --no-confirm to continue without prompting.") from exc

        if response in {"", "y", "yes"}:
            return
        if response in {"n", "no"}:
            print("Cancelled.")
            raise SystemExit(0)
        print("Please enter Y or N.")


def summarize_run_text(
    *,
    workbook_change_count: int,
    media_change_count: int,
    action_count: int,
    search_rebuilt: bool,
    status: str,
) -> str:
    if status != "completed":
        return "Build failed."
    if workbook_change_count == 0 and media_change_count == 0 and action_count == 0 and not search_rebuilt:
        return "No changes detected."

    parts: list[str] = []
    if workbook_change_count:
        parts.append(f"{workbook_change_count} workbook change{'s' if workbook_change_count != 1 else ''}")
    if media_change_count:
        parts.append(f"{media_change_count} media change{'s' if media_change_count != 1 else ''}")
    if action_count:
        parts.append(f"{action_count} planned action{'s' if action_count != 1 else ''}")
    if search_rebuilt:
        parts.append("catalogue search rebuilt")
    return "; ".join(parts) if parts else "Build completed."


def build_activity_entry(
    *,
    time_utc: str,
    status: str,
    dry_run: bool,
    planner_mode: str,
    workbook_changes: Dict[str, Set[str]],
    media_changes: Dict[str, Set[str]],
    prose_changes: Dict[str, Set[str]],
    actions: Dict[str, Any],
    results: Dict[str, Any],
) -> Dict[str, Any]:
    workbook_change_count = sum(len(values) for values in workbook_changes.values())
    media_change_count = sum(len(values) for values in media_changes.values())
    action_count = sum(
        int(value)
        for key, value in actions.items()
        if key != "rebuild_search" and isinstance(value, int)
    )
    search_rebuilt = bool(actions.get("rebuild_search"))
    return {
        "id": f"{time_utc}-build_catalogue",
        "time_utc": time_utc,
        "script": "build_catalogue",
        "status": status,
        "dry_run": dry_run,
        "planner_mode": planner_mode,
        "summary": summarize_run_text(
            workbook_change_count=workbook_change_count,
            media_change_count=media_change_count,
            action_count=action_count,
            search_rebuilt=search_rebuilt,
            status=status,
        ),
        "changes": {
            "workbook": {key: sorted(values) for key, values in workbook_changes.items()},
            "media": {key: sorted(values) for key, values in media_changes.items()},
            "prose": {key: sorted(values) for key, values in prose_changes.items()},
        },
        "actions": actions,
        "results": results,
    }


def load_json_object(path: Path) -> Dict[str, Any] | None:
    if not path.exists():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload if isinstance(payload, dict) else None


def write_json_object(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def plan_stale_artifact_paths(
    repo_root: Path,
    *,
    removed_work_ids: Set[str],
    removed_series_ids: Set[str],
    removed_detail_uids: Set[str],
    removed_moment_ids: Set[str],
) -> Dict[str, list[Path]]:
    planned: list[Path] = []
    work_details_dir = repo_root / "_work_details"

    for work_id in sorted(removed_work_ids):
        planned.append(repo_root / "_works" / f"{work_id}.md")
        planned.append(repo_root / "assets/works/index" / f"{work_id}.json")
        planned.extend(sorted(work_details_dir.glob(f"{work_id}-*.md")))
    for uid in sorted(removed_detail_uids):
        planned.append(work_details_dir / f"{uid}.md")
    for series_id in sorted(removed_series_ids):
        planned.append(repo_root / "_series" / f"{series_id}.md")
        planned.append(repo_root / "assets/series/index" / f"{series_id}.json")
    for moment_id in sorted(removed_moment_ids):
        planned.append(repo_root / "_moments" / f"{moment_id}.md")
        planned.append(repo_root / "assets/moments/index" / f"{moment_id}.json")

    seen: Set[Path] = set()
    existing: list[Path] = []
    missing: list[Path] = []
    for path in planned:
        if path in seen:
            continue
        seen.add(path)
        if path.exists():
            existing.append(path)
        else:
            missing.append(path)
    return {
        "existing": existing,
        "missing": missing,
    }


def collect_matching_paths(root: Path, patterns: Iterable[str]) -> list[Path]:
    collected: list[Path] = []
    seen: Set[Path] = set()
    if not root.exists():
        return collected
    for pattern in patterns:
        for path in sorted(root.glob(pattern)):
            if path in seen or not path.is_file():
                continue
            seen.add(path)
            collected.append(path)
    return collected


def plan_local_media_cleanup_paths(
    *,
    media_base_dir: Path | None,
    removed_work_ids: Set[str],
    removed_detail_uids: Set[str],
    removed_moment_ids: Set[str],
) -> Dict[str, list[Path] | bool]:
    if media_base_dir is None:
        return {
            "available": False,
            "existing": [],
            "missing": [],
        }

    work_input_dir = media_base_dir / media_mode_input_subdir(PIPELINE_CONFIG, "work")
    work_output_dir = media_base_dir / media_mode_output_subdir(PIPELINE_CONFIG, "work")
    detail_input_dir = media_base_dir / media_mode_input_subdir(PIPELINE_CONFIG, "work_details")
    detail_output_dir = media_base_dir / media_mode_output_subdir(PIPELINE_CONFIG, "work_details")
    moment_input_dir = media_base_dir / media_mode_input_subdir(PIPELINE_CONFIG, "moment")
    moment_output_dir = media_base_dir / media_mode_output_subdir(PIPELINE_CONFIG, "moment")
    work_files_dir = media_base_dir / media_work_files_subdir(PIPELINE_CONFIG)

    primary_subdir = str(PIPELINE_CONFIG["variants"]["primary"]["output_subdir"])
    thumb_subdir = str(PIPELINE_CONFIG["variants"]["thumb"]["output_subdir"])
    output_format = str(PIPELINE_CONFIG["encoding"]["format"])

    planned: list[Path] = []

    for work_id in sorted(removed_work_ids):
        planned.extend(collect_matching_paths(work_input_dir, [f"{work_id}.*"]))
        planned.extend(
            collect_matching_paths(
                work_output_dir / primary_subdir,
                [f"{work_id}-primary-*.{output_format}"],
            )
        )
        planned.extend(
            collect_matching_paths(
                work_output_dir / thumb_subdir,
                [f"{work_id}-thumb-*.{output_format}"],
            )
        )
        planned.extend(collect_matching_paths(work_files_dir, [f"{work_id}-*"]))

    for uid in sorted(removed_detail_uids):
        planned.extend(collect_matching_paths(detail_input_dir, [f"{uid}.*"]))
        planned.extend(
            collect_matching_paths(
                detail_output_dir / primary_subdir,
                [f"{uid}-primary-*.{output_format}"],
            )
        )
        planned.extend(
            collect_matching_paths(
                detail_output_dir / thumb_subdir,
                [f"{uid}-thumb-*.{output_format}"],
            )
        )

    for moment_id in sorted(removed_moment_ids):
        planned.extend(collect_matching_paths(moment_input_dir, [f"{moment_id}.*"]))
        planned.extend(
            collect_matching_paths(
                moment_output_dir / primary_subdir,
                [f"{moment_id}-primary-*.{output_format}"],
            )
        )
        planned.extend(
            collect_matching_paths(
                moment_output_dir / thumb_subdir,
                [f"{moment_id}-thumb-*.{output_format}"],
            )
        )

    seen: Set[Path] = set()
    existing: list[Path] = []
    missing: list[Path] = []
    for path in planned:
        if path in seen:
            continue
        seen.add(path)
        if path.exists():
            existing.append(path)
        else:
            missing.append(path)
    return {
        "available": True,
        "existing": existing,
        "missing": missing,
    }


def prune_tag_assignments_for_removed_rows(
    repo_root: Path,
    *,
    removed_series_ids: Set[str],
    removed_work_ids: Set[str],
    write: bool,
) -> Dict[str, int]:
    path = repo_root / "assets/studio/data/tag_assignments.json"
    payload = load_json_object(path)
    if payload is None:
        return {
            "series_removed": 0,
            "work_overrides_removed": 0,
            "series_touched": 0,
            "payload_written": 0,
        }

    series_map = payload.get("series")
    if not isinstance(series_map, dict):
        raise SystemExit(f"Invalid tag assignments payload: {path}")

    now_utc = utc_timestamp_now()
    series_removed = 0
    work_overrides_removed = 0
    series_touched = 0

    for series_id in sorted(removed_series_ids):
        if series_id not in series_map:
            continue
        row = series_map.get(series_id)
        if isinstance(row, dict):
            works = row.get("works")
            if isinstance(works, dict):
                work_overrides_removed += len(works)
        del series_map[series_id]
        series_removed += 1

    for series_id, row in series_map.items():
        if not isinstance(row, dict):
            continue
        works = row.get("works")
        if not isinstance(works, dict):
            continue
        removed_here = 0
        for work_id in sorted(removed_work_ids):
            if work_id not in works:
                continue
            del works[work_id]
            removed_here += 1
        if removed_here:
            row["updated_at_utc"] = now_utc
            work_overrides_removed += removed_here
            series_touched += 1

    payload_changed = bool(series_removed or work_overrides_removed)
    if payload_changed:
        payload["series"] = series_map
        payload["updated_at_utc"] = now_utc
        if not normalize_text(payload.get("tag_assignments_version")):
            payload["tag_assignments_version"] = "tag_assignments_v1"
        if write:
            write_json_object(path, payload)

    return {
        "series_removed": series_removed,
        "work_overrides_removed": work_overrides_removed,
        "series_touched": series_touched,
        "payload_written": 1 if payload_changed and write else 0,
    }


def delete_stale_artifact_paths(paths: Iterable[Path]) -> int:
    deleted = 0
    for path in paths:
        if not path.exists():
            continue
        path.unlink()
        deleted += 1
    return deleted


def main() -> int:
    media_base = env_var_value(PIPELINE_CONFIG, "media_base_dir")
    projects_base = env_var_value(PIPELINE_CONFIG, "projects_base_dir")
    default_work_input = join_base_and_subdir(media_base, media_mode_input_subdir(PIPELINE_CONFIG, "work"))
    default_work_output = join_base_and_subdir(media_base, media_mode_output_subdir(PIPELINE_CONFIG, "work"))
    default_detail_input = join_base_and_subdir(media_base, media_mode_input_subdir(PIPELINE_CONFIG, "work_details"))
    default_detail_output = join_base_and_subdir(media_base, media_mode_output_subdir(PIPELINE_CONFIG, "work_details"))
    default_moment_input = join_base_and_subdir(media_base, media_mode_input_subdir(PIPELINE_CONFIG, "moment"))
    default_moment_output = join_base_and_subdir(media_base, media_mode_output_subdir(PIPELINE_CONFIG, "moment"))

    ap = argparse.ArgumentParser(description="Build the catalogue pipeline: copy -> srcset -> generate -> search.")
    ap.add_argument("--xlsx", default="data/works.xlsx", help="Path to workbook")
    ap.add_argument(
        "--mode",
        action="append",
        choices=["work", "work_details", "moment"],
        help="Flow(s) to run. Repeat to run multiple modes. Default: all.",
    )
    ap.add_argument(
        "--input-dir",
        default=default_work_input,
        help="Input directory for source images",
    )
    ap.add_argument(
        "--output-dir",
        default=default_work_output,
        help="Output directory for srcset derivatives",
    )
    ap.add_argument(
        "--detail-input-dir",
        default=default_detail_input,
        help="Input directory for work detail source images",
    )
    ap.add_argument(
        "--detail-output-dir",
        default=default_detail_output,
        help="Output directory for work detail srcset derivatives",
    )
    ap.add_argument(
        "--moment-input-dir",
        default=default_moment_input,
        help="Input directory for moment source images",
    )
    ap.add_argument(
        "--moment-output-dir",
        default=default_moment_output,
        help="Output directory for moment srcset derivatives",
    )
    ap.add_argument(
        "--jobs",
        type=int,
        default=int(os.environ.get(SRCSET_JOBS_ENV_NAME, "4")),
        help="Parallel jobs",
    )
    ap.add_argument("--force-generate", action="store_true", help="Pass --force to generate_work_pages.py and the catalogue search rebuild")
    ap.add_argument("--dry-run", action="store_true", help="Preview mode; no writes/deletes.")
    ap.add_argument("--plan", action="store_true", help="Print the inferred execution plan and exit.")
    ap.add_argument("--no-confirm", action="store_true", help="Skip the post-plan confirmation prompt and continue immediately.")
    ap.add_argument("--full", action="store_true", help="Ignore previous planner state and rebuild all workbook-backed generation targets.")
    ap.add_argument("--reset-state", action="store_true", help=f"Remove the stored planner state at {DEFAULT_BUILD_STATE_PATH} before planning.")
    ap.add_argument(
        "--work-ids",
        default="",
        help="Comma-separated work_ids/ranges filter for this run (e.g. 66-74,38-40,1).",
    )
    ap.add_argument("--work-ids-file", default="", help="Path to work_ids file (one id per line).")
    ap.add_argument("--series-ids", default="", help="Comma-separated series_ids passed to generation.")
    ap.add_argument("--series-ids-file", default="", help="Path to series_ids file (one id per line).")
    ap.add_argument("--moment-ids", default="", help="Comma-separated moment_ids filter for this run.")
    ap.add_argument("--moment-ids-file", default="", help="Path to moment_ids file (one id per line).")
    args = ap.parse_args()
    selected_modes = set(args.mode or ["work", "work_details", "moment"])

    def require_non_empty(label: str, value: str, flag: str) -> None:
        if normalize_text(value) == "":
            raise SystemExit(
                f"Missing {label}. Set {MEDIA_BASE_DIR_ENV_NAME} or pass {flag}."
            )

    if "work" in selected_modes:
        require_non_empty("work input dir", args.input_dir, "--input-dir")
        require_non_empty("work output dir", args.output_dir, "--output-dir")
    if "work_details" in selected_modes:
        require_non_empty("work_details input dir", args.detail_input_dir, "--detail-input-dir")
        require_non_empty("work_details output dir", args.detail_output_dir, "--detail-output-dir")
    if "moment" in selected_modes:
        require_non_empty("moment input dir", args.moment_input_dir, "--moment-input-dir")
        require_non_empty("moment output dir", args.moment_output_dir, "--moment-output-dir")

    explicit_work_ids = read_optional_ids_file(args.work_ids_file)
    if args.work_ids:
        explicit_work_ids.update(parse_work_id_selection(args.work_ids))
    work_filter = explicit_work_ids if explicit_work_ids else None

    explicit_series_ids = read_optional_ids_file(args.series_ids_file)
    if args.series_ids:
        explicit_series_ids.update({normalize_text(s.strip()) for s in args.series_ids.split(",") if s.strip()})
    series_filter = explicit_series_ids if explicit_series_ids else None

    explicit_moment_ids = read_optional_ids_file(args.moment_ids_file)
    if args.moment_ids:
        explicit_moment_ids.update({normalize_text(m.strip()).lower() for m in args.moment_ids.split(",") if m.strip()})
    moment_filter = explicit_moment_ids if explicit_moment_ids else None
    manual_scope_requested = bool(work_filter is not None or series_filter is not None or moment_filter is not None)

    repo_root = Path(__file__).resolve().parents[1]
    state_path = repo_root / DEFAULT_BUILD_STATE_PATH
    xlsx_path = (repo_root / args.xlsx).resolve()
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")
    if args.reset_state and state_path.exists():
        state_path.unlink()

    previous_state = None if args.full else load_build_state(state_path)
    previous_inputs = previous_state.get("inputs", {}) if previous_state else {}
    previous_state_migration = previous_state.get("_migration", {}) if previous_state else {}
    previous_media_inputs = previous_inputs.get("media", {}) if isinstance(previous_inputs.get("media"), dict) else {}
    previous_prose_inputs = previous_inputs.get("prose", {}) if isinstance(previous_inputs.get("prose"), dict) else {}
    current_state = build_workbook_state(
        xlsx_path,
        projects_base_dir=Path(projects_base).expanduser() if projects_base else Path("."),
        previous_media=previous_media_inputs,
        previous_prose=previous_prose_inputs,
    )
    bootstrap_state = previous_state is None
    legacy_state_loaded = bool(previous_state_migration.get("legacy_state_loaded"))
    current_inputs = current_state["inputs"]

    current_works = current_inputs["works"]
    current_series = current_inputs["series"]
    current_work_details = current_inputs["work_details"]
    current_work_files = current_inputs["work_files"]
    current_work_links = current_inputs["work_links"]
    current_moments = current_inputs["moments"]
    current_media_inputs = current_inputs.get("media", {})
    current_media_work = current_media_inputs.get("work", {}) if isinstance(current_media_inputs.get("work"), dict) else {}
    current_media_work_details = current_media_inputs.get("work_details", {}) if isinstance(current_media_inputs.get("work_details"), dict) else {}
    current_media_moments = current_media_inputs.get("moment", {}) if isinstance(current_media_inputs.get("moment"), dict) else {}
    current_prose_inputs = current_inputs.get("prose", {})
    current_prose_work = current_prose_inputs.get("work", {}) if isinstance(current_prose_inputs.get("work"), dict) else {}
    current_prose_series = current_prose_inputs.get("series", {}) if isinstance(current_prose_inputs.get("series"), dict) else {}
    current_prose_moments = current_prose_inputs.get("moment", {}) if isinstance(current_prose_inputs.get("moment"), dict) else {}

    previous_works = previous_inputs.get("works", {})
    previous_series = previous_inputs.get("series", {})
    previous_work_details = previous_inputs.get("work_details", {})
    previous_work_files = previous_inputs.get("work_files", {})
    previous_work_links = previous_inputs.get("work_links", {})
    previous_moments = previous_inputs.get("moments", {})
    previous_media_work = previous_media_inputs.get("work", {}) if isinstance(previous_media_inputs.get("work"), dict) else {}
    previous_media_work_details = previous_media_inputs.get("work_details", {}) if isinstance(previous_media_inputs.get("work_details"), dict) else {}
    previous_media_moments = previous_media_inputs.get("moment", {}) if isinstance(previous_media_inputs.get("moment"), dict) else {}
    previous_prose_work = previous_prose_inputs.get("work", {}) if isinstance(previous_prose_inputs.get("work"), dict) else {}
    previous_prose_series = previous_prose_inputs.get("series", {}) if isinstance(previous_prose_inputs.get("series"), dict) else {}
    previous_prose_moments = previous_prose_inputs.get("moment", {}) if isinstance(previous_prose_inputs.get("moment"), dict) else {}
    media_tracking_available = bool(previous_state) and bool(previous_media_inputs)
    prose_tracking_available = bool(previous_state) and bool(previous_prose_inputs)

    work_diff = diff_state_entries(previous_works, current_works)
    series_diff = diff_state_entries(previous_series, current_series)
    detail_diff = diff_state_entries(previous_work_details, current_work_details)
    work_files_diff = diff_state_entries(previous_work_files, current_work_files)
    work_links_diff = diff_state_entries(previous_work_links, current_work_links)
    moments_diff = diff_state_entries(previous_moments, current_moments)
    media_work_diff = diff_media_entries(previous_media_work, current_media_work) if media_tracking_available else {"added": set(), "changed": set(), "removed": set()}
    media_work_details_diff = diff_media_entries(previous_media_work_details, current_media_work_details) if media_tracking_available else {"added": set(), "changed": set(), "removed": set()}
    media_moments_diff = diff_media_entries(previous_media_moments, current_media_moments) if media_tracking_available else {"added": set(), "changed": set(), "removed": set()}
    prose_work_diff = diff_media_entries(previous_prose_work, current_prose_work) if prose_tracking_available else {"added": set(), "changed": set(), "removed": set()}
    prose_series_diff = diff_media_entries(previous_prose_series, current_prose_series) if prose_tracking_available else {"added": set(), "changed": set(), "removed": set()}
    prose_moments_diff = diff_media_entries(previous_prose_moments, current_prose_moments) if prose_tracking_available else {"added": set(), "changed": set(), "removed": set()}

    if args.full:
        work_diff = {"added": set(current_works.keys()), "changed": set(), "removed": set()}
        series_diff = {"added": set(current_series.keys()), "changed": set(), "removed": set()}
        detail_diff = {"added": set(current_work_details.keys()), "changed": set(), "removed": set()}
        work_files_diff = {"added": set(current_work_files.keys()), "changed": set(), "removed": set()}
        work_links_diff = {"added": set(current_work_links.keys()), "changed": set(), "removed": set()}
        moments_diff = {"added": set(current_moments.keys()), "changed": set(), "removed": set()}
        media_work_diff = {"added": set(current_media_work.keys()), "changed": set(), "removed": set()}
        media_work_details_diff = {"added": set(current_media_work_details.keys()), "changed": set(), "removed": set()}
        media_moments_diff = {"added": set(current_media_moments.keys()), "changed": set(), "removed": set()}
        prose_work_diff = {"added": set(current_prose_work.keys()), "changed": set(), "removed": set()}
        prose_series_diff = {"added": set(current_prose_series.keys()), "changed": set(), "removed": set()}
        prose_moments_diff = {"added": set(current_prose_moments.keys()), "changed": set(), "removed": set()}

    draft_work_ids = ids_with_status(current_works, {"draft"})
    draft_series_ids = ids_with_status(current_series, {"draft"})
    draft_detail_uids = ids_with_status(current_work_details, {"draft"})
    draft_moment_ids = ids_with_status(current_moments, {"draft"})

    explicit_media_statuses = {"draft", "published"}
    changed_work_media_ids = media_work_diff["added"] | media_work_diff["changed"] | media_work_diff["removed"]
    changed_detail_media_uids = media_work_details_diff["added"] | media_work_details_diff["changed"] | media_work_details_diff["removed"]
    changed_moment_media_ids = media_moments_diff["added"] | media_moments_diff["changed"] | media_moments_diff["removed"]
    changed_work_prose_ids = prose_work_diff["added"] | prose_work_diff["changed"] | prose_work_diff["removed"]
    changed_series_prose_ids = prose_series_diff["added"] | prose_series_diff["changed"] | prose_series_diff["removed"]
    changed_moment_prose_ids = prose_moments_diff["added"] | prose_moments_diff["changed"] | prose_moments_diff["removed"]
    work_media_candidate_ids = set()
    if "work" in selected_modes:
        if work_filter is not None:
            work_media_candidate_ids = {
                work_id
                for work_id, payload in current_works.items()
                if work_id in work_filter and normalize_status(payload.get("status")) in explicit_media_statuses
            }
        elif not manual_scope_requested:
            work_media_candidate_ids = set(draft_work_ids)
            work_media_candidate_ids.update(changed_work_media_ids & set(current_works.keys()))

    detail_media_candidate_uids = set()
    if "work_details" in selected_modes:
        if work_filter is not None:
            detail_media_candidate_uids = {
                uid
                for uid, payload in current_work_details.items()
                if payload.get("work_id") in work_filter and normalize_status(payload.get("status")) in explicit_media_statuses
            }
        elif not manual_scope_requested:
            detail_media_candidate_uids = set(draft_detail_uids)
            detail_media_candidate_uids.update(changed_detail_media_uids & set(current_work_details.keys()))

    moment_media_candidate_ids = set()
    if "moment" in selected_modes:
        if moment_filter is not None:
            moment_media_candidate_ids = {
                moment_id
                for moment_id, payload in current_moments.items()
                if moment_id in moment_filter and normalize_status(payload.get("status")) in explicit_media_statuses
            }
        elif not manual_scope_requested:
            moment_media_candidate_ids = set(draft_moment_ids)
            moment_media_candidate_ids.update(changed_moment_media_ids & set(current_moments.keys()))

    changed_work_row_ids = work_diff["added"] | work_diff["changed"]
    changed_series_row_ids = series_diff["added"] | series_diff["changed"]
    changed_detail_uids = detail_diff["added"] | detail_diff["changed"]
    changed_moment_ids = moments_diff["added"] | moments_diff["changed"]

    changed_detail_parent_ids = work_ids_from_detail_uids(changed_detail_uids)
    changed_detail_parent_ids.update(
        previous_work_details[uid]["work_id"]
        for uid in detail_diff["removed"]
        if uid in previous_work_details and previous_work_details[uid].get("work_id")
    )

    changed_work_file_ids = (work_files_diff["added"] | work_files_diff["changed"] | work_files_diff["removed"]) & set(current_works.keys())
    changed_work_link_ids = (work_links_diff["added"] | work_links_diff["changed"] | work_links_diff["removed"]) & set(current_works.keys())

    planned_generate_work_ids: Set[str] = set()
    if "work" in selected_modes:
        if work_filter is not None:
            planned_generate_work_ids.update(
                {
                    work_id
                    for work_id, payload in current_works.items()
                    if work_id in work_filter and normalize_status(payload.get("status")) in explicit_media_statuses
                }
            )
        elif not manual_scope_requested:
            planned_generate_work_ids.update(draft_work_ids)
            planned_generate_work_ids.update(changed_work_row_ids)
            planned_generate_work_ids.update(changed_work_media_ids)
            planned_generate_work_ids.update(changed_work_prose_ids)
            planned_generate_work_ids.update(changed_work_file_ids)
            planned_generate_work_ids.update(changed_work_link_ids)
    if "work_details" in selected_modes:
        if work_filter is not None:
            planned_generate_work_ids.update(
                {
                    work_id
                    for work_id, payload in current_works.items()
                    if work_id in work_filter and normalize_status(payload.get("status")) in explicit_media_statuses
                }
            )
        elif not manual_scope_requested:
            planned_generate_work_ids.update(work_ids_from_detail_uids(draft_detail_uids))
            planned_generate_work_ids.update(changed_detail_parent_ids)
            planned_generate_work_ids.update(work_ids_from_detail_uids(changed_detail_media_uids))
    planned_generate_work_ids &= set(current_works.keys())

    planned_generate_moment_ids: Set[str] = set()
    if "moment" in selected_modes:
        if moment_filter is not None:
            planned_generate_moment_ids.update(
                {
                    moment_id
                    for moment_id, payload in current_moments.items()
                    if moment_id in moment_filter and normalize_status(payload.get("status")) in explicit_media_statuses
                }
            )
        elif not manual_scope_requested:
            planned_generate_moment_ids.update(draft_moment_ids)
            planned_generate_moment_ids.update(changed_moment_ids)
            planned_generate_moment_ids.update(changed_moment_media_ids)
            planned_generate_moment_ids.update(changed_moment_prose_ids)
    planned_generate_moment_ids &= set(current_moments.keys())

    planned_series_ids: Set[str] = set()
    if series_filter is not None:
        planned_series_ids.update(set(current_series.keys()) & series_filter)
    elif work_filter is not None and "work" in selected_modes:
        planned_series_ids.update(series_ids_for_work_ids(current_works, planned_generate_work_ids))
        planned_series_ids.update(series_ids_for_work_ids(previous_works, planned_generate_work_ids))
        planned_series_ids &= set(current_series.keys())
    elif "work" in selected_modes and not manual_scope_requested:
        planned_series_ids.update(draft_series_ids)
        planned_series_ids.update(changed_series_row_ids)
        planned_series_ids.update(changed_series_prose_ids)
        planned_series_ids.update(series_ids_for_work_ids(current_works, planned_generate_work_ids))
        planned_series_ids.update(series_ids_for_work_ids(previous_works, changed_work_row_ids | work_diff["removed"]))
        planned_series_ids &= set(current_series.keys())

    removed_work_ids = set(work_diff["removed"]) if "work" in selected_modes else set()
    removed_series_ids = set(series_diff["removed"]) if "work" in selected_modes else set()
    removed_detail_uids = set(detail_diff["removed"]) if "work_details" in selected_modes else set()
    removed_moment_ids = set(moments_diff["removed"]) if "moment" in selected_modes else set()

    if not manual_scope_requested and "work" in selected_modes:
        planned_series_ids.update(series_ids_for_work_ids(previous_works, removed_work_ids))
        planned_series_ids &= set(current_series.keys())

    stale_cleanup_requested = bool(
        not manual_scope_requested
        and (removed_work_ids or removed_series_ids or removed_detail_uids or removed_moment_ids)
    )
    stale_cleanup_plan = (
        plan_stale_artifact_paths(
            repo_root,
            removed_work_ids=removed_work_ids,
            removed_series_ids=removed_series_ids,
            removed_detail_uids=removed_detail_uids,
            removed_moment_ids=removed_moment_ids,
        )
        if stale_cleanup_requested
        else {"existing": [], "missing": []}
    )
    media_base_dir = Path(media_base).expanduser() if media_base else None
    local_media_cleanup_plan = (
        plan_local_media_cleanup_paths(
            media_base_dir=media_base_dir,
            removed_work_ids=removed_work_ids,
            removed_detail_uids=removed_detail_uids,
            removed_moment_ids=removed_moment_ids,
        )
        if stale_cleanup_requested
        else {"available": bool(media_base_dir), "existing": [], "missing": []}
    )

    plan_needs_generate = bool(planned_generate_work_ids or planned_generate_moment_ids or planned_series_ids or stale_cleanup_requested)
    non_prose_generation_requested = bool(
        manual_scope_requested
        or args.full
        or draft_work_ids
        or draft_detail_uids
        or draft_moment_ids
        or changed_work_row_ids
        or changed_series_row_ids
        or changed_detail_uids
        or changed_moment_ids
        or changed_work_media_ids
        or changed_detail_media_uids
        or changed_moment_media_ids
        or changed_work_file_ids
        or changed_work_link_ids
        or stale_cleanup_requested
    )
    plan_needs_search = bool(non_prose_generation_requested)
    planner_mode = "full" if args.full else ("bootstrap" if bootstrap_state else "incremental")

    def print_plan_summary() -> None:
        print("\n==> Build Plan")
        if args.full:
            print("Planner mode: full rebuild from current workbook state.")
        elif bootstrap_state:
            print(f"Planner mode: bootstrap (no prior state at {DEFAULT_BUILD_STATE_PATH}).")
        else:
            print(f"Planner state: {DEFAULT_BUILD_STATE_PATH}")
            print(f"Planner version: {BUILD_STATE_PLANNER_VERSION}")
            if legacy_state_loaded:
                source_schema = previous_state_migration.get("source_schema") or "unknown"
                source_planner_version = previous_state_migration.get("source_planner_version")
                if source_planner_version is None:
                    print(f"- loaded legacy planner state ({source_schema}); it will be rewritten with planner_version {BUILD_STATE_PLANNER_VERSION} on the next successful write run")
                else:
                    print(
                        f"- loaded legacy planner state ({source_schema}, planner_version {source_planner_version}); "
                        f"it will be rewritten with planner_version {BUILD_STATE_PLANNER_VERSION} on the next successful write run"
                    )
        print(f"Selected modes: {', '.join(sorted(selected_modes))}")
        print("Workbook changes:")
        print(f"- works: {len(changed_work_row_ids)} changed/new, {len(work_diff['removed'])} removed")
        print(f"- series: {len(changed_series_row_ids)} changed/new, {len(series_diff['removed'])} removed")
        print(f"- work_details: {len(changed_detail_uids)} changed/new, {len(detail_diff['removed'])} removed")
        print(f"- work_files groups: {len(work_files_diff['added'] | work_files_diff['changed'] | work_files_diff['removed'])}")
        print(f"- work_links groups: {len(work_links_diff['added'] | work_links_diff['changed'] | work_links_diff['removed'])}")
        print(f"- moments: {len(changed_moment_ids)} changed/new, {len(moments_diff['removed'])} removed")
        if media_tracking_available or args.full:
            print("Media changes:")
            print(f"- work sources: {len(changed_work_media_ids)} changed/new, {len(media_work_diff['removed'])} removed")
            print(f"- work detail sources: {len(changed_detail_media_uids)} changed/new, {len(media_work_details_diff['removed'])} removed")
            print(f"- moment sources: {len(changed_moment_media_ids)} changed/new, {len(media_moments_diff['removed'])} removed")
        else:
            print("Media changes:")
            print("- source media tracking not yet initialized in planner state; current media will be treated as baseline on the next write run")
        if prose_tracking_available or args.full:
            print("Prose changes:")
            print(f"- work prose sources: {len(changed_work_prose_ids)} changed/new, {len(prose_work_diff['removed'])} removed")
            print(f"- series prose sources: {len(changed_series_prose_ids)} changed/new, {len(prose_series_diff['removed'])} removed")
            print(f"- moment prose sources: {len(changed_moment_prose_ids)} changed/new, {len(prose_moments_diff['removed'])} removed")
        else:
            print("Prose changes:")
            print("- prose tracking not yet initialized in planner state; current prose paths will be treated as baseline on the next successful write run")
        print("Planned scope:")
        print(f"- work media candidates: {len(work_media_candidate_ids)}")
        print(f"- work detail media candidates: {len(detail_media_candidate_uids)}")
        print(f"- moment media candidates: {len(moment_media_candidate_ids)}")
        print(f"- work generation ids: {len(planned_generate_work_ids)}")
        print(f"- moment generation ids: {len(planned_generate_moment_ids)}")
        print(f"- series generation ids: {len(planned_series_ids)}")
        print(f"- stale generated files to delete: {len(stale_cleanup_plan['existing'])}")
        print(f"- stale local media files to delete: {len(local_media_cleanup_plan['existing'])}")
        print(f"- rebuild catalogue search: {'yes' if plan_needs_search else 'no'}")
        if stale_cleanup_requested:
            print("Stale cleanup:")
            if removed_work_ids:
                print(f"- removed works: {summarize_ids(removed_work_ids)}")
            if removed_series_ids:
                print(f"- removed series: {summarize_ids(removed_series_ids)}")
            if removed_detail_uids:
                print(f"- removed work_details: {summarize_ids(removed_detail_uids)}")
            if removed_moment_ids:
                print(f"- removed moments: {summarize_ids(removed_moment_ids)}")
            if stale_cleanup_plan["existing"]:
                print(f"- generated files present to delete: {len(stale_cleanup_plan['existing'])}")
            if stale_cleanup_plan["missing"]:
                print(f"- generated files already missing: {len(stale_cleanup_plan['missing'])}")
            if local_media_cleanup_plan.get("available"):
                if local_media_cleanup_plan["existing"]:
                    print(f"- local media files present to delete: {len(local_media_cleanup_plan['existing'])}")
            else:
                print("- local media cleanup skipped: DOTLINEFORM_MEDIA_BASE_DIR is not configured")

    print_plan_summary()
    if args.plan:
        return 0
    confirm_continue(no_confirm=bool(args.no_confirm))

    print("\n==> Workbook Preflight")
    raise_if_invalid_catalogue_workbook(xlsx_path)
    print("Workbook preflight passed.")

    log_event(
        repo_root,
        "pipeline_start",
        {
            "dry_run": args.dry_run,
            "plan_only": args.plan,
            "full": args.full,
            "modes": sorted(selected_modes),
            "xlsx": str(xlsx_path.relative_to(repo_root)),
            "force_generate": args.force_generate,
        },
    )

    copy_script = repo_root / "scripts/copy_draft_media_files.py"
    make_script = repo_root / "scripts/make_srcset_images.sh"
    generate_script = repo_root / "scripts/generate_work_pages.py"
    search_script = repo_root / "scripts/build_search.rb"
    py = sys.executable

    with tempfile.TemporaryDirectory(prefix="draft-pipeline-") as tmp:
        tmp_dir = Path(tmp)
        draft_ids_file = tmp_dir / "draft_ids.txt"
        copied_ids_file = tmp_dir / "copied_ids.txt"
        draft_detail_uids_file = tmp_dir / "draft_detail_uids.txt"
        copied_detail_uids_file = tmp_dir / "copied_detail_uids.txt"
        success_ids_file = tmp_dir / "success_ids.txt"
        detail_success_ids_file = tmp_dir / "detail_success_ids.txt"
        moment_success_ids_file = tmp_dir / "moment_success_ids.txt"
        draft_moment_ids_file = tmp_dir / "draft_moment_ids.txt"
        copied_moment_ids_file = tmp_dir / "copied_moment_ids.txt"
        generate_moment_ids_file = tmp_dir / "generate_moment_ids.txt"
        generate_ids_file = tmp_dir / "generate_ids.txt"

        draft_ids = set(work_media_candidate_ids)
        draft_ids_file.write_text(
            "\n".join(sorted(draft_ids)) + ("\n" if draft_ids else ""),
            encoding="utf-8",
        )
        print(f"Draft candidates (work): {len(draft_ids)}")

        draft_detail_uids = set(detail_media_candidate_uids)
        draft_detail_uids_file.write_text(
            "\n".join(sorted(draft_detail_uids)) + ("\n" if draft_detail_uids else ""),
            encoding="utf-8",
        )
        print(f"Draft detail candidates (work_details): {len(draft_detail_uids)}")

        draft_moment_ids = set(moment_media_candidate_ids)
        draft_moment_ids_file.write_text(
            "\n".join(sorted(draft_moment_ids)) + ("\n" if draft_moment_ids else ""),
            encoding="utf-8",
        )
        print(f"Draft moment candidates (moment): {len(draft_moment_ids)}")

        copied_ids = read_ids(copied_ids_file)
        generated_work_ids = set()
        if "work" in selected_modes:
            if draft_ids:
                run_step(
                    "Copy Draft Work Files",
                    [
                        py,
                        str(copy_script),
                        "--mode",
                        "work",
                        "--ids-file",
                        str(draft_ids_file),
                        "--copied-ids-file",
                        str(copied_ids_file),
                    ]
                    + ([] if args.dry_run else ["--write"]),
                    cwd=repo_root,
                )

                copied_ids = read_ids(copied_ids_file)
                if copied_ids:
                    # make_srcset_images.sh controls selection via environment variables.
                    env = os.environ.copy()
                    env[SRCSET_SELECTED_IDS_ENV_NAME] = str(copied_ids_file)
                    env[SRCSET_SUCCESS_IDS_ENV_NAME] = str(success_ids_file)

                    run_step(
                        "Generate Srcset Derivatives",
                        [
                            "bash",
                            str(make_script),
                            str(args.input_dir),
                            str(args.output_dir),
                            str(args.jobs),
                        ]
                        + (["--dry-run"] if args.dry_run else []),
                        cwd=repo_root,
                        env=env,
                    )

                    if args.dry_run:
                        generated_work_ids = copied_ids
                    else:
                        generated_work_ids = read_ids(success_ids_file)
                else:
                    print("\n==> Skip Srcset\nNo copied draft files in this run.")
            else:
                print("\n==> Skip Work Copy/Srcset\nNo work-media candidates in this run.")
        else:
            print("\n==> Skip Work\nMode not selected.")
            generated_work_ids = set()

        copied_detail_uids = read_ids(copied_detail_uids_file)
        generated_detail_uids = set()
        if "work_details" in selected_modes:
            if draft_detail_uids:
                run_step(
                    "Copy Draft Work Detail Files",
                    [
                        py,
                        str(copy_script),
                        "--mode",
                        "work_details",
                        "--ids-file",
                        str(draft_detail_uids_file),
                        "--copied-ids-file",
                        str(copied_detail_uids_file),
                    ]
                    + ([] if args.dry_run else ["--write"]),
                    cwd=repo_root,
                )

                copied_detail_uids = read_ids(copied_detail_uids_file)
                if copied_detail_uids:
                    # Restrict run to copied detail_uids and track success IDs.
                    env = os.environ.copy()
                    env[SRCSET_SELECTED_IDS_ENV_NAME] = str(copied_detail_uids_file)
                    env[SRCSET_SUCCESS_IDS_ENV_NAME] = str(detail_success_ids_file)

                    run_step(
                        "Generate Work Detail Srcset Derivatives",
                        [
                            "bash",
                            str(make_script),
                            str(args.detail_input_dir),
                            str(args.detail_output_dir),
                            str(args.jobs),
                        ]
                        + (["--dry-run"] if args.dry_run else []),
                        cwd=repo_root,
                        env=env,
                    )

                    if args.dry_run:
                        generated_detail_uids = copied_detail_uids
                    else:
                        generated_detail_uids = read_ids(detail_success_ids_file)
                else:
                    print("\n==> Skip Work Detail Srcset\nNo copied detail files in this run.")
            else:
                print("\n==> Skip Work Detail Copy/Srcset\nNo work-detail media candidates in this run.")
        else:
            print("\n==> Skip Work Details\nMode not selected.")
            generated_detail_uids = set()

        copied_moment_ids = read_ids(copied_moment_ids_file)
        generated_moment_ids = set()
        if "moment" in selected_modes:
            if draft_moment_ids:
                run_step(
                    "Copy Draft Moment Files",
                    [
                        py,
                        str(copy_script),
                        "--mode",
                        "moment",
                        "--ids-file",
                        str(draft_moment_ids_file),
                        "--copied-ids-file",
                        str(copied_moment_ids_file),
                    ]
                    + ([] if args.dry_run else ["--write"]),
                    cwd=repo_root,
                )

                copied_moment_ids = read_ids(copied_moment_ids_file)
                if copied_moment_ids:
                    # Restrict run to copied moment_ids and track success IDs.
                    env = os.environ.copy()
                    env[SRCSET_SELECTED_IDS_ENV_NAME] = str(copied_moment_ids_file)
                    env[SRCSET_SUCCESS_IDS_ENV_NAME] = str(moment_success_ids_file)

                    run_step(
                        "Generate Moment Srcset Derivatives",
                        [
                            "bash",
                            str(make_script),
                            str(args.moment_input_dir),
                            str(args.moment_output_dir),
                            str(args.jobs),
                        ]
                        + (["--dry-run"] if args.dry_run else []),
                        cwd=repo_root,
                        env=env,
                    )

                    if args.dry_run:
                        generated_moment_ids = copied_moment_ids
                    else:
                        generated_moment_ids = read_ids(moment_success_ids_file)
                else:
                    print("\n==> Skip Moment Srcset\nNo copied moment files in this run.")
            else:
                print("\n==> Skip Moment Copy/Srcset\nNo moment-media candidates in this run.")
        else:
            print("\n==> Skip Moments\nMode not selected.")
            generated_moment_ids = set()

        generate_ids = set(planned_generate_work_ids)
        generate_ids.update(generated_work_ids)
        generate_ids.update(work_ids_from_detail_uids(generated_detail_uids))

        generate_ids_file.write_text(
            "\n".join(sorted(generate_ids)) + ("\n" if generate_ids else ""),
            encoding="utf-8",
        )
        generate_moment_ids = set(planned_generate_moment_ids)
        generate_moment_ids_file.write_text(
            "\n".join(sorted(generate_moment_ids)) + ("\n" if generate_moment_ids else ""),
            encoding="utf-8",
        )

        selected_series_for_generate = set(planned_series_ids)
        cleanup_stats = {
            "generated_files_deleted": 0,
            "local_media_files_deleted": 0,
            "tag_assignment_series_removed": 0,
            "tag_assignment_work_overrides_removed": 0,
            "tag_assignment_series_touched": 0,
            "tag_assignments_written": 0,
        }
        if generate_ids or generate_moment_ids or selected_series_for_generate or stale_cleanup_requested:
            generate_cmd = [py, str(generate_script), str(xlsx_path)]
            generate_cmd += ["--work-ids-file", str(generate_ids_file)]
            generate_cmd += ["--moment-ids-file", str(generate_moment_ids_file)]
            if selected_series_for_generate:
                generate_cmd += ["--series-ids", ",".join(sorted(selected_series_for_generate))]
            elif stale_cleanup_requested and not generate_ids and not generate_moment_ids:
                generate_cmd += ["--only", "series-index-json", "--only", "works-index-json", "--only", "moments-index-json"]
            if not args.dry_run:
                generate_cmd.append("--write")
            if args.force_generate:
                generate_cmd.append("--force")
            run_step("Generate Work Pages", generate_cmd, cwd=repo_root)

            if stale_cleanup_requested:
                print("\n==> Cleanup Stale Generated Artifacts")
                existing_cleanup_paths = stale_cleanup_plan["existing"]
                missing_cleanup_paths = stale_cleanup_plan["missing"]
                print(f"Generated files to delete: {len(existing_cleanup_paths)}")
                for path in existing_cleanup_paths:
                    print(f"  - delete {path.relative_to(repo_root)}")
                for path in missing_cleanup_paths:
                    print(f"  - missing {path.relative_to(repo_root)}")
                if local_media_cleanup_plan.get("available"):
                    existing_local_media_paths = local_media_cleanup_plan["existing"]
                    print(f"Local media files to delete: {len(existing_local_media_paths)}")
                    for path in existing_local_media_paths:
                        try:
                            rel_path = path.relative_to(media_base_dir) if media_base_dir is not None else path
                        except ValueError:
                            rel_path = path
                        print(f"  - delete [media] {rel_path}")
                else:
                    existing_local_media_paths = []
                cleanup_tag_stats = prune_tag_assignments_for_removed_rows(
                    repo_root,
                    removed_series_ids=removed_series_ids,
                    removed_work_ids=removed_work_ids,
                    write=not args.dry_run,
                )
                print(
                    "Tag assignments cleanup: "
                    f"series removed={cleanup_tag_stats['series_removed']}; "
                    f"work overrides removed={cleanup_tag_stats['work_overrides_removed']}; "
                    f"series touched={cleanup_tag_stats['series_touched']}"
                )
                if args.dry_run:
                    print("DRY-RUN: stale artifact cleanup not written.")
                else:
                    cleanup_stats["generated_files_deleted"] = delete_stale_artifact_paths(existing_cleanup_paths)
                    cleanup_stats["local_media_files_deleted"] = delete_stale_artifact_paths(existing_local_media_paths)
                    cleanup_stats["tag_assignment_series_removed"] = cleanup_tag_stats["series_removed"]
                    cleanup_stats["tag_assignment_work_overrides_removed"] = cleanup_tag_stats["work_overrides_removed"]
                    cleanup_stats["tag_assignment_series_touched"] = cleanup_tag_stats["series_touched"]
                    cleanup_stats["tag_assignments_written"] = cleanup_tag_stats["payload_written"]

            search_cmd = [str(search_script), "--scope", "catalogue"]
            if not args.dry_run:
                search_cmd.append("--write")
            if args.force_generate:
                search_cmd.append("--force")
            run_step("Build Catalogue Search Index", search_cmd, cwd=repo_root)
        else:
            print("\n==> Skip Generate/Search\nPlanner found no affected targets in this run.")

    print("\nPipeline complete.")
    if not args.dry_run:
        try:
            append_build_activity(
                repo_root,
                build_activity_entry(
                    time_utc=utc_timestamp_now(),
                    status="completed",
                    dry_run=False,
                    planner_mode=planner_mode,
                    workbook_changes={
                        "works": changed_work_row_ids,
                        "series": changed_series_row_ids,
                        "work_details": changed_detail_uids,
                        "moments": changed_moment_ids,
                    },
                    media_changes={
                        "work": changed_work_media_ids,
                        "work_details": changed_detail_media_uids,
                        "moment": changed_moment_media_ids,
                    },
                    prose_changes={
                        "work": changed_work_prose_ids,
                        "series": changed_series_prose_ids,
                        "moment": changed_moment_prose_ids,
                    },
                    actions={
                        "copy_work": len(draft_ids),
                        "copy_work_details": len(draft_detail_uids),
                        "copy_moment": len(draft_moment_ids),
                        "generate_work_ids": len(generate_ids),
                        "generate_series_ids": len(selected_series_for_generate),
                        "generate_moment_ids": len(generate_moment_ids),
                        "delete_generated_files": cleanup_stats["generated_files_deleted"],
                        "delete_local_media_files": cleanup_stats["local_media_files_deleted"],
                        "clean_tag_assignment_series": cleanup_stats["tag_assignment_series_removed"],
                        "clean_tag_assignment_work_overrides": cleanup_stats["tag_assignment_work_overrides_removed"],
                        "rebuild_search": bool(generate_ids or generate_moment_ids or selected_series_for_generate or stale_cleanup_requested),
                    },
                    results={
                        "planner_state_updated": True,
                        "media_tracking_initialized": True,
                        "prose_tracking_initialized": True,
                        "planner_state_migrated": legacy_state_loaded,
                        "stale_cleanup_requested": stale_cleanup_requested,
                        "local_media_cleanup_available": bool(local_media_cleanup_plan.get("available")),
                        "tag_assignments_written": bool(cleanup_stats["tag_assignments_written"]),
                    },
                ),
            )
        except Exception:
            pass
        write_build_state(state_path, current_state)
    log_event(
        repo_root,
        "pipeline_complete",
        {
            "dry_run": args.dry_run,
            "modes": sorted(selected_modes),
            "generated_work_ids": len(generate_ids),
            "generated_moment_ids": len(generate_moment_ids),
            "stale_cleanup_requested": stale_cleanup_requested,
            "deleted_generated_files": cleanup_stats["generated_files_deleted"],
            "deleted_local_media_files": cleanup_stats["local_media_files_deleted"],
            "state_updated": not args.dry_run,
        },
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except subprocess.CalledProcessError as e:
        log_event(
            REPO_ROOT,
            "pipeline_failed",
            {
                "kind": "called_process_error",
                "return_code": e.returncode,
                "cmd": list(e.cmd),
            },
        )
        print(
            "\nPipeline failed at command: "
            + format_display_command(
                [str(part) for part in e.cmd],
                repo_root=REPO_ROOT,
                projects_base_dir=DEFAULT_PROJECTS_BASE_DIR,
                media_base_dir=DEFAULT_MEDIA_BASE_DIR,
            ),
            file=sys.stderr,
        )
        raise SystemExit(e.returncode)
    except Exception as e:  # noqa: BLE001
        log_event(
            REPO_ROOT,
            "pipeline_failed",
            {
                "kind": "exception",
                "error": str(e),
            },
        )
        raise
