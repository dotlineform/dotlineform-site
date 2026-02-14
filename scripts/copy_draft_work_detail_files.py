#!/usr/bin/env python3
"""
Copy and rename work detail files based on data/works.xlsx (WorkDetails + Works).

Source:  [PROJECTS_BASE_DIR]/projects/[project_folder]/[project_subfolder]/[project_filename]
Target:  [WORKS_BASE_DIR]/work_details/make_srcset_images/[work_id]-[detail_id][.ext]

Use --detail-uids-file to limit processing to selected detail_uids (one per line).

Usage:
  python3 scripts/copy_draft_work_detail_files.py --write

Defaults to dry-run. Pass --write to actually copy files.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

import openpyxl

# ---------
# CONFIG
# ---------
PROJECTS_BASE_DIR = Path("/Users/dlf/Library/CloudStorage/OneDrive-Personal/dotlineform")
WORKS_BASE_DIR = Path("/Users/dlf/Library/Mobile Documents/com~apple~CloudDocs/dotlineform")

WORKBOOK_PATH = Path("data/works.xlsx")
WORKS_SHEET = "Works"
DETAILS_SHEET = "WorkDetails"

DEST_RELATIVE = Path("work_details/make_srcset_images")


def normalize_id(value: Any, width: int) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.zfill(width) if s.isdigit() else s


def work_id(value: Any) -> str:
    return normalize_id(value, 5)


def detail_id(value: Any) -> str:
    return normalize_id(value, 3)


def detail_uid(raw_work_id: Any, raw_detail_id: Any) -> str:
    return f"{work_id(raw_work_id)}-{detail_id(raw_detail_id)}"


def header_map(ws) -> Dict[str, int]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return {str(h).strip(): i for i, h in enumerate(headers) if h is not None}


def read_selected_ids(path: str) -> Optional[Set[str]]:
    if not path:
        return None
    ids_path = Path(path).expanduser()
    if not ids_path.exists():
        raise SystemExit(f"detail_uids file not found: {ids_path}")
    out: Set[str] = set()
    for line in ids_path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s:
            out.add(s)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Copy work detail files from works.xlsx")
    parser.add_argument("--write", action="store_true", help="Actually copy files (default: dry-run)")
    parser.add_argument("--keep-ext", action="store_true", help="Keep source file extension on target (default)")
    parser.add_argument("--no-ext", action="store_true", help="Strip extension on target filename")
    parser.add_argument(
        "--detail-uids-file",
        default="",
        help="Optional path to detail_uids file (one uid per line). Only those rows are processed.",
    )
    parser.add_argument(
        "--copied-detail-uids-file",
        default="",
        help="Optional path to write copied detail_uids (one per line).",
    )
    args = parser.parse_args()

    keep_ext = not args.no_ext
    if args.keep_ext:
        keep_ext = True

    wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=False, data_only=True)
    if WORKS_SHEET not in wb.sheetnames:
        raise SystemExit(f"Worksheet not found: {WORKS_SHEET!r}")
    if DETAILS_SHEET not in wb.sheetnames:
        raise SystemExit(f"Worksheet not found: {DETAILS_SHEET!r}")

    works_ws = wb[WORKS_SHEET]
    details_ws = wb[DETAILS_SHEET]
    works_cols = header_map(works_ws)
    details_cols = header_map(details_ws)

    missing_works = [c for c in ["work_id", "project_folder"] if c not in works_cols]
    if missing_works:
        raise SystemExit(f"Missing required columns in Works sheet: {', '.join(missing_works)}")

    missing_details = [c for c in ["work_id", "detail_id", "project_subfolder", "project_filename"] if c not in details_cols]
    if missing_details:
        raise SystemExit(f"Missing required columns in WorkDetails sheet: {', '.join(missing_details)}")

    selected_uids = read_selected_ids(args.detail_uids_file)

    work_folder_by_id: Dict[str, str] = {}
    for row_cells in works_ws.iter_rows(min_row=2):
        row = [cell.value for cell in row_cells]
        wid = work_id(row[works_cols["work_id"]])
        folder = row[works_cols["project_folder"]]
        if wid and folder:
            work_folder_by_id[wid] = str(folder).strip()

    dest_dir = WORKS_BASE_DIR / DEST_RELATIVE
    if args.write:
        dest_dir.mkdir(parents=True, exist_ok=True)

    total = 0
    copied = 0
    copied_uids: List[str] = []
    missing_src = 0
    missing_work_folder = 0

    for row_cells in details_ws.iter_rows(min_row=2):
        row = [cell.value for cell in row_cells]
        raw_work_id = row[details_cols["work_id"]]
        raw_detail_id = row[details_cols["detail_id"]]
        subfolder = row[details_cols["project_subfolder"]]
        filename = row[details_cols["project_filename"]]

        if not (raw_work_id and raw_detail_id and filename is not None):
            continue

        wid = work_id(raw_work_id)
        did = detail_id(raw_detail_id)
        uid = f"{wid}-{did}"

        if selected_uids is not None and uid not in selected_uids:
            continue

        project_folder = work_folder_by_id.get(wid)
        if not project_folder:
            print(f"Warning: missing Works.project_folder for work_id={wid}; skipping detail {uid}")
            missing_work_folder += 1
            continue

        subfolder_str = str(subfolder).strip() if subfolder else ""
        filename_str = str(filename).strip()
        src = PROJECTS_BASE_DIR / "projects" / project_folder
        if subfolder_str:
            src = src / subfolder_str
        src = src / filename_str

        if keep_ext:
            ext = Path(filename_str).suffix
            dest = dest_dir / f"{uid}{ext}"
        else:
            dest = dest_dir / uid

        total += 1

        if not src.exists():
            print(f"Missing source: {src}")
            missing_src += 1
            continue

        if args.write:
            dest.parent.mkdir(parents=True, exist_ok=True)
            from shutil import copy2

            copy2(src, dest)
            print(f"Copied: {src} -> {dest}")
            copied += 1
            copied_uids.append(uid)
        else:
            print(f"Dry-run: {src} -> {dest}")
            copied_uids.append(uid)

    if args.copied_detail_uids_file:
        ids_path = Path(args.copied_detail_uids_file).expanduser()
        ids_path.parent.mkdir(parents=True, exist_ok=True)
        ids_path.write_text("\n".join(copied_uids) + ("\n" if copied_uids else ""), encoding="utf-8")
        print(f"Wrote copied detail IDs manifest: {ids_path} ({len(copied_uids)} ids)")

    print(
        "Detail rows: "
        f"{total}, copied: {copied}, missing source: {missing_src}, missing work folder: {missing_work_folder}"
    )
    if not args.write:
        print("Dry-run only. Re-run with --write to copy files.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
