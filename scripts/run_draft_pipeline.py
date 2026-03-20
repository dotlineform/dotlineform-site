#!/usr/bin/env python3
"""
Run the draft publish pipeline end-to-end (fail-fast):

1) copy_draft_media_files.py --mode work --write
2) make_srcset_images.sh
3) copy_draft_media_files.py --mode work_details --write
4) make_srcset_images.sh (for work details)
5) copy_draft_media_files.py --mode moment --write
6) make_srcset_images.sh (for moments)
7) generate_work_pages.py data/works.xlsx --write (for affected work IDs only)

Dry-run mode:
- copy + srcset + generate run in preview mode
- no workbook writes/deletes are performed

Flag usage summary:
- --xlsx: workbook path (repo-relative by default)
- --dry-run: preview-only mode for all steps
- --force-generate: pass --force to generate_work_pages.py
- --jobs: parallelism for make_srcset_images.sh
- --*-input-dir / --*-output-dir: source/derivative directories for each media type
- --mode: select one or more flows: work, work_details, moment (repeatable)
- --work-ids / --work-ids-file: limit work + work_details processing scope
- --series-ids / --series-ids-file: pass series filter into generation
- --moment-ids / --moment-ids-file: limit moment processing scope

Required local environment (when path flags are not passed):
- env vars are resolved from `_data/pipeline.json`
- defaults remain `DOTLINEFORM_MEDIA_BASE_DIR` and `DOTLINEFORM_PROJECTS_BASE_DIR`

Behavior note:
- When mode includes `work` and no explicit `--series-ids*` are provided,
  draft series IDs are auto-included for generation.

Manifest files (created in a temporary directory):
- draft_*: candidates from workbook (status=draft)
- copied_*: successfully copied source IDs (input to srcset step)
- *_success_ids: srcset-success IDs returned by make_srcset_images.sh
- generate_ids.txt: work IDs to regenerate (works + any work_details parents)
"""

from __future__ import annotations

import argparse
import os
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, Iterable, Set

import openpyxl

try:
    from script_logging import append_script_log
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.script_logging import append_script_log

try:
    from pipeline_config import (
        env_var_name,
        env_var_value,
        join_base_and_subdir,
        load_pipeline_config,
        media_mode_input_subdir,
        media_mode_output_subdir,
    )
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.pipeline_config import (
        env_var_name,
        env_var_value,
        join_base_and_subdir,
        load_pipeline_config,
        media_mode_input_subdir,
        media_mode_output_subdir,
    )


PIPELINE_CONFIG = load_pipeline_config(Path(__file__))
MEDIA_BASE_DIR_ENV_NAME = env_var_name(PIPELINE_CONFIG, "media_base_dir")
SRCSET_JOBS_ENV_NAME = env_var_name(PIPELINE_CONFIG, "srcset_jobs")
SRCSET_SELECTED_IDS_ENV_NAME = env_var_name(PIPELINE_CONFIG, "srcset_selected_ids_file")
SRCSET_SUCCESS_IDS_ENV_NAME = env_var_name(PIPELINE_CONFIG, "srcset_success_ids_file")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("'") and len(s) > 1:
        s = s[1:]
    return s


def normalize_status(value: Any) -> str:
    return normalize_text(value).lower()


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def slug_id(raw: Any, width: int = 5) -> str:
    if raw is None:
        raise ValueError("Missing id")
    s = normalize_text(raw)
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"\D", "", s)
    if not s:
        raise ValueError(f"Invalid id value: {raw!r}")
    return s.zfill(width)


def header_map(ws) -> Dict[str, int]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    hi: Dict[str, int] = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).strip()
        hi[key] = i
        hi[key.lower()] = i
    return hi


def first_present_col(header_index: Dict[str, int], names: list[str]) -> str | None:
    for name in names:
        if name in header_index:
            return name
    return None


def parse_series_ids_from_works_row(row: tuple[Any, ...], hi: Dict[str, int]) -> list[str]:
    parsed: list[str] = []
    series_ids_col = first_present_col(hi, ["series_ids"])
    if series_ids_col is not None:
        raw_value = row[hi[series_ids_col]]
        parsed = [normalize_text(part) for part in normalize_text(raw_value).split(",") if normalize_text(part)]

    out: list[str] = []
    seen: set[str] = set()
    for sid in parsed:
        if not sid or sid in seen:
            continue
        seen.add(sid)
        out.append(sid)
    return out


def run_step(label: str, cmd: list[str], cwd: Path, env: Dict[str, str] | None = None) -> None:
    """Run one pipeline step and fail fast on non-zero exit."""
    print(f"\n==> {label}")
    print("+", " ".join(cmd))
    subprocess.run(cmd, cwd=str(cwd), env=env, check=True)


def log_event(repo_root: Path, event: str, details: Dict[str, Any] | None = None) -> None:
    try:
        append_script_log(Path(__file__), event=event, details=details or {}, repo_root=repo_root)
    except Exception:
        # Logging failures must not stop pipeline execution.
        pass


def read_ids(path: Path) -> Set[str]:
    """Read newline-delimited IDs from a manifest file."""
    if not path.exists():
        return set()
    out: Set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s:
            out.add(s)
    return out


def read_optional_ids_file(path: str) -> Set[str]:
    if not path:
        return set()
    ids_path = Path(path).expanduser()
    if not ids_path.exists():
        raise SystemExit(f"IDs file not found: {ids_path}")
    return read_ids(ids_path)


def parse_work_id_selection(raw: str) -> Set[str]:
    """
    Parse comma-separated work-id selectors supporting individual IDs and ranges.
    Examples:
      "66,74" -> {"00066", "00074"}
      "66-74,38-40,12" -> {"00012", "00038", ..., "00074"}
    """
    selected: Set[str] = set()
    for token in (part.strip() for part in str(raw).split(",") if part.strip()):
        m = re.match(r"^(\d+)\s*-\s*(\d+)$", token)
        if m:
            start = int(m.group(1))
            end = int(m.group(2))
            if start > end:
                start, end = end, start
            for n in range(start, end + 1):
                selected.add(slug_id(n))
        else:
            selected.add(slug_id(token))
    return selected


def collect_draft_ids(
    xlsx_path: Path,
    allowed_ids: Set[str] | None = None,
    include_published: bool = False,
) -> Set[str]:
    """Collect draft work_ids from Works.status."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Works" not in wb.sheetnames:
        raise SystemExit("Worksheet not found: 'Works'")
    ws = wb["Works"]
    hi = header_map(ws)

    required = ["work_id", "status"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"Works sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_work_id = row[hi["work_id"]]
        if is_empty(raw_work_id):
            continue
        wid = slug_id(raw_work_id)
        if allowed_ids is not None and wid not in allowed_ids:
            continue
        status = normalize_status(row[hi["status"]])
        if status != "draft" and not (include_published and status == "published"):
            continue
        out.add(wid)
    return out


def collect_draft_detail_uids(
    xlsx_path: Path,
    allowed_uids: Set[str] | None = None,
    allowed_work_ids: Set[str] | None = None,
    include_published: bool = False,
) -> Set[str]:
    """Collect draft detail_uids from WorkDetails.status."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "WorkDetails" not in wb.sheetnames:
        return set()
    ws = wb["WorkDetails"]
    hi = header_map(ws)

    required = ["work_id", "detail_id", "status"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"WorkDetails sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_work_id = row[hi["work_id"]]
        raw_detail_id = row[hi["detail_id"]]
        if is_empty(raw_work_id) or is_empty(raw_detail_id):
            continue
        wid = slug_id(raw_work_id)
        did = slug_id(raw_detail_id, width=3)
        uid = f"{wid}-{did}"
        if allowed_uids is not None and uid not in allowed_uids:
            continue
        if allowed_work_ids is not None and wid not in allowed_work_ids:
            continue
        status = normalize_status(row[hi["status"]])
        if status != "draft" and not (include_published and status == "published"):
            continue
        out.add(uid)
    return out


def collect_draft_moment_ids(
    xlsx_path: Path,
    allowed_ids: Set[str] | None = None,
    include_published: bool = False,
) -> Set[str]:
    """Collect draft moment_ids from Moments.status."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Moments" not in wb.sheetnames:
        return set()
    ws = wb["Moments"]
    hi = header_map(ws)

    required = ["moment_id", "status"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"Moments sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_moment_id = row[hi["moment_id"]]
        if is_empty(raw_moment_id):
            continue
        mid = normalize_text(raw_moment_id).lower()
        if not mid:
            continue
        if allowed_ids is not None and mid not in allowed_ids:
            continue
        status = normalize_status(row[hi["status"]])
        if status != "draft" and not (include_published and status == "published"):
            continue
        out.add(mid)
    return out


def collect_draft_series_ids(xlsx_path: Path, allowed_ids: Set[str] | None = None) -> Set[str]:
    """Collect draft series_ids from Series.status."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Series" not in wb.sheetnames:
        return set()
    ws = wb["Series"]
    hi = header_map(ws)

    required = ["series_id", "status"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"Series sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_series_id = row[hi["series_id"]]
        if is_empty(raw_series_id):
            continue
        sid = normalize_text(raw_series_id)
        if not sid:
            continue
        if allowed_ids is not None and sid not in allowed_ids:
            continue
        status = normalize_status(row[hi["status"]])
        if status != "draft":
            continue
        out.add(sid)
    return out


def collect_series_ids_for_work_ids(xlsx_path: Path, work_ids: Set[str]) -> Set[str]:
    """Map work_ids to series_ids using Works sheet rows."""
    if not work_ids:
        return set()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Works" not in wb.sheetnames:
        raise SystemExit("Worksheet not found: 'Works'")
    ws = wb["Works"]
    hi = header_map(ws)

    required = ["work_id", "series_ids"]
    missing = [c for c in required if c not in hi]
    if missing:
        raise SystemExit(f"Works sheet missing required columns: {', '.join(missing)}")

    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_work_id = row[hi["work_id"]]
        if is_empty(raw_work_id):
            continue
        wid = slug_id(raw_work_id)
        if wid not in work_ids:
            continue
        for sid in parse_series_ids_from_works_row(row, hi):
            out.add(sid)
    return out


def work_ids_from_detail_uids(detail_uids: Iterable[str]) -> Set[str]:
    """Map detail_uids (work_id-detail_id) to parent work_ids."""
    out: Set[str] = set()
    for uid in detail_uids:
        s = str(uid).strip()
        if not s or "-" not in s:
            continue
        out.add(s.split("-", 1)[0])
    return out


def main() -> int:
    media_base = env_var_value(PIPELINE_CONFIG, "media_base_dir")
    default_work_input = join_base_and_subdir(media_base, media_mode_input_subdir(PIPELINE_CONFIG, "work"))
    default_work_output = join_base_and_subdir(media_base, media_mode_output_subdir(PIPELINE_CONFIG, "work"))
    default_detail_input = join_base_and_subdir(media_base, media_mode_input_subdir(PIPELINE_CONFIG, "work_details"))
    default_detail_output = join_base_and_subdir(media_base, media_mode_output_subdir(PIPELINE_CONFIG, "work_details"))
    default_moment_input = join_base_and_subdir(media_base, media_mode_input_subdir(PIPELINE_CONFIG, "moment"))
    default_moment_output = join_base_and_subdir(media_base, media_mode_output_subdir(PIPELINE_CONFIG, "moment"))

    ap = argparse.ArgumentParser(description="Run copy -> make_srcset -> generate pipeline.")
    ap.add_argument("--xlsx", default="data/works.xlsx", help="Path to workbook")
    ap.add_argument(
        "--mode",
        action="append",
        choices=["work", "work_details", "moment"],
        help="Flow(s) to run. Repeat to run multiple modes. Default: all.",
    )
    ap.add_argument(
        "--input-dir",
        default=default_work_input,
        help="Input directory for source images",
    )
    ap.add_argument(
        "--output-dir",
        default=default_work_output,
        help="Output directory for srcset derivatives",
    )
    ap.add_argument(
        "--detail-input-dir",
        default=default_detail_input,
        help="Input directory for work detail source images",
    )
    ap.add_argument(
        "--detail-output-dir",
        default=default_detail_output,
        help="Output directory for work detail srcset derivatives",
    )
    ap.add_argument(
        "--moment-input-dir",
        default=default_moment_input,
        help="Input directory for moment source images",
    )
    ap.add_argument(
        "--moment-output-dir",
        default=default_moment_output,
        help="Output directory for moment srcset derivatives",
    )
    ap.add_argument(
        "--jobs",
        type=int,
        default=int(os.environ.get(SRCSET_JOBS_ENV_NAME, "4")),
        help="Parallel jobs",
    )
    ap.add_argument("--force-generate", action="store_true", help="Pass --force to generate_work_pages.py")
    ap.add_argument("--dry-run", action="store_true", help="Preview mode; no writes/deletes.")
    ap.add_argument(
        "--work-ids",
        default="",
        help="Comma-separated work_ids/ranges filter for this run (e.g. 66-74,38-40,1).",
    )
    ap.add_argument("--work-ids-file", default="", help="Path to work_ids file (one id per line).")
    ap.add_argument("--series-ids", default="", help="Comma-separated series_ids passed to generation.")
    ap.add_argument("--series-ids-file", default="", help="Path to series_ids file (one id per line).")
    ap.add_argument("--moment-ids", default="", help="Comma-separated moment_ids filter for this run.")
    ap.add_argument("--moment-ids-file", default="", help="Path to moment_ids file (one id per line).")
    args = ap.parse_args()
    selected_modes = set(args.mode or ["work", "work_details", "moment"])

    def require_non_empty(label: str, value: str, flag: str) -> None:
        if normalize_text(value) == "":
            raise SystemExit(
                f"Missing {label}. Set {MEDIA_BASE_DIR_ENV_NAME} or pass {flag}."
            )

    if "work" in selected_modes:
        require_non_empty("work input dir", args.input_dir, "--input-dir")
        require_non_empty("work output dir", args.output_dir, "--output-dir")
    if "work_details" in selected_modes:
        require_non_empty("work_details input dir", args.detail_input_dir, "--detail-input-dir")
        require_non_empty("work_details output dir", args.detail_output_dir, "--detail-output-dir")
    if "moment" in selected_modes:
        require_non_empty("moment input dir", args.moment_input_dir, "--moment-input-dir")
        require_non_empty("moment output dir", args.moment_output_dir, "--moment-output-dir")

    explicit_work_ids = read_optional_ids_file(args.work_ids_file)
    if args.work_ids:
        explicit_work_ids.update(parse_work_id_selection(args.work_ids))
    work_filter = explicit_work_ids if explicit_work_ids else None

    explicit_series_ids = read_optional_ids_file(args.series_ids_file)
    if args.series_ids:
        explicit_series_ids.update({normalize_text(s.strip()) for s in args.series_ids.split(",") if s.strip()})
    series_filter = explicit_series_ids if explicit_series_ids else None
    auto_series_filter = None

    explicit_moment_ids = read_optional_ids_file(args.moment_ids_file)
    if args.moment_ids:
        explicit_moment_ids.update({normalize_text(m.strip()).lower() for m in args.moment_ids.split(",") if m.strip()})
    moment_filter = explicit_moment_ids if explicit_moment_ids else None

    repo_root = Path(__file__).resolve().parents[1]
    xlsx_path = (repo_root / args.xlsx).resolve()
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")
    log_event(
        repo_root,
        "pipeline_start",
        {
            "dry_run": args.dry_run,
            "modes": sorted(selected_modes),
            "xlsx": str(xlsx_path.relative_to(repo_root)),
            "force_generate": args.force_generate,
        },
    )
    if series_filter is None and "work" in selected_modes:
        auto_series_filter = collect_draft_series_ids(xlsx_path=xlsx_path)

    copy_script = repo_root / "scripts/copy_draft_media_files.py"
    make_script = repo_root / "scripts/make_srcset_images.sh"
    generate_script = repo_root / "scripts/generate_work_pages.py"
    py = sys.executable

    with tempfile.TemporaryDirectory(prefix="draft-pipeline-") as tmp:
        tmp_dir = Path(tmp)
        draft_ids_file = tmp_dir / "draft_ids.txt"
        copied_ids_file = tmp_dir / "copied_ids.txt"
        draft_detail_uids_file = tmp_dir / "draft_detail_uids.txt"
        copied_detail_uids_file = tmp_dir / "copied_detail_uids.txt"
        success_ids_file = tmp_dir / "success_ids.txt"
        detail_success_ids_file = tmp_dir / "detail_success_ids.txt"
        moment_success_ids_file = tmp_dir / "moment_success_ids.txt"
        draft_moment_ids_file = tmp_dir / "draft_moment_ids.txt"
        copied_moment_ids_file = tmp_dir / "copied_moment_ids.txt"
        generate_moment_ids_file = tmp_dir / "generate_moment_ids.txt"
        generate_ids_file = tmp_dir / "generate_ids.txt"

        draft_ids = set()
        if "work" in selected_modes:
            draft_ids = collect_draft_ids(
                xlsx_path,
                allowed_ids=work_filter,
                include_published=bool(work_filter is not None),
            )
        draft_ids_file.write_text(
            "\n".join(sorted(draft_ids)) + ("\n" if draft_ids else ""),
            encoding="utf-8",
        )
        print(f"Draft candidates (work): {len(draft_ids)}")

        draft_detail_uids = set()
        if "work_details" in selected_modes:
            draft_detail_uids = collect_draft_detail_uids(
                xlsx_path,
                allowed_uids=None,
                allowed_work_ids=work_filter,
                include_published=bool(work_filter is not None),
            )
        draft_detail_uids_file.write_text(
            "\n".join(sorted(draft_detail_uids)) + ("\n" if draft_detail_uids else ""),
            encoding="utf-8",
        )
        print(f"Draft detail candidates (work_details): {len(draft_detail_uids)}")

        draft_moment_ids = set()
        if "moment" in selected_modes:
            draft_moment_ids = collect_draft_moment_ids(
                xlsx_path,
                allowed_ids=moment_filter,
                include_published=bool(moment_filter is not None),
            )
        draft_moment_ids_file.write_text(
            "\n".join(sorted(draft_moment_ids)) + ("\n" if draft_moment_ids else ""),
            encoding="utf-8",
        )
        print(f"Draft moment candidates (moment): {len(draft_moment_ids)}")

        copied_ids = read_ids(copied_ids_file)
        generated_work_ids = set()
        if "work" in selected_modes:
            run_step(
                "Copy Draft Work Files",
                [
                    py,
                    str(copy_script),
                    "--mode",
                    "work",
                    "--ids-file",
                    str(draft_ids_file),
                    "--copied-ids-file",
                    str(copied_ids_file),
                ]
                + ([] if args.dry_run else ["--write"]),
                cwd=repo_root,
            )

            copied_ids = read_ids(copied_ids_file)
            if copied_ids:
                # make_srcset_images.sh controls selection via environment variables.
                env = os.environ.copy()
                env[SRCSET_SELECTED_IDS_ENV_NAME] = str(copied_ids_file)
                env[SRCSET_SUCCESS_IDS_ENV_NAME] = str(success_ids_file)

                run_step(
                    "Generate Srcset Derivatives",
                    [
                        "bash",
                        str(make_script),
                        str(args.input_dir),
                        str(args.output_dir),
                        str(args.jobs),
                    ]
                    + (["--dry-run"] if args.dry_run else []),
                    cwd=repo_root,
                    env=env,
                )

                if args.dry_run:
                    generated_work_ids = copied_ids
                else:
                    generated_work_ids = read_ids(success_ids_file)
            else:
                print("\n==> Skip Srcset\nNo copied draft files in this run.")
        else:
            print("\n==> Skip Work\nMode not selected.")
            generated_work_ids = set()

        copied_detail_uids = read_ids(copied_detail_uids_file)
        generated_detail_uids = set()
        if "work_details" in selected_modes:
            run_step(
                "Copy Draft Work Detail Files",
                [
                    py,
                    str(copy_script),
                    "--mode",
                    "work_details",
                    "--ids-file",
                    str(draft_detail_uids_file),
                    "--copied-ids-file",
                    str(copied_detail_uids_file),
                ]
                + ([] if args.dry_run else ["--write"]),
                cwd=repo_root,
            )

            copied_detail_uids = read_ids(copied_detail_uids_file)
            if copied_detail_uids:
                # Restrict run to copied detail_uids and track success IDs.
                env = os.environ.copy()
                env[SRCSET_SELECTED_IDS_ENV_NAME] = str(copied_detail_uids_file)
                env[SRCSET_SUCCESS_IDS_ENV_NAME] = str(detail_success_ids_file)

                run_step(
                    "Generate Work Detail Srcset Derivatives",
                    [
                        "bash",
                        str(make_script),
                        str(args.detail_input_dir),
                        str(args.detail_output_dir),
                        str(args.jobs),
                    ]
                    + (["--dry-run"] if args.dry_run else []),
                    cwd=repo_root,
                    env=env,
                )

                if args.dry_run:
                    generated_detail_uids = copied_detail_uids
                else:
                    generated_detail_uids = read_ids(detail_success_ids_file)
            else:
                print("\n==> Skip Work Detail Srcset\nNo copied detail files in this run.")
        else:
            print("\n==> Skip Work Details\nMode not selected.")
            generated_detail_uids = set()

        copied_moment_ids = read_ids(copied_moment_ids_file)
        generated_moment_ids = set()
        if "moment" in selected_modes:
            run_step(
                "Copy Draft Moment Files",
                [
                    py,
                    str(copy_script),
                    "--mode",
                    "moment",
                    "--ids-file",
                    str(draft_moment_ids_file),
                    "--copied-ids-file",
                    str(copied_moment_ids_file),
                ]
                + ([] if args.dry_run else ["--write"]),
                cwd=repo_root,
            )

            copied_moment_ids = read_ids(copied_moment_ids_file)
            if copied_moment_ids:
                # Restrict run to copied moment_ids and track success IDs.
                env = os.environ.copy()
                env[SRCSET_SELECTED_IDS_ENV_NAME] = str(copied_moment_ids_file)
                env[SRCSET_SUCCESS_IDS_ENV_NAME] = str(moment_success_ids_file)

                run_step(
                    "Generate Moment Srcset Derivatives",
                    [
                        "bash",
                        str(make_script),
                        str(args.moment_input_dir),
                        str(args.moment_output_dir),
                        str(args.jobs),
                    ]
                    + (["--dry-run"] if args.dry_run else []),
                    cwd=repo_root,
                    env=env,
                )

                if args.dry_run:
                    generated_moment_ids = copied_moment_ids
                else:
                    generated_moment_ids = read_ids(moment_success_ids_file)
            else:
                print("\n==> Skip Moment Srcset\nNo copied moment files in this run.")
        else:
            print("\n==> Skip Moments\nMode not selected.")
            generated_moment_ids = set()

        generate_ids = set(generated_work_ids)
        generate_ids.update(work_ids_from_detail_uids(generated_detail_uids))

        generate_ids_file.write_text(
            "\n".join(sorted(generate_ids)) + ("\n" if generate_ids else ""),
            encoding="utf-8",
        )
        # Moment generation should be driven by selected draft moments, not copied image IDs.
        # This ensures moments without images are still generated.
        generate_moment_ids_file.write_text(
            "\n".join(sorted(draft_moment_ids)) + ("\n" if draft_moment_ids else ""),
            encoding="utf-8",
        )

        generate_cmd = [py, str(generate_script), str(xlsx_path)]
        generate_cmd += ["--work-ids-file", str(generate_ids_file)]
        generate_cmd += ["--moment-ids-file", str(generate_moment_ids_file)]
        selected_series_for_generate = series_filter
        if selected_series_for_generate is None:
            selected_series_for_generate = set(auto_series_filter or set())
            selected_series_for_generate.update(
                collect_series_ids_for_work_ids(xlsx_path=xlsx_path, work_ids=generate_ids)
            )
        if selected_series_for_generate:
            generate_cmd += ["--series-ids", ",".join(sorted(selected_series_for_generate))]
        if not args.dry_run:
            generate_cmd.append("--write")
        if args.force_generate:
            generate_cmd.append("--force")
        run_step("Generate Work Pages", generate_cmd, cwd=repo_root)

    print("\nPipeline complete.")
    log_event(
        repo_root,
        "pipeline_complete",
        {
            "dry_run": args.dry_run,
            "modes": sorted(selected_modes),
            "generated_work_ids": len(generate_ids),
            "generated_moment_ids": len(draft_moment_ids),
        },
    )
    return 0


if __name__ == "__main__":
    repo_root = Path(__file__).resolve().parents[1]
    try:
        raise SystemExit(main())
    except subprocess.CalledProcessError as e:
        log_event(
            repo_root,
            "pipeline_failed",
            {
                "kind": "called_process_error",
                "return_code": e.returncode,
                "cmd": list(e.cmd),
            },
        )
        print(f"\nPipeline failed at command: {' '.join(e.cmd)}", file=sys.stderr)
        raise SystemExit(e.returncode)
    except Exception as e:  # noqa: BLE001
        log_event(
            repo_root,
            "pipeline_failed",
            {
                "kind": "exception",
                "error": str(e),
            },
        )
        raise
