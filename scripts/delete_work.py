#!/usr/bin/env python3
"""
Delete a single work from generated repo artifacts when Works.status == delete.

Safe by default:
- dry-run unless you pass --write
- accepts exactly one --work-id
- only proceeds when the matching Works row has status 'delete'
- sets Works.status to 'deleted' after successful writes

Deletion scope:
- _works/<work_id>.md
- _work_details/<work_id>-*.md
- assets/works/index/<work_id>.json
- assets/data/series_index.json
- assets/data/works_index.json
- assets/studio/data/tag_assignments.json (per-work overrides only)

Intentionally left untouched:
- assets/work_details/img/*
- canonical work prose under the external projects prose folder
- staged media under $DOTLINEFORM_MEDIA_BASE_DIR/works/files/<work_id>-*
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

import openpyxl

try:
    from script_logging import append_script_log
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.script_logging import append_script_log


WORK_ID_INPUT_RE = re.compile(r"^\d{1,5}$")
WORK_ID_RE = re.compile(r"^\d{5}$")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("'") and len(s) > 1:
        s = s[1:]
    return s


def normalize_status(value: Any) -> str:
    return normalize_text(value).lower()


def normalize_work_id(raw: str) -> str:
    value = normalize_text(raw)
    if not WORK_ID_INPUT_RE.fullmatch(value):
        raise SystemExit("--work-id must be a single numeric work_id (for example 00123)")
    return value.zfill(5)


def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def backup_stamp_now() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%d-%H%M%S")


def canonicalize_for_hash(value: Any) -> Any:
    if isinstance(value, dict):
        out: Dict[str, Any] = {}
        for key in sorted(value.keys(), key=lambda k: str(k)):
            out[str(key)] = canonicalize_for_hash(value[key])
        return out
    if isinstance(value, list):
        return [canonicalize_for_hash(item) for item in value]
    if isinstance(value, tuple):
        return [canonicalize_for_hash(item) for item in value]
    if isinstance(value, float):
        if value == 0.0:
            return 0
        if value.is_integer():
            return int(value)
        return float(f"{value:.15g}")
    return value


def compute_payload_version(payload: Any) -> str:
    canonical = json.dumps(
        canonicalize_for_hash(payload),
        sort_keys=True,
        ensure_ascii=False,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return f"blake2b-{hashlib.blake2b(canonical, digest_size=16).hexdigest()}"


def build_header_index(rows: list[tuple[Any, ...]]) -> Dict[str, int]:
    if not rows:
        return {}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    index: Dict[str, int] = {}
    for i, header in enumerate(headers):
        if not header:
            continue
        index[header] = i
        index[header.lower()] = i
    return index


def cell(row: tuple[Any, ...], header_index: Dict[str, int], col_name: str) -> Any:
    idx = header_index.get(col_name)
    return None if idx is None or idx >= len(row) else row[idx]


def find_repo_root(explicit_root: str) -> Path:
    if explicit_root:
        repo_root = Path(explicit_root).expanduser().resolve()
        if not (repo_root / "_config.yml").exists():
            raise SystemExit(f"--repo-root does not look like repo root (missing _config.yml): {repo_root}")
        return repo_root

    for start in [Path.cwd(), Path(__file__).resolve().parent]:
        current = start.resolve()
        for candidate in [current, *current.parents]:
            if (candidate / "_config.yml").exists():
                return candidate
    raise SystemExit("Could not auto-detect repo root. Pass --repo-root.")


def log_event(event: str, details: Optional[Dict[str, Any]] = None) -> None:
    try:
        append_script_log(Path(__file__), event=event, details=details or {})
    except Exception:
        pass


def load_json_object(path: Path, label: str) -> Dict[str, Any]:
    if not path.exists():
        raise SystemExit(f"Missing {label}: {path}")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise SystemExit(f"Failed to parse {label}: {path} ({exc})")
    if not isinstance(payload, dict):
        raise SystemExit(f"Invalid {label} payload (expected object): {path}")
    return payload


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def write_json_atomic(path: Path, payload: Dict[str, Any]) -> None:
    ensure_parent(path)
    fd, temp_name = tempfile.mkstemp(prefix=f"{path.name}.", suffix=".tmp", dir=str(path.parent))
    temp_path = Path(temp_name)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2, sort_keys=False)
            handle.write("\n")
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def write_workbook_atomic(path: Path, workbook: openpyxl.Workbook) -> None:
    ensure_parent(path)
    fd, temp_name = tempfile.mkstemp(prefix=f"{path.name}.", suffix=".tmp", dir=str(path.parent))
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def backup_existing_paths(paths: Iterable[Path], repo_root: Path, backup_root: Path) -> Dict[Path, Path]:
    backups: Dict[Path, Path] = {}
    for path in paths:
        if not path.exists() or path in backups:
            continue
        try:
            rel_path = path.resolve().relative_to(repo_root.resolve())
        except ValueError:
            rel_path = Path(path.name)
        backup_path = backup_root / rel_path
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, backup_path)
        backups[path] = backup_path
    return backups


def restore_backups(backups: Dict[Path, Path]) -> None:
    for path, backup_path in backups.items():
        if not backup_path.exists():
            continue
        path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(backup_path, path)


def finalize_series_index_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    series_map = payload.get("series")
    if not isinstance(series_map, dict):
        raise SystemExit("Invalid series_index.json payload: missing series map")
    schema = str((payload.get("header") or {}).get("schema") or "series_index_v2")
    payload["header"] = {
        "schema": schema,
        "version": compute_payload_version({"schema": schema, "series": series_map}),
        "generated_at_utc": utc_now(),
        "count": len(series_map),
    }
    return payload


def finalize_works_index_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    works_map = payload.get("works")
    if not isinstance(works_map, dict):
        raise SystemExit("Invalid works_index.json payload: missing works map")
    schema = str((payload.get("header") or {}).get("schema") or "works_index_v4")
    payload["header"] = {
        "schema": schema,
        "version": compute_payload_version({"schema": schema, "works": works_map}),
        "generated_at_utc": utc_now(),
        "count": len(works_map),
    }
    return payload


def remove_work_from_series_index(payload: Dict[str, Any], work_id: str) -> Dict[str, int]:
    series_map = payload.get("series")
    if not isinstance(series_map, dict):
        raise SystemExit("Invalid series_index.json payload: missing series map")

    membership_updates = 0
    primary_nulls = 0
    for row in series_map.values():
        if not isinstance(row, dict):
            continue

        works = row.get("works")
        if isinstance(works, list):
            filtered = [item for item in works if normalize_text(item) != work_id]
            if filtered != works:
                row["works"] = filtered
                membership_updates += 1

        if normalize_text(row.get("primary_work_id")) == work_id:
            row["primary_work_id"] = None
            primary_nulls += 1

    return {
        "membership_updates": membership_updates,
        "primary_nulls": primary_nulls,
    }


def remove_work_from_works_index(payload: Dict[str, Any], work_id: str) -> bool:
    works_map = payload.get("works")
    if not isinstance(works_map, dict):
        raise SystemExit("Invalid works_index.json payload: missing works map")
    if work_id not in works_map:
        return False
    del works_map[work_id]
    return True


def remove_work_overrides(payload: Dict[str, Any], work_id: str) -> Dict[str, int]:
    series_map = payload.get("series")
    if not isinstance(series_map, dict):
        raise SystemExit("Invalid tag_assignments.json payload: missing series map")

    touched_series = 0
    removed_overrides = 0
    now_utc = utc_now()

    for row in series_map.values():
        if not isinstance(row, dict):
            continue
        works = row.get("works")
        if not isinstance(works, dict) or work_id not in works:
            continue
        del works[work_id]
        removed_overrides += 1
        touched_series += 1
        row["updated_at_utc"] = now_utc
        if "works" not in row or not isinstance(row.get("works"), dict):
            row["works"] = {}

    if removed_overrides > 0:
        payload["updated_at_utc"] = now_utc
    if "tag_assignments_version" not in payload:
        payload["tag_assignments_version"] = "tag_assignments_v1"

    return {
        "touched_series": touched_series,
        "removed_overrides": removed_overrides,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--work-id", required=True, help="Single work_id to delete")
    parser.add_argument("--write", action="store_true", help="Apply changes (otherwise dry-run)")
    parser.add_argument("--repo-root", default="", help="Optional repo root override")
    parser.add_argument("--xlsx", default="data/works.xlsx", help="Workbook path relative to repo root")
    parser.add_argument("--works-sheet", default="Works", help="Worksheet name containing work metadata")
    args = parser.parse_args()

    work_id = normalize_work_id(args.work_id)
    repo_root = find_repo_root(args.repo_root)
    xlsx_path = (repo_root / args.xlsx).resolve()
    works_dir = repo_root / "_works"
    work_details_dir = repo_root / "_work_details"
    work_json_dir = repo_root / "assets/works/index"
    series_index_path = repo_root / "assets/data/series_index.json"
    works_index_path = repo_root / "assets/data/works_index.json"
    tag_assignments_path = repo_root / "assets/studio/data/tag_assignments.json"

    log_event("delete_work_start", {"argv": sys.argv[1:], "work_id": work_id, "write": bool(args.write)})

    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    workbook = openpyxl.load_workbook(xlsx_path)
    if args.works_sheet not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found in workbook: {args.works_sheet}")
    works_ws = workbook[args.works_sheet]
    works_rows = list(works_ws.iter_rows(values_only=True))
    if not works_rows:
        raise SystemExit(f"Works sheet '{args.works_sheet}' is empty")

    works_hi = build_header_index(works_rows)
    required_columns = ["work_id", "status"]
    missing_columns = [name for name in required_columns if name not in works_hi]
    if missing_columns:
        raise SystemExit(f"Works sheet missing required columns: {', '.join(missing_columns)}")

    target_row_cells = None
    target_row_number = None
    target_status_idx = works_hi["status"]

    for row, row_cells in zip(works_rows[1:], works_ws.iter_rows(min_row=2), strict=False):
        raw_work_id = normalize_text(cell(row, works_hi, "work_id"))
        if not raw_work_id:
            continue
        if normalize_work_id(raw_work_id) != work_id:
            continue
        if target_row_cells is not None:
            raise SystemExit(f"Works sheet has duplicate work_id rows for {work_id}")
        target_row_cells = row_cells
        target_row_number = row_cells[0].row

    if target_row_cells is None or target_row_number is None:
        raise SystemExit(f"Work not found in {args.works_sheet}: {work_id}")

    current_status = normalize_status(target_row_cells[target_status_idx].value)
    if current_status != "delete":
        raise SystemExit(
            f"Refusing to delete {work_id}: Works.status must be 'delete' (found '{current_status or 'blank'}')"
        )

    series_index_payload = load_json_object(series_index_path, "series_index.json")
    works_index_payload = load_json_object(works_index_path, "works_index.json")
    tag_assignments_payload = load_json_object(tag_assignments_path, "tag_assignments.json")

    series_stats = remove_work_from_series_index(series_index_payload, work_id)
    works_index_removed = remove_work_from_works_index(works_index_payload, work_id)
    override_stats = remove_work_overrides(tag_assignments_payload, work_id)

    finalize_series_index_payload(series_index_payload)
    finalize_works_index_payload(works_index_payload)

    work_page_path = works_dir / f"{work_id}.md"
    work_json_path = work_json_dir / f"{work_id}.json"
    detail_page_paths = sorted(work_details_dir.glob(f"{work_id}-*.md"))
    delete_paths = [work_page_path, work_json_path, *detail_page_paths]
    existing_delete_paths = [path for path in delete_paths if path.exists()]
    missing_delete_paths = [path for path in delete_paths if not path.exists()]

    json_updates: Dict[Path, Dict[str, Any]] = {}
    if (series_stats["membership_updates"] + series_stats["primary_nulls"] + series_stats["thumb_nulls"]) > 0:
        json_updates[series_index_path] = series_index_payload
    if works_index_removed:
        json_updates[works_index_path] = works_index_payload
    if override_stats["removed_overrides"] > 0:
        json_updates[tag_assignments_path] = tag_assignments_payload

    print(f"Delete target: {work_id}")
    print(f"Workbook row: {target_row_number} ({args.works_sheet}.status delete -> deleted)")
    print(
        "Series index updates: "
        f"membership rows touched={series_stats['membership_updates']}; "
        f"primary_work_id nulled={series_stats['primary_nulls']}"
    )
    print(f"Works index updates: removed entry={1 if works_index_removed else 0}")
    print(
        "Tag assignment updates: "
        f"series touched={override_stats['touched_series']}; "
        f"work overrides removed={override_stats['removed_overrides']}"
    )
    print(f"Delete generated files: {len(existing_delete_paths)} found; {len(missing_delete_paths)} already missing.")
    for path in existing_delete_paths:
        print(f"  - delete {path.relative_to(repo_root)}")
    for path in missing_delete_paths:
        print(f"  - missing {path.relative_to(repo_root)}")

    if not args.write:
        print("DRY-RUN: no files were changed.")
        log_event(
            "delete_work_dry_run",
            {
                "work_id": work_id,
                "row": target_row_number,
                "delete_paths": [str(path.relative_to(repo_root)) for path in existing_delete_paths],
                "missing_paths": [str(path.relative_to(repo_root)) for path in missing_delete_paths],
                "series_stats": series_stats,
                "works_index_removed": works_index_removed,
                "override_stats": override_stats,
            },
        )
        return

    backup_root = repo_root / "var" / "delete_work" / "backups" / backup_stamp_now()
    backup_targets = [xlsx_path, *json_updates.keys(), *existing_delete_paths]
    backups = backup_existing_paths(backup_targets, repo_root=repo_root, backup_root=backup_root)

    try:
        for path, payload in json_updates.items():
            write_json_atomic(path, payload)

        target_row_cells[target_status_idx].value = "deleted"
        write_workbook_atomic(xlsx_path, workbook)

        for path in existing_delete_paths:
            path.unlink()
    except Exception as exc:
        restore_backups(backups)
        raise SystemExit(
            "Delete failed after partial changes; restored backed up files from "
            f"{backup_root.relative_to(repo_root)} ({exc})"
        )

    print(f"WRITE complete. Backups: {backup_root.relative_to(repo_root)}")
    log_event(
        "delete_work_write",
        {
            "work_id": work_id,
            "row": target_row_number,
            "backup_root": str(backup_root.relative_to(repo_root)),
            "delete_paths": [str(path.relative_to(repo_root)) for path in existing_delete_paths],
            "missing_paths": [str(path.relative_to(repo_root)) for path in missing_delete_paths],
            "series_stats": series_stats,
            "works_index_removed": works_index_removed,
            "override_stats": override_stats,
        },
    )


if __name__ == "__main__":
    main()
