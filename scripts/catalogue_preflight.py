from __future__ import annotations

from pathlib import Path
import re
from typing import Any, Dict, List, Sequence

import openpyxl

try:
    from moment_sources import parse_front_matter
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.moment_sources import parse_front_matter

try:
    from pipeline_config import env_var_value, load_pipeline_config, source_moments_root_subdir
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.pipeline_config import env_var_value, load_pipeline_config, source_moments_root_subdir

try:
    from series_ids import normalize_series_id, parse_series_ids
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.series_ids import normalize_series_id, parse_series_ids


ACTIONABLE_STATUSES = {"draft", "published"}
PIPELINE_CONFIG = load_pipeline_config(Path(__file__))


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("'") and len(s) > 1:
        s = s[1:]
    return s


def normalize_status(value: Any) -> str:
    return normalize_text(value).lower()


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def slug_id(raw: Any, width: int = 5) -> str:
    if raw is None:
        raise ValueError("Missing id")
    s = normalize_text(raw)
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"\D", "", s)
    if not s:
        raise ValueError(f"Invalid id value: {raw!r}")
    return s.zfill(width)


def is_slug_safe(value: str) -> bool:
    return bool(re.match(r"^[a-z0-9]+(?:-[a-z0-9]+)*$", value))


def header_map(ws) -> Dict[str, int]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    hi: Dict[str, int] = {}
    for i, header in enumerate(headers):
        if header is None:
            continue
        key = normalize_text(header)
        if not key:
            continue
        hi[key] = i
        hi[key.lower()] = i
    return hi


def cell(row: Sequence[Any], header_index: Dict[str, int], col_name: str) -> Any:
    idx = header_index.get(col_name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _ensure_sheet(wb, sheet_name: str, errors: List[str]):
    if sheet_name not in wb.sheetnames:
        errors.append(f"Sheet not found in workbook: {sheet_name}")
        return None
    return wb[sheet_name]


def _ensure_columns(sheet_name: str, hi: Dict[str, int], required: Sequence[str], errors: List[str]) -> bool:
    missing = [name for name in required if name not in hi]
    if missing:
        errors.append(f"{sheet_name} sheet missing required columns: {', '.join(missing)}")
        return False
    return True


def validate_catalogue_workbook(
    xlsx_path: Path,
    *,
    works_sheet: str = "Works",
    series_sheet: str = "Series",
    work_details_sheet: str = "WorkDetails",
    projects_base_dir: Path | None = None,
) -> List[str]:
    errors: List[str] = []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    works_ws = _ensure_sheet(wb, works_sheet, errors)
    series_ws = _ensure_sheet(wb, series_sheet, errors)
    work_details_ws = _ensure_sheet(wb, work_details_sheet, errors)

    if errors:
        return errors

    works_hi = header_map(works_ws)
    series_hi = header_map(series_ws)
    work_details_hi = header_map(work_details_ws)

    if not _ensure_columns(works_sheet, works_hi, ["work_id", "status", "series_ids"], errors):
        return errors
    if not _ensure_columns(series_sheet, series_hi, ["series_id", "status", "primary_work_id"], errors):
        return errors
    if not _ensure_columns(work_details_sheet, work_details_hi, ["work_id", "detail_id", "status"], errors):
        return errors

    all_series_ids: set[str] = set()
    all_work_ids: set[str] = set()
    work_series_ids_by_work_id: Dict[str, List[str]] = {}

    for row_number, row in enumerate(works_ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_work_id = cell(row, works_hi, "work_id")
        if is_empty(raw_work_id):
            continue
        try:
            work_id = slug_id(raw_work_id)
        except ValueError as exc:
            errors.append(f"{works_sheet} row {row_number}: invalid work_id {raw_work_id!r} ({exc})")
            continue

        all_work_ids.add(work_id)
        try:
            series_ids = parse_series_ids(cell(row, works_hi, "series_ids"))
        except ValueError as exc:
            errors.append(f"{works_sheet} row {row_number}: {exc}")
            series_ids = []
        work_series_ids_by_work_id[work_id] = series_ids

        if normalize_status(cell(row, works_hi, "status")) not in ACTIONABLE_STATUSES:
            continue

    for row_number, row in enumerate(series_ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_series_id = cell(row, series_hi, "series_id")
        if is_empty(raw_series_id):
            continue
        try:
            series_id = normalize_series_id(raw_series_id)
        except ValueError as exc:
            errors.append(f"{series_sheet} row {row_number}: {exc}")
            continue
        all_series_ids.add(series_id)
        if normalize_status(cell(row, series_hi, "status")) not in ACTIONABLE_STATUSES:
            continue

        raw_primary_work_id = cell(row, series_hi, "primary_work_id")
        if is_empty(raw_primary_work_id):
            errors.append(f"{series_sheet} row {row_number}: series {series_id!r} missing primary_work_id")
            continue
        try:
            primary_work_id = slug_id(raw_primary_work_id)
        except ValueError as exc:
            errors.append(
                f"{series_sheet} row {row_number}: series {series_id!r} has invalid primary_work_id {raw_primary_work_id!r} ({exc})"
            )
            continue

        if primary_work_id not in all_work_ids:
            errors.append(
                f"{series_sheet} row {row_number}: series {series_id!r} primary_work_id {primary_work_id!r} not found in {works_sheet}"
            )
            continue

        series_membership = work_series_ids_by_work_id.get(primary_work_id, [])
        if series_id not in series_membership:
            errors.append(
                f"{series_sheet} row {row_number}: series {series_id!r} primary_work_id {primary_work_id!r} is not in that work's series_ids"
            )

    for row_number, row in enumerate(works_ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_work_id = cell(row, works_hi, "work_id")
        if is_empty(raw_work_id):
            continue
        try:
            work_id = slug_id(raw_work_id)
        except ValueError:
            continue
        if normalize_status(cell(row, works_hi, "status")) not in ACTIONABLE_STATUSES:
            continue

        for series_id in work_series_ids_by_work_id.get(work_id, []):
            if series_id not in all_series_ids:
                errors.append(
                    f"{works_sheet} row {row_number}: work {work_id!r} references unknown series_id {series_id!r}"
                )

    for row_number, row in enumerate(work_details_ws.iter_rows(min_row=2, values_only=True), start=2):
        if normalize_status(cell(row, work_details_hi, "status")) not in ACTIONABLE_STATUSES:
            continue
        raw_work_id = cell(row, work_details_hi, "work_id")
        raw_detail_id = cell(row, work_details_hi, "detail_id")
        if is_empty(raw_work_id) or is_empty(raw_detail_id):
            continue
        try:
            work_id = slug_id(raw_work_id)
        except ValueError as exc:
            errors.append(f"{work_details_sheet} row {row_number}: invalid work_id {raw_work_id!r} ({exc})")
            continue
        try:
            slug_id(raw_detail_id, width=3)
        except ValueError as exc:
            errors.append(f"{work_details_sheet} row {row_number}: invalid detail_id {raw_detail_id!r} ({exc})")
            continue
        if work_id not in all_work_ids:
            errors.append(
                f"{work_details_sheet} row {row_number}: work_id {work_id!r} not found in {works_sheet}"
            )

    configured_projects_base_dir = (
        projects_base_dir
        if projects_base_dir is not None
        else Path(env_var_value(PIPELINE_CONFIG, "projects_base_dir")).expanduser()
        if env_var_value(PIPELINE_CONFIG, "projects_base_dir")
        else None
    )
    if configured_projects_base_dir is not None:
        moments_root = configured_projects_base_dir / source_moments_root_subdir(PIPELINE_CONFIG)
        if moments_root.exists():
            for path in sorted(moments_root.glob("*.md")):
                moment_id = path.stem.strip().lower()
                if not moment_id:
                    continue
                if not is_slug_safe(moment_id):
                    errors.append(f"Moment source file {path.name!r}: moment_id is not slug-safe")
                    continue
                front_matter = parse_front_matter(path)
                status = normalize_status(front_matter.get("status"))
                if status and status not in ACTIONABLE_STATUSES:
                    errors.append(
                        f"Moment source file {path.name!r}: invalid status {front_matter.get('status')!r}; expected draft or published"
                    )

    return errors


def raise_if_invalid_catalogue_workbook(
    xlsx_path: Path,
    *,
    works_sheet: str = "Works",
    series_sheet: str = "Series",
    work_details_sheet: str = "WorkDetails",
    projects_base_dir: Path | None = None,
) -> None:
    errors = validate_catalogue_workbook(
        xlsx_path,
        works_sheet=works_sheet,
        series_sheet=series_sheet,
        work_details_sheet=work_details_sheet,
        projects_base_dir=projects_base_dir,
    )
    if not errors:
        return

    ordered = sorted(dict.fromkeys(errors))
    lines = ["Catalogue workbook preflight failed:", ""]
    lines.extend(f"- {line}" for line in ordered)
    lines.append("")
    lines.append("Fix the workbook errors above and re-run the command.")
    raise SystemExit("\n".join(lines))
