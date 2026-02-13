#!/usr/bin/env python3
"""
Copy and rename work files based on data/works.xlsx (Works sheet).

Source:  [PROJECTS_BASE_DIR]/projects/[project_folder]/[project_filename]
Target:  [WORKS_BASE_DIR]/works/make_srcset_images/[work_id][.ext]

Use --work-ids-file to limit processing to selected work_ids (one per line).

Usage:
  python3 scripts/copy_draft_work_files.py --write

Defaults to dry-run. Pass --write to actually copy files.
"""

from __future__ import annotations

import argparse
import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

import openpyxl

# ---------
# CONFIG
# ---------
# Hard-coded base folders
PROJECTS_BASE_DIR = Path("/Users/dlf/Library/CloudStorage/OneDrive-Personal/dotlineform")
WORKS_BASE_DIR = Path("/Users/dlf/Library/Mobile Documents/com~apple~CloudDocs/dotlineform")

WORKBOOK_PATH = Path("data/works.xlsx")
SHEET_NAME = "Works"

DEST_RELATIVE = Path("works/make_srcset_images")


def normalize_work_id(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    # Handle numeric-like IDs from Excel, e.g. 455.0
    if s.endswith(".0"):
        s = s[:-2]
    return s.zfill(5) if s.isdigit() else s


def header_map(ws) -> Dict[str, int]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return {str(h).strip(): i for i, h in enumerate(headers) if h is not None}


def main() -> int:
    parser = argparse.ArgumentParser(description="Copy draft work files from works.xlsx")
    parser.add_argument("--write", action="store_true", help="Actually copy files (default: dry-run)")
    parser.add_argument("--keep-ext", action="store_true", help="Keep source file extension on target (default)")
    parser.add_argument("--no-ext", action="store_true", help="Strip extension on target filename")
    parser.add_argument(
        "--work-ids-file",
        default="",
        help="Optional path to work_ids file (one id per line). Only those rows are processed.",
    )
    parser.add_argument(
        "--copied-ids-file",
        default="",
        help="Optional path to write copied work_ids (one per line).",
    )
    args = parser.parse_args()

    keep_ext = True
    if args.no_ext:
        keep_ext = False
    elif args.keep_ext:
        keep_ext = True

    wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=False, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Worksheet not found: {SHEET_NAME!r}")

    ws = wb[SHEET_NAME]
    cols = header_map(ws)

    required = ["work_id", "project_folder", "project_filename"]
    missing = [c for c in required if c not in cols]
    if missing:
        raise SystemExit(f"Missing required columns in Works sheet: {', '.join(missing)}")

    selected_ids: Optional[Set[str]] = None
    if args.work_ids_file:
        ids_path = Path(args.work_ids_file).expanduser()
        if not ids_path.exists():
            raise SystemExit(f"work_ids file not found: {ids_path}")
        selected_ids = set()
        for line in ids_path.read_text(encoding="utf-8").splitlines():
            s = normalize_work_id(line)
            if s:
                selected_ids.add(s)

    dest_dir = WORKS_BASE_DIR / DEST_RELATIVE
    if args.write:
        dest_dir.mkdir(parents=True, exist_ok=True)

    total = 0
    copied = 0
    copied_ids: List[str] = []
    missing_src = 0

    for row_cells in ws.iter_rows(min_row=2):
        row = [cell.value for cell in row_cells]
        work_id = row[cols["work_id"]]
        project_folder = row[cols["project_folder"]]
        project_filename = row[cols["project_filename"]]

        if not (work_id and project_folder and project_filename):
            continue

        work_id_str = normalize_work_id(work_id)
        if selected_ids is not None and work_id_str not in selected_ids:
            continue
        src = PROJECTS_BASE_DIR / "projects" / str(project_folder).strip() / str(project_filename).strip()

        if keep_ext:
            ext = Path(str(project_filename).strip()).suffix
            dest = dest_dir / f"{work_id_str}{ext}"
        else:
            dest = dest_dir / f"{work_id_str}"

        total += 1

        if not src.exists():
            print(f"Missing source: {src}")
            missing_src += 1
            continue

        if args.write:
            dest.parent.mkdir(parents=True, exist_ok=True)
            # Preserve metadata where possible
            from shutil import copy2

            copy2(src, dest)
            print(f"Copied: {src} -> {dest}")
            copied += 1
            copied_ids.append(work_id_str)
        else:
            print(f"Dry-run: {src} -> {dest}")
            copied_ids.append(work_id_str)

    if args.copied_ids_file:
        ids_path = Path(args.copied_ids_file).expanduser()
        ids_path.parent.mkdir(parents=True, exist_ok=True)
        ids_path.write_text("\n".join(copied_ids) + ("\n" if copied_ids else ""), encoding="utf-8")
        print(f"Wrote copied IDs manifest: {ids_path} ({len(copied_ids)} ids)")

    print(f"Draft rows: {total}, copied: {copied}, missing source: {missing_src}")
    if not args.write:
        print("Dry-run only. Re-run with --write to copy files.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
