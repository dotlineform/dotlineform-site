#!/usr/bin/env python3
"""
Generate Jekyll work/moment pages and JSON artifacts from an Excel workbook.

This repo stores works as a Jekyll collection in `_works/`. The generator writes one Markdown
file per work (e.g. `_works/00286.md`) with YAML front matter populated from these worksheets.

Series index JSON is written to assets/data/series_index.json.
Work-details JSON index files are written to assets/works/index/<work_id>.json (work-driven; one per selected work).
Lightweight works index JSON is written to assets/data/works_index.json (object keyed by work_id).
Recent publications JSON is written to assets/data/recent_index.json.
Studio-only work storage index JSON is written to assets/studio/data/work_storage_index.json (object keyed by work_id).
Moment JSON index files are written to assets/moments/index/<moment_id>.json (one per selected moment).
Lightweight moments index JSON is written to assets/data/moments_index.json (object keyed by moment_id).

- Works: base work metadata (1 row per work)
- Series: series master data (1 row per series_id)
- WorkDetails: additional detail images associated with a work
- WorkFiles: downloadable files associated with a work
- WorkLinks: published links associated with a work
- Moments: standalone moment entries sourced from `moments/*.md` front matter

YAML typing rules enforced by this script (so Excel cells do NOT need quoting):
- Numbers are emitted unquoted for: year, height_cm, width_cm, depth_cm
- Everything else is emitted as a quoted string (including fields like year_display)
- Empty cells become YAML null

Safe by default:
- dry-run unless you pass --write
- will not overwrite unless --force
- status gating (Works/Series):
  - draft -> process (candidate to publish)
  - published -> skip unless --force
  - unknown -> skip
  - when writing with --write: set status=published; set published_date=today if status changed or --force

specify work_ids to process with --work-ids (comma-separated list)
  - Only those IDs are processed; others are skipped early.
  - Status filtering still applies to the selected IDs unless you also pass --force.
Usage:
    python3 scripts/generate_work_pages.py --work-ids 00001,00002 --write
    python3 scripts/generate_work_pages.py --work-ids-file tmp/work_ids.txt --write
    python3 scripts/generate_work_pages.py --series-ids curve-poems,dots --write
    python3 scripts/generate_work_pages.py --moment-ids blue-sky,compiled --write

Common flags:
- --write: persist generated files + workbook status/date updates
- --force: regenerate even when checksum/hash matches existing output
- --work-ids / --work-ids-file: limit work/work_details generation scope
- --series-ids / --series-ids-file: limit series page/JSON scope
- --moment-ids / --moment-ids-file: limit moments generation scope
- --moments-output-dir: moment page destination
- --moments-json-dir: moment JSON output destination
- --moments-index-json-path: moments index JSON output destination
- --projects-base-dir: base path used for work/work_details dimension lookups, WorkFiles source lookup, and moment source discovery
- --media-base-dir: base path used for staging work download files into works/files

Path variables used by the script:
- projects_root = [projects-base-dir]/projects (work + work_details + work_files source lookup)
- moments_root = [projects-base-dir]/moments (moment prose source lookup)
- moments_images_root = [projects-base-dir]/moments/images (moment source image lookup)

"""

from __future__ import annotations

import argparse
import datetime as dt
import math
import os
import re
import shutil
import subprocess
from pathlib import Path
from typing import Any, Dict, List, Optional

import hashlib
import json
import sys

import openpyxl


RECENT_INDEX_SCHEMA = "recent_index_v1"
RECENT_INDEX_LIMIT = 50

try:
    from display_paths import format_display_path
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.display_paths import format_display_path

try:
    from script_logging import append_script_log
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.script_logging import append_script_log

try:
    from pipeline_config import (
        env_var_name,
        env_var_value,
        load_pipeline_config,
        media_work_files_subdir,
        source_moments_images_subdir,
        source_moments_root_subdir,
        source_works_prose_subdir,
        source_works_root_subdir,
    )
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.pipeline_config import (
        env_var_name,
        env_var_value,
        load_pipeline_config,
        media_work_files_subdir,
        source_moments_images_subdir,
        source_moments_root_subdir,
        source_works_prose_subdir,
        source_works_root_subdir,
    )


try:
    from catalogue_preflight import raise_if_invalid_catalogue_workbook
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.catalogue_preflight import raise_if_invalid_catalogue_workbook

try:
    from series_ids import normalize_series_id
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.series_ids import normalize_series_id

try:
    from moment_sources import load_moment_sources_manifest, scan_moment_source_files
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.moment_sources import load_moment_sources_manifest, scan_moment_source_files

try:
    from moment_sources import update_moment_source_front_matter
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.moment_sources import update_moment_source_front_matter


PIPELINE_CONFIG = load_pipeline_config(Path(__file__))
PROJECTS_BASE_DIR_ENV_NAME = env_var_name(PIPELINE_CONFIG, "projects_base_dir")
MEDIA_BASE_DIR_ENV_NAME = env_var_name(PIPELINE_CONFIG, "media_base_dir")


# ----------------------------
# Helpers (ID/date/YAML parsing)
# ----------------------------
# These functions exist to normalise common Excel quirks (e.g. numeric IDs like 361.0)
# and to keep YAML output safe/consistent.
def slug_id(raw: Any, width: int = 5) -> str:
    if raw is None:
        raise ValueError("Missing id")
    s = normalize_text(raw)
    # Handle numeric IDs like 361.0 from Excel
    s = re.sub(r"\.0$", "", s)
    # NOTE: This strips ALL non-digits. If your IDs ever include prefixes/suffixes, change this logic.
    s = re.sub(r"\D", "", s)  # keep digits only
    if not s:
        raise ValueError(f"Invalid id value: {raw!r}")
    return s.zfill(width)


def is_slug_safe(s: str) -> bool:
    return bool(re.match(r"^[a-z0-9]+(?:-[a-z0-9]+)*$", s))


def require_slug_safe(label: str, raw: Any) -> str:
    """Validate that `raw` is a slug-safe id and return it as a string."""
    if raw is None:
        raise ValueError(f"Missing {label}")
    s = normalize_text(raw)
    if not s:
        raise ValueError(f"Missing {label}")
    if not is_slug_safe(s):
        raise ValueError(f"{label} is not slug-safe: {s!r}")
    return s


def parse_date(raw: Any) -> Optional[str]:
    if raw is None or str(raw).strip() == "":
        return None
    if isinstance(raw, dt.datetime):
        return raw.date().isoformat()
    if isinstance(raw, dt.date):
        return raw.isoformat()
    s = normalize_text(raw)
    # NOTE: Prefer real Excel date cells or ISO strings in the workbook; anything else may pass through unchanged.
    # Accept YYYY-M-D and normalise to YYYY-MM-DD if possible
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
    if m:
        y, mo, d = map(int, m.groups())
        return dt.date(y, mo, d).isoformat()
    # Last resort: leave as-is (but you should fix upstream)
    return s


def parse_list(raw: Any, sep: str = ",") -> List[str]:
    if raw is None:
        return []
    s = normalize_text(raw)
    if not s:
        return []
    return [item.strip() for item in s.split(sep) if item.strip()]


def parse_work_id_selection(raw: str) -> set[str]:
    """
    Parse comma-separated work-id selectors supporting individual IDs and ranges.
    Examples:
      "66,74" -> {"00066", "00074"}
      "66-74,38-40,12" -> {"00012", "00038", ..., "00074"}
    """
    selected: set[str] = set()
    for token in (part.strip() for part in str(raw).split(",") if part.strip()):
        m = re.match(r"^(\d+)\s*-\s*(\d+)$", token)
        if m:
            start = int(m.group(1))
            end = int(m.group(2))
            if start > end:
                start, end = end, start
            for n in range(start, end + 1):
                selected.add(slug_id(n))
        else:
            selected.add(slug_id(token))
    return selected


def normalize_status(value: Any) -> str:
    if value is None:
        return ""
    return normalize_text(value).lower()


def normalize_text(value: Any) -> str:
    """Normalize Excel text by trimming and stripping a leading apostrophe prefix."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("'") and len(s) > 1:
        s = s[1:]
    return s


def numeric_aware_sort_key(value: Any, width: int = 3) -> str:
    """
    Build a numeric-aware string sort key by left-padding each run of digits.
    Example (width=3): "test 2.10" -> "test 002.010"
    """
    s = normalize_text(value)
    if not s:
        return ""
    return re.sub(r"\d+", lambda m: m.group(0).zfill(width), s)


def slug_anchor(value: Any) -> str:
    """
    Create a stable lowercase slug for anchor/query ids.
    Mirrors the client-side slug behavior used in work detail section ids.
    """
    s = normalize_text(value).lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"^-+|-+$", "", s)
    return s


def yaml_quote(s: str) -> str:
    """Quote a string safely for YAML."""
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{s}"'


NUMERIC_KEYS = {"year", "height_cm", "width_cm", "depth_cm", "width_px", "height_px"}
BOOLEAN_KEYS = set()


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def log_event(event: str, details: Optional[Dict[str, Any]] = None) -> None:
    try:
        append_script_log(Path(__file__), event=event, details=details or {})
    except Exception:
        # Logging failures must not block generation.
        pass


def coerce_numeric(value: Any) -> Optional[float]:
    """Best-effort numeric coercion for dimension fields; returns None if not parseable."""
    if is_empty(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except Exception:
        return None


def coerce_int(value: Any) -> Optional[int]:
    """Best-effort integer coercion for year; returns None if not parseable."""
    if is_empty(value):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    try:
        return int(str(value).strip())
    except Exception:
        return None


def coerce_string(value: Any) -> Optional[str]:
    """Coerce any non-empty value to a trimmed string (for quoted YAML output)."""
    if is_empty(value):
        return None
    s = normalize_text(value)
    return s if s != "" else None


def dump_scalar(key: str, value: Any) -> str:
    """Dump a scalar YAML key/value with typing rules."""
    if value is None:
        if key in BOOLEAN_KEYS:
            return f"{key}: false"
        return f"{key}: null"

    if key in BOOLEAN_KEYS:
        return f"{key}: {'true' if bool(value) else 'false'}"

    if key in NUMERIC_KEYS:
        # year is an int; dimensions are floats but we render ints without .0
        if key == "year":
            iv = coerce_int(value)
            return f"{key}: {iv}" if iv is not None else f"{key}: null"
        fv = coerce_numeric(value)
        if fv is None:
            return f"{key}: null"
        if fv.is_integer():
            return f"{key}: {int(fv)}"
        return f"{key}: {fv}"

    # Everything else is a quoted string
    sv = coerce_string(value)
    return f"{key}: {yaml_quote(sv)}" if sv is not None else f"{key}: null"


def dump_list_of_strings(key: str, values: List[str]) -> List[str]:
    lines: List[str] = [f"{key}:"]
    for v in values:
        lines.append(f"  - {yaml_quote(str(v))}")
    return lines


def dump_list_of_dicts(key: str, items: List[Dict[str, Any]], field_order: Optional[List[str]] = None) -> List[str]:
    """Dump a YAML list of dicts with stable key ordering."""
    lines: List[str] = [f"{key}:"]
    for item in items:
        # First line for each item
        lines.append("  -")
        keys = field_order if field_order else list(item.keys())
        for k in keys:
            if k not in item:
                continue
            # All nested fields here are strings by design; emit as quoted string or null
            sv = coerce_string(item.get(k))
            if sv is None:
                lines.append(f"    {k}: null")
            else:
                lines.append(f"    {k}: {yaml_quote(sv)}")
    return lines



def build_front_matter(fields: Dict[str, Any]) -> str:
    """Build YAML front matter in block style (supports lists of dicts)."""
    # Policy: we intentionally emit explicit empty values to keep the front matter schema stable.
    # - Empty/blank scalars become `key: null`
    # - Empty lists become `key: []`
    # This makes diffs, validation, and layout logic simpler and easier to test.
    lines: List[str] = ["---"]

    for k, v in fields.items():
        if isinstance(v, list):
            if not v:
                # Emit empty list explicitly (rare; keeps schema predictable)
                lines.append(f"{k}: []")
                continue

            # Detect list of dicts vs list of strings
            if all(isinstance(x, dict) for x in v):
                lines.extend(dump_list_of_dicts(k, v))
            else:
                # List of scalars -> strings
                lines.extend(dump_list_of_strings(k, [str(x) for x in v if not is_empty(x)]))
            continue

        # Scalars
        lines.append(dump_scalar(k, v))

    lines.append("---")
    return "\n".join(lines) + "\n"


# ----------------------------
# Canonical schema (Works sheet)
# ----------------------------
# Define the Works->front matter mapping once so adding a new field is a one-line change.
# Each entry is: (front_matter_key, excel_column_name, coercer)
# Coercers should return the Python type we want before YAML emission:
# - strings (quoted), ints/floats (unquoted for configured numeric keys), or None (-> null)
WORKS_SCHEMA: List[tuple[str, str, Any]] = [
    ("artist", "artist", coerce_string),
    ("title", "title", coerce_string),
    ("year", "year", coerce_int),
    ("year_display", "year_display", coerce_string),
    ("storage", "storage_location", coerce_string),
    ("medium_type", "medium_type", coerce_string),
    ("medium_caption", "medium_caption", coerce_string),
    ("duration", "duration", coerce_string),
    ("height_cm", "height_cm", coerce_numeric),
    ("width_cm", "width_cm", coerce_numeric),
    ("depth_cm", "depth_cm", coerce_numeric),
    ("width_px", "width_px", coerce_int),
    ("height_px", "height_px", coerce_int),
    # tags handled separately (csv list)
    # checksum is always computed, not sourced from Excel
]


def build_works_front_matter(works_row: tuple, works_hi: Dict[str, int]) -> Dict[str, Any]:
    """Build the scalar portion of Works front matter (excluding work_id, tags, images, attachments)."""
    fm: Dict[str, Any] = {}
    for fm_key, col_name, coercer in WORKS_SCHEMA:
        raw = works_row[works_hi[col_name]] if col_name in works_hi else None
        fm[fm_key] = coercer(raw)
    return fm


def build_download_entry(filename: Any, label: Any) -> Dict[str, str]:
    filename_value = coerce_string(filename)
    label_value = coerce_string(label)
    if filename_value is None:
        raise ValueError("Missing filename")
    if label_value is None:
        raise ValueError("Missing label")
    source_name = Path(filename_value).name
    stem = Path(source_name).stem
    suffix = "".join(Path(source_name).suffixes)
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "-", stem).strip("-._")
    safe_stem = re.sub(r"-{2,}", "-", safe_stem)
    if safe_stem == "":
        safe_stem = "file"
    return {
        "source_filename": source_name,
        "filename": f"{safe_stem}{suffix}",
        "label": label_value,
    }


def build_link_entry(url: Any, label: Any) -> Dict[str, str]:
    url_value = coerce_string(url)
    label_value = coerce_string(label)
    if url_value is None:
        raise ValueError("Missing URL")
    if label_value is None:
        raise ValueError("Missing label")
    return {
        "url": url_value,
        "label": label_value,
    }


# ----------------------------
# Checksum helpers
# ----------------------------

def compute_work_checksum(front_matter: Dict[str, Any]) -> str:
    """Compute a deterministic checksum for a work from its front matter (excluding checksum itself)."""
    payload = dict(front_matter)
    payload.pop("checksum", None)

    # Canonical JSON for hashing (sorted keys ensures deterministic output)
    canonical = json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    h = hashlib.blake2b(canonical, digest_size=16)
    return h.hexdigest()


def extract_existing_checksum(path: Path) -> Optional[str]:
    """Extract `checksum` from the YAML front matter of an existing work page."""
    try:
        text = path.read_text(encoding="utf-8")
    except Exception:
        return None

    # Only inspect the first YAML front matter block
    if not text.startswith("---"):
        return None

    parts = text.split("---", 2)
    if len(parts) < 3:
        return None

    fm_text = parts[1]
    for line in fm_text.splitlines():
        if not line.startswith("checksum:"):
            continue
        _, raw = line.split(":", 1)
        raw = raw.strip()
        if raw == "null" or raw == "":
            return None
        if raw.startswith('"') and raw.endswith('"') and len(raw) >= 2:
            return raw[1:-1]
        return raw

    return None


# ----------------------------
# JSON payload helpers
# ----------------------------


def compute_work_details_hash(work_id: str, sections: List[Dict[str, Any]]) -> str:
    """Compute deterministic hash for a work-details JSON payload."""
    payload = {"work_id": work_id, "sections": sections}
    canonical = json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return hashlib.blake2b(canonical, digest_size=16).hexdigest()


def canonicalize_for_hash(value: Any) -> Any:
    """Canonicalize values for deterministic hashing."""
    if isinstance(value, dict):
        out: Dict[str, Any] = {}
        for key in sorted(value.keys(), key=lambda k: str(k)):
            out[str(key)] = canonicalize_for_hash(value[key])
        return out
    if isinstance(value, list):
        return [canonicalize_for_hash(item) for item in value]
    if isinstance(value, tuple):
        return [canonicalize_for_hash(item) for item in value]
    if isinstance(value, float):
        if not math.isfinite(value):
            return value
        if value == 0.0:
            return 0
        if value.is_integer():
            return int(value)
        return float(f"{value:.15g}")
    return value


def compute_payload_hash_hex(payload: Any) -> str:
    """Compute deterministic blake2b hex hash for a canonicalized payload."""
    canonical = json.dumps(
        canonicalize_for_hash(payload),
        sort_keys=True,
        ensure_ascii=False,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return hashlib.blake2b(canonical, digest_size=16).hexdigest()


def compute_payload_version(payload: Any) -> str:
    """Compute deterministic blake2b content version token."""
    return f"blake2b-{compute_payload_hash_hex(payload)}"


def compact_json_value(value: Any, *, prune_empty_dicts: bool = True) -> Any:
    """Drop null object fields recursively while preserving empty lists."""
    if isinstance(value, dict):
        out: Dict[str, Any] = {}
        for key, item in value.items():
            compacted = compact_json_value(item, prune_empty_dicts=prune_empty_dicts)
            if compacted is None:
                continue
            out[key] = compacted
        if prune_empty_dicts and not out:
            return None
        return out
    if isinstance(value, list):
        out_list = []
        for item in value:
            compacted = compact_json_value(item, prune_empty_dicts=prune_empty_dicts)
            if compacted is None:
                continue
            out_list.append(compacted)
        return out_list
    return value


def compact_json_object(payload: Dict[str, Any]) -> Dict[str, Any]:
    compacted = compact_json_value(payload, prune_empty_dicts=False)
    return compacted if isinstance(compacted, dict) else {}


def normalize_recent_entry(entry: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(entry, dict):
        return None
    kind = normalize_text(entry.get("kind")).lower()
    if kind not in {"series", "work"}:
        return None
    target_id = normalize_text(entry.get("target_id"))
    title = coerce_string(entry.get("title"))
    caption = coerce_string(entry.get("caption"))
    published_date = parse_date(entry.get("published_date"))
    thumb_id = coerce_string(entry.get("thumb_id"))
    recorded_at_utc = coerce_string(entry.get("recorded_at_utc"))
    session_order = coerce_int(entry.get("session_order"))
    if not target_id or not title or not published_date:
        return None
    return compact_json_object({
        "id": f"{kind}:{target_id}",
        "kind": kind,
        "target_id": target_id,
        "title": title,
        "caption": caption,
        "published_date": published_date,
        "thumb_id": thumb_id,
        "recorded_at_utc": recorded_at_utc,
        "session_order": session_order,
    })


def load_recent_entries(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    entries_raw = payload.get("entries") if isinstance(payload, dict) else None
    if not isinstance(entries_raw, list):
        return []
    entries: List[Dict[str, Any]] = []
    for raw in entries_raw:
        normalized = normalize_recent_entry(raw)
        if normalized is not None:
            entries.append(normalized)
    return entries


def sort_recent_entries(entries: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def key(entry: Dict[str, Any]) -> tuple[str, str, int, str, str]:
        return (
            str(entry.get("published_date") or ""),
            str(entry.get("recorded_at_utc") or ""),
            -(coerce_int(entry.get("session_order")) or 0),
            str(entry.get("title") or ""),
            str(entry.get("target_id") or ""),
        )

    return sorted(entries, key=key, reverse=True)


def build_recent_index_payload(entries: List[Dict[str, Any]]) -> Dict[str, Any]:
    sorted_entries = sort_recent_entries(entries)[:RECENT_INDEX_LIMIT]
    version_payload = compact_json_object({
        "schema": RECENT_INDEX_SCHEMA,
        "entries": sorted_entries,
    })
    version = compute_payload_version(version_payload)
    return compact_json_object({
        "header": {
            "schema": RECENT_INDEX_SCHEMA,
            "version": version,
            "generated_at_utc": utc_timestamp_now(),
            "count": len(sorted_entries),
        },
        "entries": sorted_entries,
    })


def extract_existing_header_scalar(path: Path, key: str) -> Optional[str]:
    """Extract header.<key> from an existing JSON payload."""
    try:
        obj = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(obj, dict):
        return None
    header = obj.get("header")
    if not isinstance(header, dict):
        return None
    hv = header.get(key)
    if hv is None:
        return None
    s = str(hv).strip()
    return s or None


def render_markdown_with_jekyll(markdown_path: Path) -> str:
    """Render a markdown file to HTML using the repo's local Jekyll stack."""
    renderer_script = Path(__file__).resolve().with_name("render_markdown_with_jekyll.rb")
    if not renderer_script.exists():
        raise SystemExit(f"Markdown renderer helper not found: {renderer_script}")

    bundle_path = shutil.which("bundle")
    if bundle_path is None:
        raise SystemExit("Bundler not found in PATH; ensure the local Jekyll toolchain is available before generating moment JSON.")

    proc = subprocess.run(
        [bundle_path, "exec", "ruby", str(renderer_script), str(markdown_path.resolve())],
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        detail = (proc.stderr or proc.stdout).strip()
        raise SystemExit(f"Failed to render markdown with Jekyll: {markdown_path}\n{detail}")
    return proc.stdout


def parse_sips_pixel_dims(output: str) -> tuple[Optional[int], Optional[int]]:
    width = None
    height = None
    for line in output.splitlines():
        m_w = re.search(r"pixelWidth:\s*([0-9]+)", line)
        if m_w:
            width = int(m_w.group(1))
        m_h = re.search(r"pixelHeight:\s*([0-9]+)", line)
        if m_h:
            height = int(m_h.group(1))
    return width, height


def read_image_dims_px(path: Path) -> tuple[Optional[int], Optional[int]]:
    """Read pixel dimensions from an image file using macOS `sips` when available."""
    if not path.exists():
        return None, None
    if shutil.which("sips") is None:
        return None, None
    proc = subprocess.run(
        ["sips", "-g", "pixelWidth", "-g", "pixelHeight", str(path)],
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        return None, None
    return parse_sips_pixel_dims(proc.stdout)


def utc_timestamp_now() -> str:
    """Return current UTC timestamp formatted as YYYY-MM-DDTHH:MM:SSZ."""
    return dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def load_tag_assignments_payload(path: Path) -> Dict[str, Any]:
    """
    Load tag assignments JSON payload.
    If file is missing, return a default payload shape.
    """
    if not path.exists():
        return {
            "tag_assignments_version": "tag_assignments_v1",
            "updated_at_utc": utc_timestamp_now(),
            "series": {},
        }

    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise SystemExit(f"Failed to parse tag assignments JSON: {path} ({exc})")

    if not isinstance(payload, dict):
        raise SystemExit(f"Invalid tag assignments payload (expected object): {path}")

    if not isinstance(payload.get("series"), dict):
        payload["series"] = {}
    if not coerce_string(payload.get("tag_assignments_version")):
        payload["tag_assignments_version"] = "tag_assignments_v1"
    if not coerce_string(payload.get("updated_at_utc")):
        payload["updated_at_utc"] = utc_timestamp_now()
    for series_id, row in list(payload["series"].items()):
        if not isinstance(row, dict):
            payload["series"][series_id] = {
                "tags": [],
                "works": {},
                "updated_at_utc": utc_timestamp_now(),
            }
            continue
        if not isinstance(row.get("tags"), list):
            row["tags"] = []
        if "works" not in row or not isinstance(row.get("works"), dict):
            row["works"] = {}
    return payload


def load_tag_registry_payload(path: Path) -> Dict[str, Any]:
    """Load tag registry JSON payload or return a default shape when missing."""
    if not path.exists():
        return {
            "tag_registry_version": "tag_registry_v1",
            "updated_at_utc": utc_timestamp_now(),
            "tags": [],
        }

    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise SystemExit(f"Failed to parse tag registry JSON: {path} ({exc})")

    if not isinstance(payload, dict):
        raise SystemExit(f"Invalid tag registry payload (expected object): {path}")

    if not isinstance(payload.get("tags"), list):
        payload["tags"] = []
    if not coerce_string(payload.get("tag_registry_version")):
        payload["tag_registry_version"] = "tag_registry_v1"
    if not coerce_string(payload.get("updated_at_utc")):
        payload["updated_at_utc"] = utc_timestamp_now()
    return payload


# ----------------------------
# Main program
# ----------------------------
# High-level flow:
# 1) Parse CLI args (column mapping + output options)
# 2) Load workbook + sheets
# 3) Build header->index maps from row 1 of each sheet
# 4) Iterate Works rows -> build front matter + body -> write one file per work
def main() -> None:
    # CLI arguments define how we map Excel columns to front matter fields, and where output files go.
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default="data/works.xlsx", help="Path to Excel workbook (.xlsx)")

    # Worksheet names
    ap.add_argument("--works-sheet", default="Works", help="Worksheet name for base work metadata")
    ap.add_argument("--series-sheet", default="Series", help="Worksheet name for series master data")
    ap.add_argument("--series-sort-sheet", default="SeriesSort", help="Worksheet name for custom per-series sorting rules")
    ap.add_argument("--work-details-sheet", default="WorkDetails", help="Worksheet name for work detail metadata")
    ap.add_argument("--work-files-sheet", default="WorkFiles", help="Worksheet name for work download-file metadata")
    ap.add_argument("--work-links-sheet", default="WorkLinks", help="Worksheet name for work published-link metadata")

    # Output
    ap.add_argument("--output-dir", default="_works", help="Output folder for generated work pages")
    ap.add_argument("--series-output-dir", default="_series", help="Output folder for generated series pages")
    ap.add_argument("--series-json-dir", default="assets/series/index", help="Output folder for generated per-series JSON files")
    ap.add_argument("--series-index-json-path", default="assets/data/series_index.json", help="Output path for generated series index JSON")
    ap.add_argument("--work-details-output-dir", default="_work_details", help="Output folder for generated work detail pages")
    ap.add_argument("--works-json-dir", default="assets/works/index", help="Output folder for generated per-work detail JSON index files")
    ap.add_argument("--works-index-json-path", default="assets/data/works_index.json", help="Output path for generated lightweight works index JSON")
    ap.add_argument("--recent-index-json-path", default="assets/data/recent_index.json", help="Output path for generated recent publications index JSON")
    ap.add_argument("--work-storage-index-json-path", default="assets/studio/data/work_storage_index.json", help="Output path for generated Studio-only work storage index JSON")
    ap.add_argument("--moments-output-dir", default="_moments", help="Output folder for generated moment pages")
    ap.add_argument("--moments-json-dir", default="assets/moments/index", help="Output folder for generated per-moment JSON index files")
    ap.add_argument("--moments-index-json-path", default="assets/data/moments_index.json", help="Output path for generated lightweight moments index JSON")
    ap.add_argument(
        "--projects-base-dir",
        default=env_var_value(PIPELINE_CONFIG, "projects_base_dir"),
        help="Base folder containing the projects directory used to resolve WorkDetails and WorkFiles source files",
    )
    ap.add_argument(
        "--media-base-dir",
        default=env_var_value(PIPELINE_CONFIG, "media_base_dir"),
        help="Base folder containing staged media outputs such as works/files",
    )

    # Write controls
    ap.add_argument("--write", action="store_true", help="Actually write files (otherwise dry-run)")
    ap.add_argument("--force", action="store_true", help="Overwrite existing files")
    ap.add_argument(
        "--work-ids",
        default="",
        help=(
            "Comma-separated work_ids/ranges to process "
            "(e.g. 00001,00002 or 66-74,38-40). If set, only these IDs are processed."
        ),
    )
    ap.add_argument(
        "--work-ids-file",
        default="",
        help="Path to work_ids file (one id per line). If set, only these IDs are processed.",
    )
    ap.add_argument(
        "--series-ids",
        default="",
        help="Comma-separated series_ids to process for series page/index generation.",
    )
    ap.add_argument(
        "--series-ids-file",
        default="",
        help="Path to series_ids file (one id per line). If set, only these series are processed.",
    )
    ap.add_argument(
        "--moment-ids",
        default="",
        help="Comma-separated moment_ids to process.",
    )
    ap.add_argument(
        "--moment-ids-file",
        default="",
        help="Path to moment_ids file (one id per line). If set, only these moments are processed.",
    )
    ap.add_argument(
        "--moment-sources-manifest",
        default="",
        help="Optional JSON manifest of resolved moment source records.",
    )
    ap.add_argument(
        "--only",
        action="append",
        default=[],
        help=(
            "Limit run to selected artifacts. Repeat flag and/or pass comma-separated values. "
            "Allowed: work-pages,work-files,work-links,series-pages,series-index-json,work-details-pages,work-json,works-index-json,recent-index-json,moments,moments-index-json. "
            "Aggregate index JSON artifacts are always rebuilt on every run."
        ),
    )
    args = ap.parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    projects_base_dir_display = Path(args.projects_base_dir).expanduser() if normalize_text(args.projects_base_dir) else None
    media_base_dir_display = Path(args.media_base_dir).expanduser() if normalize_text(args.media_base_dir) else None

    def display_path(path: Path | str) -> str:
        return format_display_path(
            path,
            repo_root=repo_root,
            projects_base_dir=projects_base_dir_display,
            media_base_dir=media_base_dir_display,
        )

    def display_projects_path(path: Path | str) -> str:
        return format_display_path(
            path,
            repo_root=repo_root,
            projects_base_dir=projects_base_dir_display,
        )

    def display_media_path(path: Path | str) -> str:
        return format_display_path(
            path,
            repo_root=repo_root,
            media_base_dir=media_base_dir_display,
        )

    log_event(
        "generate_start",
        {
            "argv": sys.argv[1:],
            "write": bool(args.write),
            "force": bool(args.force),
        },
    )

    valid_artifacts = {
        "work-pages",
        "work-files",
        "work-links",
        "series-pages",
        "series-index-json",
        "work-details-pages",
        "work-json",
        "works-index-json",
        "recent-index-json",
        "moments",
        "moments-index-json",
    }
    selected_artifacts: Optional[set[str]] = None
    if args.only:
        requested: set[str] = set()
        for raw in args.only:
            for part in str(raw).split(","):
                item = part.strip().lower()
                if item:
                    requested.add(item)
        invalid = sorted(item for item in requested if item not in valid_artifacts)
        if invalid:
            raise SystemExit(
                "Invalid --only value(s): "
                + ", ".join(invalid)
                + ". Allowed: "
                + ", ".join(sorted(valid_artifacts))
            )
        selected_artifacts = requested

    def artifact_enabled(name: str) -> bool:
        if selected_artifacts is None:
            return True
        return name in selected_artifacts

    run_work_pages = artifact_enabled("work-pages")
    run_work_files = artifact_enabled("work-files")
    run_work_links = artifact_enabled("work-links")
    run_series_pages = artifact_enabled("series-pages")
    run_series_index_json = True
    run_work_details_pages = artifact_enabled("work-details-pages")
    run_work_json = artifact_enabled("work-json") or run_work_pages
    run_works_index_json = True
    run_moments_artifact = artifact_enabled("moments")
    run_moments_index_json = True
    run_studio_series_pages = False  # retired: use /studio/series-tag-editor/?series=<id>

    needs_projects_base = run_work_files or run_work_details_pages or run_work_json or run_series_pages or run_moments_artifact or run_moments_index_json
    if needs_projects_base and normalize_text(args.projects_base_dir) == "":
        raise SystemExit(
            f"Missing projects base directory. Set {PROJECTS_BASE_DIR_ENV_NAME} "
            "or pass --projects-base-dir."
        )
    if run_work_files and normalize_text(args.media_base_dir) == "":
        raise SystemExit(
            f"Missing media base directory. Set {MEDIA_BASE_DIR_ENV_NAME} "
            "or pass --media-base-dir."
        )

    # Resolve the workbook path and fail fast if it is missing.
    xlsx_path = Path(args.xlsx).expanduser()
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    # `data_only=True` reads the last-calculated values of formula cells.
    # Keep a second non-data_only workbook for writes so formulas are preserved on save.
    # If your sheet relies on formulas that haven't been calculated/saved, cached values may be None.
    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_write = openpyxl.load_workbook(xlsx_path, data_only=False) if args.write else None

    if args.works_sheet not in wb_values.sheetnames:
        raise SystemExit(f"Sheet not found in workbook: {args.works_sheet}")
    works_ws = wb_write[args.works_sheet] if wb_write is not None else wb_values[args.works_sheet]

    def read_sheet_rows(sheet_name: str) -> List[tuple]:
        if sheet_name not in wb_values.sheetnames:
            raise SystemExit(f"Sheet not found in workbook: {sheet_name}")
        ws = wb_values[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        return rows

    def build_header_index(rows: List[tuple]) -> Dict[str, int]:
        if not rows:
            return {}
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        hi: Dict[str, int] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            # Preserve original header and also store a lowercase alias for case-insensitive lookups.
            hi[h] = i
            hi[h.lower()] = i
        return hi

    def cell(row: tuple, header_index: Dict[str, int], col_name: str) -> Any:
        i = header_index.get(col_name)
        return None if i is None or i >= len(row) else row[i]

    def require_series_primary_work_id(
        sid: str,
        sr: tuple,
        *,
        ordered_work_ids: Optional[List[str]] = None,
    ) -> str:
        """Return a required primary_work_id for a series and validate membership when provided."""
        if "primary_work_id" not in series_hi:
            raise ValueError("Series sheet missing primary_work_id column")
        raw = cell(sr, series_hi, "primary_work_id")
        if is_empty(raw):
            raise ValueError(f"Series '{sid}' missing primary_work_id")
        wid = slug_id(raw)
        if ordered_work_ids and wid not in ordered_work_ids:
            raise ValueError(f"Series '{sid}' primary_work_id '{wid}' is not in its works list")
        return wid

    def read_cell_value(row: tuple, row_cells: tuple | None, header_index: Dict[str, int], col_name: str) -> Any:
        i = header_index.get(col_name)
        if i is None:
            return None
        if i < len(row):
            return row[i]
        if row_cells is not None and i < len(row_cells):
            return row_cells[i].value
        return None

    def first_present_col(header_index: Dict[str, int], names: List[str]) -> Optional[str]:
        for n in names:
            if n in header_index:
                return n
        return None

    def parse_work_series_ids(row: tuple) -> List[str]:
        series_ids_col = first_present_col(works_hi, ["series_ids"])
        if series_ids_col is None:
            return []
        series_ids: List[str] = []
        seen_series_ids: set[str] = set()
        for raw_series_id in parse_list(cell(row, works_hi, series_ids_col)):
            try:
                sid = normalize_series_id(raw_series_id)
            except ValueError:
                continue
            if sid in seen_series_ids:
                continue
            seen_series_ids.add(sid)
            series_ids.append(sid)
        return series_ids

    # Output directory:
    # - Use `works` for a normal pages folder.
    # - Use `_works` if you're using a Jekyll collection.
    out_dir = Path(args.output_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)

    series_out_dir = Path(args.series_output_dir).expanduser()
    series_out_dir.mkdir(parents=True, exist_ok=True)
    series_json_dir = Path(args.series_json_dir).expanduser()
    series_json_dir.mkdir(parents=True, exist_ok=True)
    studio_series_out_dir: Optional[Path] = None
    if run_studio_series_pages:
        studio_series_out_dir = Path("_studio_series").expanduser()
        studio_series_out_dir.mkdir(parents=True, exist_ok=True)

    tag_assignments_path = Path("assets/studio/data/tag_assignments.json").expanduser()
    tag_assignments_path.parent.mkdir(parents=True, exist_ok=True)
    series_index_json_path = Path(args.series_index_json_path).expanduser()
    series_index_json_path.parent.mkdir(parents=True, exist_ok=True)
    work_details_out_dir = Path(args.work_details_output_dir).expanduser()
    work_details_out_dir.mkdir(parents=True, exist_ok=True)
    works_json_dir = Path(args.works_json_dir).expanduser()
    works_json_dir.mkdir(parents=True, exist_ok=True)
    works_index_json_path = Path(args.works_index_json_path).expanduser()
    works_index_json_path.parent.mkdir(parents=True, exist_ok=True)
    recent_index_json_path = Path(args.recent_index_json_path).expanduser()
    recent_index_json_path.parent.mkdir(parents=True, exist_ok=True)
    work_storage_index_json_path = Path(args.work_storage_index_json_path).expanduser()
    work_storage_index_json_path.parent.mkdir(parents=True, exist_ok=True)
    moments_out_dir = Path(args.moments_output_dir).expanduser()
    moments_out_dir.mkdir(parents=True, exist_ok=True)
    moments_json_dir = Path(args.moments_json_dir).expanduser()
    moments_json_dir.mkdir(parents=True, exist_ok=True)
    moments_index_json_path = Path(args.moments_index_json_path).expanduser()
    moments_index_json_path.parent.mkdir(parents=True, exist_ok=True)
    projects_base_dir = Path(args.projects_base_dir).expanduser() if normalize_text(args.projects_base_dir) != "" else Path(".")
    projects_root = projects_base_dir / source_works_root_subdir(PIPELINE_CONFIG)
    media_base_dir = Path(args.media_base_dir).expanduser() if normalize_text(args.media_base_dir) != "" else None
    work_files_stage_dir = (media_base_dir / media_work_files_subdir(PIPELINE_CONFIG)) if media_base_dir is not None else None
    if run_work_files and work_files_stage_dir is not None:
        work_files_stage_dir.mkdir(parents=True, exist_ok=True)

    # Load all worksheets up-front.
    works_rows = read_sheet_rows(args.works_sheet)
    series_rows = read_sheet_rows(args.series_sheet)
    series_sort_rows = read_sheet_rows(args.series_sort_sheet) if args.series_sort_sheet in wb_values.sheetnames else []
    work_details_rows = read_sheet_rows(args.work_details_sheet) if args.work_details_sheet in wb_values.sheetnames else []
    work_files_rows = read_sheet_rows(args.work_files_sheet)
    work_links_rows = read_sheet_rows(args.work_links_sheet)
    series_ws = wb_write[args.series_sheet] if wb_write is not None else wb_values[args.series_sheet]
    work_details_ws = (
        wb_write[args.work_details_sheet] if wb_write is not None else wb_values[args.work_details_sheet]
    ) if args.work_details_sheet in wb_values.sheetnames else None
    work_files_ws = wb_write[args.work_files_sheet] if wb_write is not None else wb_values[args.work_files_sheet]
    work_links_ws = wb_write[args.work_links_sheet] if wb_write is not None else wb_values[args.work_links_sheet]

    if not works_rows:
        raise SystemExit(f"Works sheet '{args.works_sheet}' is empty")

    works_hi = build_header_index(works_rows)
    series_hi = build_header_index(series_rows) if series_rows else {}
    series_sort_hi = build_header_index(series_sort_rows) if series_sort_rows else {}
    work_details_hi = build_header_index(work_details_rows) if work_details_rows else {}
    work_files_hi = build_header_index(work_files_rows) if work_files_rows else {}
    work_links_hi = build_header_index(work_links_rows) if work_links_rows else {}

    if "status" not in works_hi:
        raise SystemExit("Works sheet missing required column: status")
    if "series_ids" not in works_hi:
        raise SystemExit("Works sheet missing required column: series_ids")
    if series_sort_rows:
        required_series_sort = ["series_id", "sort_fields"]
        missing_series_sort = [c for c in required_series_sort if c not in series_sort_hi]
        if missing_series_sort:
            raise SystemExit(f"{args.series_sort_sheet} sheet missing required columns: {', '.join(missing_series_sort)}")
    if work_details_rows and "status" not in work_details_hi:
        raise SystemExit("WorkDetails sheet missing required column: status")
    required_work_files = ["work_id", "filename", "label", "status", "published_date"]
    missing_work_files = [c for c in required_work_files if c not in work_files_hi]
    if missing_work_files:
        raise SystemExit(f"{args.work_files_sheet} sheet missing required columns: {', '.join(missing_work_files)}")
    required_work_links = ["work_id", "url", "label", "status", "published_date"]
    missing_work_links = [c for c in required_work_links if c not in work_links_hi]
    if missing_work_links:
        raise SystemExit(f"{args.work_links_sheet} sheet missing required columns: {', '.join(missing_work_links)}")
    series_duplicate_rows: Dict[str, List[int]] = {}
    if series_rows and len(series_rows) > 1 and "series_id" in series_hi:
        first_row_by_series_id: Dict[str, int] = {}
        for row_number, row in enumerate(series_rows[1:], start=2):
            raw_series_id = cell(row, series_hi, "series_id")
            if is_empty(raw_series_id):
                continue
            try:
                sid = normalize_series_id(raw_series_id)
            except ValueError:
                continue
            if sid not in first_row_by_series_id:
                first_row_by_series_id[sid] = row_number
                continue
            series_duplicate_rows.setdefault(sid, [first_row_by_series_id[sid]]).append(row_number)
        if series_duplicate_rows:
            duplicate_summary = "; ".join(
                f"{sid} (rows {', '.join(str(row_number) for row_number in row_numbers)})"
                for sid, row_numbers in sorted(series_duplicate_rows.items())
            )
            print(
                "Warning: Series sheet has duplicate series_id values; later rows overwrite earlier rows in generated output: "
                f"{duplicate_summary}"
            )

    # Pre-index series titles by series_id
    series_title_by_id: Dict[str, str] = {}
    for r in series_rows[1:] if len(series_rows) > 1 else []:
        sid_raw = cell(r, series_hi, "series_id")
        if is_empty(sid_raw):
            continue
        try:
            sid = normalize_series_id(sid_raw)
        except ValueError:
            continue
        title_raw = cell(r, series_hi, "title")
        title = coerce_string(title_raw)
        if title is None:
            continue
        series_title_by_id[sid] = title

    # Pre-index unique project folders by series_id from Works.
    series_project_folders_by_id: Dict[str, List[str]] = {}
    if "project_folder" in works_hi:
        project_folder_sets_by_series: Dict[str, set[str]] = {}
        for wr in works_rows[1:]:
            folder = coerce_string(cell(wr, works_hi, "project_folder"))
            series_ids = parse_work_series_ids(wr)
            if not series_ids or folder is None:
                continue
            for sid in series_ids:
                project_folder_sets_by_series.setdefault(sid, set()).add(folder)
        for sid, folder_set in project_folder_sets_by_series.items():
            series_project_folders_by_id[sid] = sorted(folder_set, key=lambda v: v.lower())

    # Compile canonical series_sort per work:
    # - default is work_id
    # - optional custom per-series rules come from SeriesSort(series_id, sort_fields)
    # - sort_fields supports comma-separated keys and optional '-' prefix for descending
    # - 'title' aliases to 'title_sort' for numeric-aware ordering
    # - work_id is always appended as final ascending tiebreaker
    works_sortable_fields = {fm_key for fm_key, _, _ in WORKS_SCHEMA}
    works_sortable_fields.update({"work_id", "series_title", "title_sort"})
    numeric_sort_fields = {"year", "height_cm", "width_cm", "depth_cm"}
    work_meta_by_id: Dict[str, Dict[str, Any]] = {}
    work_ids_by_series_all: Dict[str, List[str]] = {}
    for wr in works_rows[1:]:
        wid_raw = cell(wr, works_hi, "work_id")
        if is_empty(wid_raw):
            continue
        wid = slug_id(wid_raw)
        meta = build_works_front_matter(wr, works_hi)
        series_ids = parse_work_series_ids(wr)
        sid = series_ids[0] if series_ids else ""
        meta["work_id"] = wid
        meta["series_ids"] = series_ids
        meta["series_id"] = sid
        meta["series_title"] = series_title_by_id.get(sid) if sid else None
        work_meta_by_id[wid] = meta
        for series_id in series_ids:
            work_ids_by_series_all.setdefault(series_id, []).append(wid)

    series_sort_by_series_id: Dict[str, Dict[str, str]] = {
        sid: {wid: wid for wid in work_ids}
        for sid, work_ids in work_ids_by_series_all.items()
    }
    # Effective per-series sort_fields in user-facing terms (e.g. "title,work_id").
    # Defaults to work_id when no custom SeriesSort row exists.
    series_sort_fields_by_series_id: Dict[str, List[str]] = {
        sid: ["work_id"] for sid in work_ids_by_series_all.keys()
    }

    if series_sort_rows and len(series_sort_rows) > 1:
        seen_series_ids: set[str] = set()
        for sr in series_sort_rows[1:]:
            sid_raw = cell(sr, series_sort_hi, "series_id")
            if is_empty(sid_raw):
                continue
            sid = normalize_series_id(sid_raw)
            if sid in seen_series_ids:
                raise SystemExit(f"{args.series_sort_sheet} has duplicate series_id: {sid}")
            seen_series_ids.add(sid)

            sort_fields_raw = coerce_string(cell(sr, series_sort_hi, "sort_fields"))
            if sort_fields_raw is None or sort_fields_raw.strip() == "":
                raise SystemExit(f"{args.series_sort_sheet} has empty sort_fields for series_id: {sid}")

            parsed_fields: List[tuple[str, bool]] = []
            display_fields: List[str] = []
            for raw_token in sort_fields_raw.split(","):
                token = normalize_text(raw_token)
                if token == "":
                    continue
                desc = token.startswith("-")
                field = token[1:] if desc else token
                field = normalize_text(field).lower()
                display_field = field
                if field == "title":
                    field = "title_sort"
                    display_field = "title"
                elif field == "title_sort":
                    display_field = "title"
                if field not in works_sortable_fields:
                    raise SystemExit(
                        f"{args.series_sort_sheet} has unknown sort field '{field}' for series_id '{sid}'"
                    )
                if field == "work_id":
                    continue
                parsed_fields.append((field, desc))
                display_fields.append(f"-{display_field}" if desc else display_field)

            parsed_fields.append(("work_id", False))
            display_fields.append("work_id")
            series_sort_fields_by_series_id[sid] = display_fields
            series_work_ids = list(work_ids_by_series_all.get(sid, []))
            if not series_work_ids:
                continue

            def sortable_value(wid: str, field: str) -> Any:
                if field == "title_sort":
                    value = numeric_aware_sort_key(work_meta_by_id[wid].get("title"))
                else:
                    value = work_meta_by_id[wid].get(field)
                if field in numeric_sort_fields:
                    nv = coerce_numeric(value)
                    return float("-inf") if nv is None else nv
                return normalize_text(value).lower()

            for field, desc in reversed(parsed_fields):
                series_work_ids.sort(
                    key=lambda current_wid, current_field=field: sortable_value(current_wid, current_field),
                    reverse=desc,
                )

            rank_width = max(3, len(str(len(series_work_ids))))
            for idx, wid in enumerate(series_work_ids, start=1):
                series_sort_value = f"{idx:0{rank_width}d}-{wid}"
                series_sort_by_series_id.setdefault(sid, {})[wid] = series_sort_value

    # Pre-index project folder and prose filename by work_id.
    work_project_folder_by_id: Dict[str, str] = {}
    has_project_folder_col = "project_folder" in works_hi
    if (run_work_pages or run_work_json) and not has_project_folder_col:
        raise SystemExit("Works sheet missing required column for prose migration: project_folder")
    if has_project_folder_col:
        for wr in works_rows[1:]:
            wid_raw = cell(wr, works_hi, "work_id")
            pf_raw = cell(wr, works_hi, "project_folder")
            if is_empty(wid_raw) or is_empty(pf_raw):
                continue
            work_project_folder_by_id[slug_id(wid_raw)] = normalize_text(pf_raw)

    work_prose_file_by_id: Dict[str, str] = {}
    has_work_prose_file_col = "work_prose_file" in works_hi
    if has_work_prose_file_col:
        for wr in works_rows[1:]:
            wid_raw = cell(wr, works_hi, "work_id")
            prose_raw = cell(wr, works_hi, "work_prose_file")
            if is_empty(wid_raw) or is_empty(prose_raw):
                continue
            work_prose_file_by_id[slug_id(wid_raw)] = Path(normalize_text(prose_raw)).name

    works_prose_subdir = source_works_prose_subdir(PIPELINE_CONFIG)

    def resolve_work_prose_source_path(wid: str) -> Optional[Path]:
        project_folder = work_project_folder_by_id.get(wid)
        prose_filename = work_prose_file_by_id.get(wid)
        if not project_folder or not prose_filename:
            return None
        return projects_root / project_folder / works_prose_subdir / prose_filename

    has_series_prose_file_col = "series_prose_file" in series_hi
    if run_series_pages and not has_series_prose_file_col:
        raise SystemExit("Series sheet missing required column for prose migration: series_prose_file")

    def resolve_series_prose_source_path(series_id: str, sr: tuple, *, ordered_work_ids: Optional[List[str]] = None) -> Optional[Path]:
        primary_work_id = require_series_primary_work_id(
            series_id,
            sr,
            ordered_work_ids=ordered_work_ids,
        )
        project_folder = work_project_folder_by_id.get(primary_work_id)
        prose_raw = cell(sr, series_hi, "series_prose_file") if has_series_prose_file_col else None
        prose_filename = Path(normalize_text(prose_raw)).name if not is_empty(prose_raw) else None
        if not project_folder or not prose_filename:
            return None
        return projects_root / project_folder / works_prose_subdir / prose_filename

    work_file_entries_by_work_id: Dict[str, List[Dict[str, Any]]] = {}
    if len(work_files_rows) > 1:
        for row_number, (wf_row, wf_cells) in enumerate(
            zip(work_files_rows[1:], work_files_ws.iter_rows(min_row=2), strict=False),
            start=2,
        ):
            wid_raw = cell(wf_row, work_files_hi, "work_id")
            if is_empty(wid_raw):
                continue
            wid = slug_id(wid_raw)
            try:
                download_entry = build_download_entry(
                    read_cell_value(wf_row, wf_cells, work_files_hi, "filename"),
                    read_cell_value(wf_row, wf_cells, work_files_hi, "label"),
                )
            except ValueError as exc:
                raise SystemExit(f"{args.work_files_sheet} row {row_number}: {exc}") from exc
            work_file_entries_by_work_id.setdefault(wid, []).append(
                {
                    "source_filename": download_entry["source_filename"],
                    "filename": download_entry["filename"],
                    "label": download_entry["label"],
                    "row_number": row_number,
                    "row_cells": wf_cells,
                }
            )

    work_link_entries_by_work_id: Dict[str, List[Dict[str, Any]]] = {}
    if len(work_links_rows) > 1:
        for row_number, (wl_row, wl_cells) in enumerate(
            zip(work_links_rows[1:], work_links_ws.iter_rows(min_row=2), strict=False),
            start=2,
        ):
            wid_raw = cell(wl_row, work_links_hi, "work_id")
            if is_empty(wid_raw):
                continue
            wid = slug_id(wid_raw)
            try:
                link_entry = build_link_entry(
                    read_cell_value(wl_row, wl_cells, work_links_hi, "url"),
                    read_cell_value(wl_row, wl_cells, work_links_hi, "label"),
                )
            except ValueError as exc:
                raise SystemExit(f"{args.work_links_sheet} row {row_number}: {exc}") from exc
            work_link_entries_by_work_id.setdefault(wid, []).append(
                {
                    "url": link_entry["url"],
                    "label": link_entry["label"],
                    "row_number": row_number,
                    "row_cells": wl_cells,
                }
            )

    works_field_order = [
        "work_id",
        "title",
        "year",
        "year_display",
        "series_id",
        "series_ids",
        "series_title",
        "series_sort",
        "storage",
        "medium_type",
        "medium_caption",
        "duration",
        "links",
        "height_cm",
        "width_cm",
        "depth_cm",
        "width_px",
        "height_px",
        "downloads",
        "artist",
    ]

    def build_canonical_work_record(wid: str) -> Optional[Dict[str, Any]]:
        base = work_meta_by_id.get(wid)
        if base is None:
            return None
        fm: Dict[str, Any] = {"work_id": wid}
        fm.update(base)
        downloads = [
            {"filename": entry["filename"], "label": entry["label"]}
                for entry in work_file_entries_by_work_id.get(wid, [])
        ]
        if downloads:
            fm["downloads"] = downloads
        links = [
            {"url": entry["url"], "label": entry["label"]}
            for entry in work_link_entries_by_work_id.get(wid, [])
        ]
        if links:
            fm["links"] = links
        raw_series_ids = fm.get("series_ids")
        series_ids = [coerce_string(item) for item in raw_series_ids] if isinstance(raw_series_ids, list) else []
        series_ids = [item for item in series_ids if item is not None]
        sid = series_ids[0] if series_ids else coerce_string(fm.get("series_id"))
        fm["series_id"] = sid
        fm["series_ids"] = series_ids
        fm["series_title"] = series_title_by_id.get(sid) if sid is not None else None
        fm["series_sort"] = series_sort_by_series_id.get(sid, {}).get(wid, wid) if sid is not None else wid

        fm_ordered: Dict[str, Any] = {}
        for key in works_field_order:
            if key in fm:
                fm_ordered[key] = fm[key]
        for key, value in fm.items():
            if key not in fm_ordered:
                fm_ordered[key] = value
        fm = fm_ordered
        fm["checksum"] = compute_work_checksum(fm)
        return fm

    def build_canonical_detail_record(
        wid: str,
        did: str,
        title: Optional[str],
        project_subfolder: Optional[str],
        width_px: Optional[int],
        height_px: Optional[int],
    ) -> Dict[str, Any]:
        detail_uid = f"{wid}-{did}"
        dfm: Dict[str, Any] = {
            "work_id": wid,
            "detail_id": did,
            "detail_uid": detail_uid,
            "title": title,
            "project_subfolder": project_subfolder,
            "width_px": width_px,
            "height_px": height_px,
            "layout": "work_details",
        }
        return compact_json_object(dfm)

    def build_work_index_record(work_record: Dict[str, Any]) -> Dict[str, Any]:
        wid = str(work_record.get("work_id", ""))
        title_value = coerce_string(work_record.get("title"))
        year_value = work_record.get("year")
        year_display_value = coerce_string(work_record.get("year_display"))
        return compact_json_object({
            "work_id": wid,
            "title": title_value,
            "year": year_value,
            "year_display": year_display_value if year_display_value is not None else (str(year_value) if year_value is not None else None),
            "series_ids": list(work_record.get("series_ids", [])) if isinstance(work_record.get("series_ids"), list) else [],
        })

    def build_work_storage_index_record(work_record: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        storage_value = coerce_string(work_record.get("storage"))
        if storage_value is None:
            return None
        return compact_json_object({
            "storage": storage_value,
        })

    def build_work_json_record(work_record: Dict[str, Any]) -> Dict[str, Any]:
        public_record = dict(work_record)
        public_record.pop("series_id", None)
        public_record.pop("series_title", None)
        public_record.pop("series_sort", None)
        public_record.pop("title_sort", None)
        public_record.pop("checksum", None)
        return compact_json_object(public_record)

    def build_series_json_record(series_record: Dict[str, Any]) -> Dict[str, Any]:
        public_record = dict(series_record)
        public_record.pop("layout", None)
        public_record.pop("checksum", None)
        return compact_json_object(public_record)

    def build_moment_json_record(moment_record: Dict[str, Any]) -> Dict[str, Any]:
        public_record = dict(moment_record)
        public_record.pop("layout", None)
        public_record.pop("checksum", None)
        return compact_json_object(public_record)

    def build_moment_index_record(moment_record: Dict[str, Any]) -> Dict[str, Any]:
        moment_id_value = coerce_string(moment_record.get("moment_id"))
        title_value = coerce_string(moment_record.get("title"))
        date_value = coerce_string(moment_record.get("date"))
        date_display_value = coerce_string(moment_record.get("date_display"))
        images_value = moment_record.get("images")
        thumb_id_value = moment_id_value if isinstance(images_value, list) and len(images_value) > 0 else None
        return compact_json_object({
            "moment_id": moment_id_value,
            "title": title_value,
            "date": date_value,
            "date_display": date_display_value,
            "thumb_id": thumb_id_value,
        })

    def build_sections_from_detail_records(detail_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        section_index: Dict[str, int] = {}
        sections: List[Dict[str, Any]] = []
        for detail in detail_records:
            project_subfolder = coerce_string(detail.get("project_subfolder")) or ""
            if project_subfolder not in section_index:
                section_index[project_subfolder] = len(sections)
                sections.append(
                    {
                        "project_subfolder": project_subfolder,
                        "details": [],
                    }
                )
            sections[section_index[project_subfolder]]["details"].append(dict(detail))
        for sec in sections:
            details = sec.get("details")
            if isinstance(details, list):
                details.sort(key=lambda item: str(item.get("detail_id", "")))
        return sections

    written = 0
    skipped = 0
    downloads_copied = 0
    downloads_missing = 0
    work_files_status_updated = 0
    work_files_published_date_updated = 0
    work_files_status_idx = work_files_hi.get("status")
    work_files_published_date_idx = work_files_hi.get("published_date")
    work_files_published_date_missing_warned = False
    work_links_status_updated = 0
    work_links_published_date_updated = 0
    work_links_status_idx = work_links_hi.get("status")
    work_links_published_date_idx = work_links_hi.get("published_date")
    work_links_published_date_missing_warned = False
    run_work_processing = run_work_pages or run_work_files or run_work_links
    run_work_selection_scope = run_work_processing or run_work_json
    run_work_dimension_refresh = run_work_json

    def is_actionable_status(status_value: str) -> bool:
        if status_value == "draft":
            return True
        if status_value == "published" and args.force:
            return True
        return False

    # Optional filtering: allow a specific list of work_ids (from file or comma-separated arg).
    selected_ids = None
    explicit_work_filter = bool(args.work_ids_file or args.work_ids)
    if args.work_ids_file:
        ids_path = Path(args.work_ids_file).expanduser()
        if not ids_path.exists():
            raise SystemExit(f"work_ids file not found: {ids_path}")
        selected_ids = {slug_id(line.strip()) for line in ids_path.read_text(encoding="utf-8").splitlines() if line.strip()}
    elif args.work_ids:
        selected_ids = parse_work_id_selection(args.work_ids)

    selected_series_ids = None
    if args.series_ids_file:
        sids_path = Path(args.series_ids_file).expanduser()
        if not sids_path.exists():
            raise SystemExit(f"series_ids file not found: {sids_path}")
        try:
            selected_series_ids = {
                normalize_series_id(line.strip())
                for line in sids_path.read_text(encoding="utf-8").splitlines()
                if line.strip()
            }
        except ValueError as exc:
            raise SystemExit(f"Invalid series_ids file value: {exc}") from exc
    elif args.series_ids:
        try:
            selected_series_ids = {
                normalize_series_id(sid.strip())
                for sid in args.series_ids.split(",")
                if sid.strip()
            }
        except ValueError as exc:
            raise SystemExit(f"Invalid --series-ids value: {exc}") from exc

    selected_moment_ids = None
    if args.moment_ids_file:
        mids_path = Path(args.moment_ids_file).expanduser()
        if not mids_path.exists():
            raise SystemExit(f"moment_ids file not found: {mids_path}")
        selected_moment_ids = {
            require_slug_safe("moment_id", normalize_text(line).lower())
            for line in mids_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        }
    elif args.moment_ids:
        selected_moment_ids = {
            require_slug_safe("moment_id", normalize_text(mid).lower())
            for mid in args.moment_ids.split(",")
            if mid.strip()
        }
    explicit_moment_filter = bool(args.moment_ids_file or args.moment_ids)
    moment_sources_manifest_records: Dict[str, Dict[str, Any]] = {}
    if args.moment_sources_manifest:
        manifest_path = Path(args.moment_sources_manifest).expanduser()
        if not manifest_path.exists():
            raise SystemExit(f"moment sources manifest not found: {manifest_path}")
        moment_sources_manifest_records = load_moment_sources_manifest(manifest_path)

    # If caller scopes by series but does not provide an explicit work filter:
    # - when work artifacts are explicitly selected via --only, derive selected work_ids from those series
    # - otherwise skip work-page processing by default (backward compatible behavior)
    if selected_series_ids is not None and not explicit_work_filter:
        if selected_artifacts is not None and run_work_selection_scope:
            selected_ids = set()
            for r in works_rows[1:]:
                raw_work_id = cell(r, works_hi, "work_id")
                if is_empty(raw_work_id):
                    continue
                wid = slug_id(raw_work_id)
                series_ids = parse_work_series_ids(r)
                if any(sid in selected_series_ids for sid in series_ids):
                    selected_ids.add(wid)
        else:
            selected_ids = set()
    # If caller scopes by work/series and does not provide an explicit moments filter,
    # skip moments generation by default (unless moments was explicitly selected via --only).
    run_moments = run_moments_artifact
    if (
        selected_artifacts is None
        and (explicit_work_filter or selected_series_ids is not None)
        and not explicit_moment_filter
    ):
        run_moments = False
        run_moments_index_json = False

    raise_if_invalid_catalogue_workbook(
        xlsx_path,
        works_sheet=args.works_sheet,
        series_sheet=args.series_sheet,
        work_details_sheet=args.work_details_sheet,
        projects_base_dir=Path(args.projects_base_dir).expanduser(),
    )

    works_width_px_idx = works_hi.get("width_px")
    works_height_px_idx = works_hi.get("height_px")
    if args.write and run_work_dimension_refresh:
        if works_width_px_idx is None:
            works_width_px_idx = works_ws.max_column
            works_ws.cell(row=1, column=works_width_px_idx + 1, value="width_px")
            works_hi["width_px"] = works_width_px_idx
        if works_height_px_idx is None:
            works_height_px_idx = works_ws.max_column
            works_ws.cell(row=1, column=works_height_px_idx + 1, value="height_px")
            works_hi["height_px"] = works_height_px_idx

    work_dimensions_updated = 0
    work_project_folder_missing_warned = False
    if run_work_dimension_refresh:
        for wr, wr_cells in zip(works_rows[1:], works_ws.iter_rows(min_row=2), strict=False):
            raw_work_id = cell(wr, works_hi, "work_id")
            if is_empty(raw_work_id):
                continue
            wid = slug_id(raw_work_id)
            if selected_ids is not None and wid not in selected_ids:
                continue
            status = normalize_status(cell(wr, works_hi, "status"))
            if status not in {"draft", "published"}:
                continue

            width_px = coerce_int(read_cell_value(wr, wr_cells, works_hi, "width_px")) if "width_px" in works_hi else None
            height_px = coerce_int(read_cell_value(wr, wr_cells, works_hi, "height_px")) if "height_px" in works_hi else None
            project_filename = coerce_string(read_cell_value(wr, wr_cells, works_hi, "project_filename")) if "project_filename" in works_hi else None

            src_path: Optional[Path] = None
            if project_filename:
                if Path(project_filename).is_absolute():
                    src_path = Path(project_filename)
                else:
                    project_folder = work_project_folder_by_id.get(wid)
                    if project_folder:
                        src_path = projects_root / project_folder / project_filename
                    elif not work_project_folder_missing_warned:
                        if not has_project_folder_col:
                            print("Warning: Works sheet has no project_folder column; cannot persist work image dimensions.")
                        else:
                            print("Warning: missing Works.project_folder for one or more works; cannot persist those image dimensions.")
                        work_project_folder_missing_warned = True

            if src_path is not None:
                src_w, src_h = read_image_dims_px(src_path)
                if src_w is not None and src_h is not None:
                    width_px = src_w
                    height_px = src_h
                    if args.write and works_width_px_idx is not None and works_height_px_idx is not None:
                        prev_w = wr_cells[works_width_px_idx].value if works_width_px_idx < len(wr_cells) else None
                        prev_h = wr_cells[works_height_px_idx].value if works_height_px_idx < len(wr_cells) else None
                        if prev_w != src_w or prev_h != src_h:
                            wr_cells[works_width_px_idx].value = src_w
                            wr_cells[works_height_px_idx].value = src_h
                            work_dimensions_updated += 1
                else:
                    print(f"Warning: could not read dimensions for work primary source image: {display_projects_path(src_path)}")
            elif project_filename:
                print(f"Warning: could not resolve work primary source image path for {wid} ({project_filename})")

            meta = work_meta_by_id.get(wid)
            if meta is not None:
                meta["width_px"] = width_px
                meta["height_px"] = height_px

    canonical_work_record_by_id: Dict[str, Dict[str, Any]] = {}
    for wid in sorted(work_meta_by_id.keys()):
        record = build_canonical_work_record(wid)
        if record is not None:
            canonical_work_record_by_id[wid] = record

    total = 0
    if run_work_processing:
        for r in works_rows[1:]:
            raw_work_id = cell(r, works_hi, "work_id")
            if is_empty(raw_work_id):
                continue
            wid = slug_id(raw_work_id)
            if selected_ids is not None and wid not in selected_ids:
                continue
            status = normalize_status(cell(r, works_hi, "status"))
            if is_actionable_status(status):
                total += 1

    processed = 0
    status_updated = 0
    published_date_updated = 0
    published_date_idx = works_hi.get("published_date")
    published_date_missing_warned = False
    today = dt.date.today()
    work_publish_transitions: List[Dict[str, Any]] = []

    # Iterate each Works row and emit one Markdown file per work.
    if run_work_processing:
        for r, row_cells in zip(works_rows[1:], works_ws.iter_rows(min_row=2), strict=False):
            status = normalize_status(cell(r, works_hi, "status"))
            raw_work_id = cell(r, works_hi, "work_id")
            if is_empty(raw_work_id):
                skipped += 1
                continue
            wid = slug_id(raw_work_id)

            if selected_ids is not None and wid not in selected_ids:
                skipped += 1
                continue

            if not is_actionable_status(status):
                skipped += 1
                continue

            processed += 1
            prefix = f"[{processed}/{total}] "
            if run_work_pages:
                canonical_work_fm = build_canonical_work_record(wid)
                if canonical_work_fm is None:
                    skipped += 1
                    continue
                checksum = str(canonical_work_fm.get("checksum"))
                # Canonical work metadata lives in JSON artifacts; keep _works lightweight.
                work_page_fm: Dict[str, Any] = {
                    "work_id": wid,
                    "title": coerce_string(canonical_work_fm.get("title")),
                    "layout": "work",
                    "checksum": checksum,
                }

                work_page_content = build_front_matter(work_page_fm) + "\n"
                out_path = out_dir / f"{wid}.md"

                def write_page(path: Path, label: str, page_content: str) -> bool:
                    exists = path.exists()
                    existing_checksum = extract_existing_checksum(path) if exists else None
                    if (existing_checksum is not None) and (existing_checksum == checksum) and (not args.force):
                        existing_content: Optional[str] = None
                        try:
                            existing_content = path.read_text(encoding="utf-8")
                        except Exception:
                            existing_content = None
                        if existing_content == page_content:
                            print(f"{prefix}SKIP ({label}; checksum+content match): {display_path(path)}")
                            return False
                    if args.write:
                        path.write_text(page_content, encoding="utf-8")
                        print(f"{prefix}WRITE ({label}): {display_path(path)}")
                    else:
                        print(f"{prefix}DRY-RUN: would write {display_path(path)} (overwrite={exists})")
                    return True

                if write_page(out_path, "work", work_page_content):
                    written += 1
                    if args.write:
                        status_idx = works_hi["status"]
                        status_was = normalize_status(row_cells[status_idx].value)
                        if status_was != "published":
                            row_cells[status_idx].value = "published"
                            status_updated += 1
                        if (status_was != "published") or args.force:
                            if published_date_idx is not None:
                                row_cells[published_date_idx].value = today
                                published_date_updated += 1
                            elif not published_date_missing_warned:
                                print("Warning: Works sheet missing published_date column; skipping date updates.")
                                published_date_missing_warned = True
                        if status_was != "published":
                            work_meta = work_meta_by_id.get(wid, {})
                            series_ids = work_meta.get("series_ids") if isinstance(work_meta, dict) else []
                            primary_series_id = series_ids[0] if isinstance(series_ids, list) and series_ids else ""
                            work_publish_transitions.append({
                                "work_id": wid,
                                "title": coerce_string(work_meta.get("title")) if isinstance(work_meta, dict) else None,
                                "primary_series_id": primary_series_id,
                                "series_title": series_title_by_id.get(primary_series_id) if primary_series_id else None,
                                "published_date": today.isoformat(),
                            })
                else:
                    skipped += 1

            if run_work_files:
                project_folder = work_project_folder_by_id.get(wid)
                work_file_entries = work_file_entries_by_work_id.get(wid, [])
                for entry in work_file_entries:
                    source_filename = entry["source_filename"]
                    download_name = entry["filename"]
                    row_cells = entry["row_cells"]
                    if project_folder:
                        download_src = projects_root / project_folder / source_filename
                    else:
                        download_src = None
                    if work_files_stage_dir is None:
                        download_dest = None
                    else:
                        download_dest = work_files_stage_dir / f"{wid}-{download_name}"

                    if download_src is None:
                        print(f"{prefix}Warning: cannot resolve download source for {wid} ({download_name})")
                        downloads_missing += 1
                        continue
                    if not download_src.exists():
                        print(f"{prefix}Warning: missing download source: {display_projects_path(download_src)}")
                        downloads_missing += 1
                        continue
                    if download_dest is None:
                        print(f"{prefix}Warning: cannot resolve work file staging destination for {wid} ({download_name})")
                        downloads_missing += 1
                        continue

                    if args.write:
                        download_dest.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(download_src, download_dest)
                        print(f"{prefix}COPY download: {display_projects_path(download_src)} -> {display_media_path(download_dest)}")
                        downloads_copied += 1
                        if work_files_status_idx is not None:
                            status_was = normalize_status(row_cells[work_files_status_idx].value)
                            if status_was != "published":
                                row_cells[work_files_status_idx].value = "published"
                                work_files_status_updated += 1
                            if (status_was != "published") or args.force:
                                if work_files_published_date_idx is not None:
                                    row_cells[work_files_published_date_idx].value = today
                                    work_files_published_date_updated += 1
                                elif not work_files_published_date_missing_warned:
                                    print(f"Warning: {args.work_files_sheet} sheet missing published_date column; skipping date updates.")
                                    work_files_published_date_missing_warned = True
                    else:
                        print(f"{prefix}DRY-RUN: would copy download {display_projects_path(download_src)} -> {display_media_path(download_dest)}")
                        downloads_copied += 1

            if run_work_links:
                work_link_entries = work_link_entries_by_work_id.get(wid, [])
                for entry in work_link_entries:
                    row_cells = entry["row_cells"]
                    if args.write:
                        if work_links_status_idx is not None:
                            status_was = normalize_status(row_cells[work_links_status_idx].value)
                            if status_was != "published":
                                row_cells[work_links_status_idx].value = "published"
                                work_links_status_updated += 1
                            if (status_was != "published") or args.force:
                                if work_links_published_date_idx is not None:
                                    row_cells[work_links_published_date_idx].value = today
                                    work_links_published_date_updated += 1
                                elif not work_links_published_date_missing_warned:
                                    print(f"Warning: {args.work_links_sheet} sheet missing published_date column; skipping date updates.")
                                    work_links_published_date_missing_warned = True

    if args.write and (
        status_updated > 0
        or work_dimensions_updated > 0
        or work_files_status_updated > 0
        or work_files_published_date_updated > 0
        or work_links_status_updated > 0
        or work_links_published_date_updated > 0
    ):
        assert wb_write is not None
        wb_write.save(xlsx_path)
        if status_updated > 0:
            print(f"Updated status to 'published' for {status_updated} row(s).")
        if published_date_updated > 0:
            print(f"Set published_date for {published_date_updated} row(s).")
        if work_files_status_updated > 0:
            print(f"Updated {args.work_files_sheet} status to 'published' for {work_files_status_updated} row(s).")
        if work_files_published_date_updated > 0:
            print(f"Set {args.work_files_sheet} published_date for {work_files_published_date_updated} row(s).")
        if work_links_status_updated > 0:
            print(f"Updated {args.work_links_sheet} status to 'published' for {work_links_status_updated} row(s).")
        if work_links_published_date_updated > 0:
            print(f"Set {args.work_links_sheet} published_date for {work_links_published_date_updated} row(s).")
        if work_dimensions_updated > 0:
            print(f"Updated work width_px/height_px for {work_dimensions_updated} row(s).")
    if run_work_processing:
        print(
            f"\nDone. {'Would write' if not args.write else 'Wrote'}: "
            f"{written} works. Skipped: {skipped} works."
        )
        print(
            f"Downloads {'to copy' if not args.write else 'copied'}: {downloads_copied}."
            f" Missing/unresolved: {downloads_missing}."
        )
        print(f"Workbook: {display_path(xlsx_path)}")
        if args.write:
            print("Note: if the workbook is open in Excel, close and reopen it to see changes.")
    else:
        print("Work pages/files skipped: not selected by --only.")

    # Determine series scope for this run:
    # - If caller explicitly scoped series via --series-ids, honor that.
    # - If caller scoped only works (--work-ids/--work-ids-file), skip series pages by default.
    series_page_selected_ids = selected_series_ids
    if explicit_work_filter and selected_series_ids is None:
        if selected_artifacts is None:
            run_series_pages = False

    # ----------------------------
    # Series page generation (Series)
    # ----------------------------
    # Series worksheet required columns:
    # - series_id
    # - title
    # Optional columns:
    # - year_display (preferred display value)
    # - year (numeric; also fallback for display when year_display column absent)

    series_publish_transitions: List[Dict[str, Any]] = []

    if not series_rows or len(series_rows) < 2:
        print("No series pages to generate (Series sheet empty).")
    else:
        def is_actionable_series_status(status_value: str) -> bool:
            if status_value == "draft":
                return True
            if status_value == "published" and args.force:
                return True
            return False

        series_written = 0
        series_skipped = 0
        series_json_written = 0
        series_json_skipped = 0
        series_status_updated = 0
        series_published_date_updated = 0
        series_published_date_idx = series_hi.get("published_date")
        series_published_date_missing_warned = False
        studio_series_written = 0
        studio_series_skipped = 0
        tag_assignments_payload = load_tag_assignments_payload(tag_assignments_path)
        tag_assignments_series = tag_assignments_payload.get("series", {})
        tag_assignments_changed = False
        tag_assignments_added = 0
        s_total = 0
        for sr in series_rows[1:]:
            sid_raw = cell(sr, series_hi, "series_id")
            if is_empty(sid_raw):
                continue
            sid = normalize_series_id(sid_raw)
            if series_page_selected_ids is not None and sid not in series_page_selected_ids:
                continue
            status = normalize_status(cell(sr, series_hi, "status"))
            if is_actionable_series_status(status):
                s_total += 1
        s_processed = 0

        if run_series_pages:
            for sr, sr_cells in zip(series_rows[1:], series_ws.iter_rows(min_row=2), strict=False):
                sid_raw = cell(sr, series_hi, "series_id")
                if is_empty(sid_raw):
                    series_skipped += 1
                    continue
                series_id = normalize_series_id(sid_raw)
                if series_page_selected_ids is not None and series_id not in series_page_selected_ids:
                    series_skipped += 1
                    continue

                status = normalize_status(cell(sr, series_hi, "status"))
                if not is_actionable_series_status(status):
                    series_skipped += 1
                    continue

                s_processed += 1
                prefix_s = f"[series {s_processed}/{s_total}] "

                title_raw = cell(sr, series_hi, "title")
                series_title = coerce_string(title_raw) or series_id

                # Numeric year (optional)
                year = coerce_int(cell(sr, series_hi, "year")) if "year" in series_hi else None

                # year_display handling:
                # - If Series sheet has a year_display column, use it (may be null).
                # - If it does NOT have year_display, fall back to numeric year rendered as text
                year_display: Optional[str]
                if "year_display" in series_hi:
                    year_display = coerce_string(cell(sr, series_hi, "year_display"))
                else:
                    # Fall back to numeric year rendered as text
                    year_display = str(year) if year is not None else None

                series_work_ids_sorted = sorted(work_ids_by_series_all.get(series_id, []))
                primary_work_id = require_series_primary_work_id(
                    series_id,
                    sr,
                    ordered_work_ids=series_work_ids_sorted,
                )
                published_date = parse_date(cell(sr, series_hi, "published_date")) if "published_date" in series_hi else None
                series_record = compact_json_object({
                    "series_id": series_id,
                    "layout": "series",
                    "status": status,
                    "published_date": published_date,
                    "title": series_title,
                    "sort_fields": ",".join(series_sort_fields_by_series_id.get(series_id, ["work_id"])),
                    "series_type": coerce_string(cell(sr, series_hi, "series_type")) if "series_type" in series_hi else None,
                    "year": year,
                    "year_display": year_display,
                    "notes": coerce_string(cell(sr, series_hi, "notes")) if "notes" in series_hi else None,
                    "project_folders": series_project_folders_by_id.get(series_id, []),
                    "works": series_work_ids_sorted,
                    "primary_work_id": primary_work_id,
                })

                public_series_record = build_series_json_record(series_record)
                source_prose_path = resolve_series_prose_source_path(
                    series_id,
                    sr,
                    ordered_work_ids=series_work_ids_sorted,
                )
                content_html: Optional[str] = None
                if source_prose_path is not None:
                    if source_prose_path.exists():
                        content_html = render_markdown_with_jekyll(source_prose_path)
                    else:
                        print(f"[series {series_id}] WARNING: missing source prose {display_projects_path(source_prose_path)}; omitting prose content.")

                payload_version = compute_payload_version(
                    compact_json_object({"series": public_series_record, "content_html": content_html})
                )
                sfm: Dict[str, Any] = {
                    "series_id": series_id,
                    "title": series_title,
                    "layout": "series",
                    "checksum": compute_work_checksum(series_record),
                }
                series_content = build_front_matter(sfm) + "\n"

                series_path = series_out_dir / f"{series_id}.md"
                existing_text = None
                if series_path.exists():
                    try:
                        existing_text = series_path.read_text(encoding="utf-8")
                    except Exception:
                        existing_text = None

                needs_write = (existing_text != series_content)
                if (not needs_write) and (not args.force):
                    print(f"{prefix_s}SKIP (no change): {display_path(series_path)}")
                    series_skipped += 1
                else:
                    if args.write:
                        series_path.write_text(series_content, encoding="utf-8")
                        print(f"{prefix_s}WRITE: {display_path(series_path)}")
                        series_written += 1
                        status_idx = series_hi.get("status")
                        if status_idx is not None:
                            status_was = normalize_status(sr_cells[status_idx].value)
                            if status_was != "published":
                                sr_cells[status_idx].value = "published"
                                series_status_updated += 1
                            if (status_was != "published") or args.force:
                                if series_published_date_idx is not None:
                                    sr_cells[series_published_date_idx].value = today
                                    series_published_date_updated += 1
                                elif not series_published_date_missing_warned:
                                    print("Warning: Series sheet missing published_date column; skipping date updates.")
                                    series_published_date_missing_warned = True
                            if status_was != "published":
                                series_publish_transitions.append({
                                    "series_id": series_id,
                                    "title": series_title,
                                    "published_date": today.isoformat(),
                                    "primary_work_id": primary_work_id,
                                    "work_count": len(series_work_ids_sorted),
                                })
                    else:
                        print(f"{prefix_s}DRY-RUN: would write {display_path(series_path)} (overwrite={series_path.exists()})")
                        series_written += 1

                payload = compact_json_object({
                    "header": {
                        "schema": "series_record_v1",
                        "version": payload_version,
                        "generated_at_utc": utc_timestamp_now(),
                        "series_id": series_id,
                        "count": len(series_work_ids_sorted),
                    },
                    "series": public_series_record,
                    "content_html": content_html,
                })
                out_json_path = series_json_dir / f"{series_id}.json"
                out_exists = out_json_path.exists()
                existing_payload_version = extract_existing_header_scalar(out_json_path, "version") if out_exists else None
                if (existing_payload_version is not None) and (existing_payload_version == payload_version) and (not args.force):
                    series_json_skipped += 1
                else:
                    if args.write:
                        out_json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
                        print(f"[Series JSON {s_processed}/{s_total}] WRITE: {display_path(out_json_path)}")
                        series_json_written += 1
                    else:
                        print(f"[Series JSON {s_processed}/{s_total}] DRY-RUN: would write {display_path(out_json_path)} (overwrite={out_exists})")
                        series_json_written += 1

                if run_studio_series_pages:
                    if studio_series_out_dir is None:
                        raise RuntimeError("studio_series_out_dir is not initialized")
                    tsm: Dict[str, Any] = {
                        "layout": "studio_series",
                        "series_id": series_id,
                        "title": series_title,
                        "sort_fields": ",".join(series_sort_fields_by_series_id.get(series_id, ["work_id"])),
                        "year": year,
                        "year_display": year_display,
                        "project_folders": series_project_folders_by_id.get(series_id, []),
                        "notes": coerce_string(cell(sr, series_hi, "notes")) if "notes" in series_hi else None,
                        "primary_work_id": primary_work_id,
                    }
                    studio_series_content = build_front_matter(tsm) + "\n"
                    studio_series_path = studio_series_out_dir / f"{series_id}.md"

                    existing_studio_series_text = None
                    if studio_series_path.exists():
                        try:
                            existing_studio_series_text = studio_series_path.read_text(encoding="utf-8")
                        except Exception:
                            existing_studio_series_text = None

                    studio_series_needs_write = (existing_studio_series_text != studio_series_content)
                    prefix_ts = f"[studio-series {s_processed}/{s_total}] "
                    if (not studio_series_needs_write) and (not args.force):
                        print(f"{prefix_ts}SKIP (no change): {studio_series_path}")
                        studio_series_skipped += 1
                    else:
                        if args.write:
                            studio_series_path.write_text(studio_series_content, encoding="utf-8")
                            print(f"{prefix_ts}WRITE: {display_path(studio_series_path)}")
                            studio_series_written += 1
                        else:
                            print(f"{prefix_ts}DRY-RUN: would write {display_path(studio_series_path)} (overwrite={studio_series_path.exists()})")
                            studio_series_written += 1

                if series_id not in tag_assignments_series:
                    tag_assignments_series[series_id] = {
                        "tags": [],
                        "works": {},
                        "updated_at_utc": utc_timestamp_now(),
                    }
                    tag_assignments_changed = True
                    tag_assignments_added += 1
                else:
                    assignment_row = tag_assignments_series.get(series_id)
                    if not isinstance(assignment_row, dict):
                        tag_assignments_series[series_id] = {
                            "tags": [],
                            "works": {},
                            "updated_at_utc": utc_timestamp_now(),
                        }
                        tag_assignments_changed = True
                    else:
                        if not isinstance(assignment_row.get("tags"), list):
                            assignment_row["tags"] = []
                            tag_assignments_changed = True
                        if "works" not in assignment_row or not isinstance(assignment_row.get("works"), dict):
                            assignment_row["works"] = {}
                            tag_assignments_changed = True
        else:
            if selected_artifacts is not None and not artifact_enabled("series-pages"):
                print("Series pages skipped: not selected by --only.")
            else:
                print("Series pages skipped: --work-ids scope active (use --series-ids to include series page rebuild).")
            print("Studio series pages retired: skipped.")
            print("Tag assignments sync skipped: follows series-pages selection.")

        if run_series_pages:
            if tag_assignments_changed:
                tag_assignments_payload["series"] = tag_assignments_series
                tag_assignments_payload["updated_at_utc"] = utc_timestamp_now()
                tag_assignments_text = json.dumps(tag_assignments_payload, indent=2, ensure_ascii=False) + "\n"
                if args.write:
                    tag_assignments_path.write_text(tag_assignments_text, encoding="utf-8")
                    print(
                        f"Tag assignments sync: WRITE {display_path(tag_assignments_path)} "
                        f"(added missing entries: {tag_assignments_added})."
                    )
                else:
                    print(
                        f"Tag assignments sync: DRY-RUN would write {display_path(tag_assignments_path)} "
                        f"(added missing entries: {tag_assignments_added})."
                    )
            else:
                print("Tag assignments sync: no missing series entries.")

        if args.write and (series_status_updated > 0 or series_published_date_updated > 0):
            assert wb_write is not None
            wb_write.save(xlsx_path)
            if series_status_updated > 0:
                print(f"Updated series status to 'published' for {series_status_updated} row(s).")
            if series_published_date_updated > 0:
                print(f"Set series published_date for {series_published_date_updated} row(s).")
        print(f"Series pages done. {'Would write' if not args.write else 'Wrote'}: {series_written}. Skipped: {series_skipped}.")
        print(
            f"Series JSON done. {'Would write' if not args.write else 'Wrote'}: "
            f"{series_json_written}. Skipped: {series_json_skipped}."
        )
        if run_studio_series_pages:
            print(
                f"Studio series pages done. {'Would write' if not args.write else 'Wrote'}: "
                f"{studio_series_written}. Skipped: {studio_series_skipped}."
            )
        else:
            print("Studio series pages retired: skipped.")

    work_rows_by_series_for_index: Dict[str, List[tuple[str, str]]] = {}
    for wr in works_rows[1:]:
        wid_raw = cell(wr, works_hi, "work_id")
        if is_empty(wid_raw):
            continue
        status = normalize_status(cell(wr, works_hi, "status"))
        if status not in {"draft", "published"}:
            continue
        wid = slug_id(wid_raw)
        for sid in parse_work_series_ids(wr):
            series_sort = series_sort_by_series_id.get(sid, {}).get(wid, wid)
            work_rows_by_series_for_index.setdefault(sid, []).append((series_sort, wid))

    ordered_work_ids_by_series_for_index: Dict[str, List[str]] = {}
    for sid, rows in work_rows_by_series_for_index.items():
        rows_sorted = sorted(rows, key=lambda item: (item[0], item[1]))
        ordered_work_ids_by_series_for_index[sid] = [wid for _, wid in rows_sorted]

    series_payload_unsorted: Dict[str, Dict[str, Any]] = {}
    for sr in series_rows[1:] if len(series_rows) > 1 else []:
        sid_raw = cell(sr, series_hi, "series_id")
        if is_empty(sid_raw):
            continue
        sid = normalize_series_id(sid_raw)
        status = normalize_status(cell(sr, series_hi, "status"))
        if status not in {"draft", "published"}:
            continue

        title_raw = cell(sr, series_hi, "title")
        series_title = coerce_string(title_raw) or sid
        year = coerce_int(cell(sr, series_hi, "year")) if "year" in series_hi else None
        if "year_display" in series_hi:
            year_display = coerce_string(cell(sr, series_hi, "year_display"))
        else:
            year_display = str(year) if year is not None else None
        published_date = parse_date(cell(sr, series_hi, "published_date")) if "published_date" in series_hi else None

        ordered_work_ids = ordered_work_ids_by_series_for_index.get(sid, [])
        primary_work_id = require_series_primary_work_id(
            sid,
            sr,
            ordered_work_ids=ordered_work_ids,
        )

        sort_fields = ",".join(series_sort_fields_by_series_id.get(sid, ["work_id"]))
        series_payload_unsorted[sid] = compact_json_object({
            "series_id": sid,
            "layout": "series",
            "status": status,
            "published_date": published_date,
            "title": series_title,
            "sort_fields": sort_fields,
            "series_type": coerce_string(cell(sr, series_hi, "series_type")) if "series_type" in series_hi else None,
            "year": year,
            "year_display": year_display,
            "primary_work_id": primary_work_id,
            "notes": coerce_string(cell(sr, series_hi, "notes")) if "notes" in series_hi else None,
            "project_folders": series_project_folders_by_id.get(sid, []),
            "works": ordered_work_ids,
        })

    series_payload: Dict[str, Dict[str, Any]] = {
        sid: series_payload_unsorted[sid] for sid in sorted(series_payload_unsorted.keys())
    }

    series_version_payload = compact_json_object({
        "schema": "series_index_v2",
        "series": series_payload,
    })
    series_version = compute_payload_version(series_version_payload)
    series_index_payload = compact_json_object({
        "header": {
            "schema": "series_index_v2",
            "version": series_version,
            "generated_at_utc": utc_timestamp_now(),
            "count": len(series_payload),
        },
        "series": series_payload,
    })

    exists = series_index_json_path.exists()
    existing_version = extract_existing_header_scalar(series_index_json_path, "version") if exists else None
    if (existing_version is not None) and (existing_version == series_version) and (not args.force):
        print("Series index JSON done. Wrote: 0. Skipped: 1.")
    else:
        if args.write:
            series_index_json_path.write_text(
                json.dumps(series_index_payload, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            print(f"Series index JSON done. Wrote: 1. Skipped: 0. Path: {display_path(series_index_json_path)}")
        else:
            print(
                "Series index JSON done. Would write: 1. Skipped: 0. "
                f"Path: {display_path(series_index_json_path)} (overwrite={exists})"
            )

    # ----------------------------
    # Work detail page generation + per-work detail JSON (WorkDetails)
    # ----------------------------
    if not work_details_rows or len(work_details_rows) < 2 or work_details_ws is None:
        if run_work_details_pages or run_work_json or run_works_index_json:
            print("No work detail pages/JSON/index rows found (WorkDetails sheet empty or missing).")
        else:
            print("Work detail pages/JSON skipped: not selected by --only.")
    else:
        required_details = ["work_id", "detail_id", "title", "status", "project_subfolder", "project_filename"]
        missing_details = [c for c in required_details if c not in work_details_hi]
        if missing_details:
            raise SystemExit(f"WorkDetails sheet missing required columns: {', '.join(missing_details)}")

        projects_base_dir = Path(args.projects_base_dir).expanduser()
        projects_root = projects_base_dir / source_works_root_subdir(PIPELINE_CONFIG)

        # Ensure width/height columns exist when writing so dimensions persist in the sheet.
        width_px_idx = work_details_hi.get("width_px")
        height_px_idx = work_details_hi.get("height_px")
        if args.write and run_work_details_pages:
            if width_px_idx is None:
                width_px_idx = work_details_ws.max_column
                work_details_ws.cell(row=1, column=width_px_idx + 1, value="width_px")
                work_details_hi["width_px"] = width_px_idx
            if height_px_idx is None:
                height_px_idx = work_details_ws.max_column
                work_details_ws.cell(row=1, column=height_px_idx + 1, value="height_px")
                work_details_hi["height_px"] = height_px_idx

        # Build known works from the Works sheet to validate foreign-key references.
        known_work_ids: set[str] = set()
        for wr in works_rows[1:]:
            wid_raw = cell(wr, works_hi, "work_id")
            if is_empty(wid_raw):
                continue
            known_work_ids.add(slug_id(wid_raw))

        def is_actionable_detail_status(status_value: str) -> bool:
            if status_value == "draft":
                return True
            if status_value == "published" and args.force:
                return True
            return False

        if run_work_details_pages:
            details_written = 0
            details_skipped = 0
            details_status_updated = 0
            details_published_date_updated = 0
            details_dimensions_updated = 0
            details_published_date_idx = work_details_hi.get("published_date")
            details_published_date_missing_warned = False
            project_folder_missing_warned = False
            details_total = 0

            for dr in work_details_rows[1:]:
                wid_raw = cell(dr, work_details_hi, "work_id")
                if is_empty(wid_raw):
                    continue
                wid = slug_id(wid_raw)
                if selected_ids is not None and wid not in selected_ids:
                    continue
                status = normalize_status(cell(dr, work_details_hi, "status"))
                if is_actionable_detail_status(status):
                    details_total += 1

            details_processed = 0
            for dr, dr_cells in zip(work_details_rows[1:], work_details_ws.iter_rows(min_row=2), strict=False):
                wid_raw = cell(dr, work_details_hi, "work_id")
                did_raw = cell(dr, work_details_hi, "detail_id")
                if is_empty(wid_raw) or is_empty(did_raw):
                    details_skipped += 1
                    continue

                wid = slug_id(wid_raw)
                if selected_ids is not None and wid not in selected_ids:
                    details_skipped += 1
                    continue

                if wid not in known_work_ids:
                    print(f"Warning: skipping work detail for unknown work_id {wid}: detail_id={did_raw}")
                    details_skipped += 1
                    continue

                status = normalize_status(cell(dr, work_details_hi, "status"))
                if not is_actionable_detail_status(status):
                    details_skipped += 1
                    continue

                details_processed += 1
                prefix_d = f"[details {details_processed}/{details_total}] "

                did = slug_id(did_raw, width=3)
                detail_uid = f"{wid}-{did}"
                title = coerce_string(cell(dr, work_details_hi, "title"))
                project_subfolder = coerce_string(cell(dr, work_details_hi, "project_subfolder"))
                project_filename = coerce_string(cell(dr, work_details_hi, "project_filename"))
                width_px = coerce_int(cell(dr, work_details_hi, "width_px")) if "width_px" in work_details_hi else None
                height_px = coerce_int(cell(dr, work_details_hi, "height_px")) if "height_px" in work_details_hi else None

                # Resolve source image and persist dimensions back to WorkDetails for stable future rebuilds.
                project_folder = work_project_folder_by_id.get(wid)
                src_path: Optional[Path] = None
                if project_folder and project_filename:
                    src_path = projects_root / project_folder
                    if project_subfolder:
                        src_path = src_path / project_subfolder
                    src_path = src_path / project_filename
                elif not project_folder_missing_warned:
                    if not has_project_folder_col:
                        print("Warning: Works sheet has no project_folder column; cannot persist WorkDetails image dimensions.")
                    else:
                        print("Warning: missing Works.project_folder for one or more WorkDetails rows; cannot persist those image dimensions.")
                    project_folder_missing_warned = True

                if src_path is not None:
                    src_w, src_h = read_image_dims_px(src_path)
                    if src_w is not None and src_h is not None:
                        width_px = src_w
                        height_px = src_h
                        if args.write and width_px_idx is not None and height_px_idx is not None:
                            prev_w = dr_cells[width_px_idx].value if width_px_idx < len(dr_cells) else None
                            prev_h = dr_cells[height_px_idx].value if height_px_idx < len(dr_cells) else None
                            if prev_w != src_w or prev_h != src_h:
                                dr_cells[width_px_idx].value = src_w
                                dr_cells[height_px_idx].value = src_h
                                details_dimensions_updated += 1
                    else:
                        print(f"Warning: could not read dimensions for detail source image: {display_projects_path(src_path)}")
                elif project_filename:
                    print(f"Warning: could not resolve detail source image path for {detail_uid} ({project_filename})")

                # Canonical detail metadata stays in per-work JSON/index artifacts.
                # Keep _work_details pages minimal (routing identifiers + title only).
                dfm: Dict[str, Any] = {
                    "work_id": wid,
                    "detail_id": did,
                    "detail_uid": detail_uid,
                    "title": title,
                }

                d_content = build_front_matter(dfm)
                d_path = work_details_out_dir / f"{detail_uid}.md"
                d_exists = d_path.exists()
                if d_exists and not args.force:
                    existing_content = d_path.read_text(encoding="utf-8")
                    if existing_content == d_content:
                        details_skipped += 1
                        continue

                if args.write:
                    d_path.write_text(d_content, encoding="utf-8")
                    print(f"{prefix_d}WRITE: {display_path(d_path)}")
                    details_written += 1

                    status_idx = work_details_hi.get("status")
                    if status_idx is not None:
                        status_was = normalize_status(dr_cells[status_idx].value)
                        if status_was != "published":
                            dr_cells[status_idx].value = "published"
                            details_status_updated += 1
                        if (status_was != "published") or args.force:
                            if details_published_date_idx is not None:
                                dr_cells[details_published_date_idx].value = today
                                details_published_date_updated += 1
                            elif not details_published_date_missing_warned:
                                print("Warning: WorkDetails sheet missing published_date column; skipping date updates.")
                                details_published_date_missing_warned = True
                else:
                    print(f"{prefix_d}DRY-RUN: would write {display_path(d_path)} (overwrite={d_exists})")
                    details_written += 1

            if args.write and (details_status_updated > 0 or details_published_date_updated > 0 or details_dimensions_updated > 0):
                assert wb_write is not None
                wb_write.save(xlsx_path)
                if details_status_updated > 0:
                    print(f"Updated work detail status to 'published' for {details_status_updated} row(s).")
                if details_published_date_updated > 0:
                    print(f"Set work detail published_date for {details_published_date_updated} row(s).")
                if details_dimensions_updated > 0:
                    print(f"Updated work detail width_px/height_px for {details_dimensions_updated} row(s).")

            print(
                f"Work detail pages done. {'Would write' if not args.write else 'Wrote'}: {details_written}. Skipped: {details_skipped}."
            )
        else:
            print("Work detail pages skipped: not selected by --only.")

        if run_work_json:
            # Build per-work JSON from Works rows (work-driven).
            # Detail sections are sourced from currently published detail rows only.
            encountered_work_ids: List[str] = []
            encountered_work_id_set: set[str] = set()
            detail_records_by_work: Dict[str, List[Dict[str, Any]]] = {}
            detail_status_idx = work_details_hi.get("status")

            for wr in works_rows[1:]:
                wid_raw = cell(wr, works_hi, "work_id")
                if is_empty(wid_raw):
                    continue
                wid = slug_id(wid_raw)
                if selected_ids is not None and wid not in selected_ids:
                    continue
                status = normalize_status(cell(wr, works_hi, "status"))
                if status not in {"draft", "published"}:
                    continue
                if wid not in canonical_work_record_by_id:
                    continue
                if wid not in encountered_work_id_set:
                    encountered_work_ids.append(wid)
                    encountered_work_id_set.add(wid)

            for dr, dr_cells in zip(work_details_rows[1:], work_details_ws.iter_rows(min_row=2), strict=False):
                wid_raw = cell(dr, work_details_hi, "work_id")
                did_raw = cell(dr, work_details_hi, "detail_id")
                if is_empty(wid_raw) or is_empty(did_raw):
                    continue

                wid = slug_id(wid_raw)
                if wid not in encountered_work_id_set:
                    continue

                status_val = dr_cells[detail_status_idx].value if detail_status_idx is not None else cell(dr, work_details_hi, "status")
                if normalize_status(status_val) != "published":
                    continue

                did = slug_id(did_raw, width=3)
                detail_record = build_canonical_detail_record(
                    wid=wid,
                    did=did,
                    title=coerce_string(read_cell_value(dr, dr_cells, work_details_hi, "title")),
                    project_subfolder=coerce_string(read_cell_value(dr, dr_cells, work_details_hi, "project_subfolder")),
                    width_px=coerce_int(dr_cells[work_details_hi["width_px"]].value) if "width_px" in work_details_hi and work_details_hi["width_px"] < len(dr_cells) else None,
                    height_px=coerce_int(dr_cells[work_details_hi["height_px"]].value) if "height_px" in work_details_hi and work_details_hi["height_px"] < len(dr_cells) else None,
                )
                detail_records_by_work.setdefault(wid, []).append(detail_record)

            wj_written = 0
            wj_skipped = 0
            wj_total = len(encountered_work_ids)
            wj_processed = 0
            generated_at_utc = utc_timestamp_now()

            for wid in encountered_work_ids:
                wj_processed += 1
                prefix_wj = f"[Work JSON {wj_processed}/{wj_total}] "

                source_prose_path = resolve_work_prose_source_path(wid)

                sections = build_sections_from_detail_records(detail_records_by_work.get(wid, []))
                details_total = sum(len(s.get("details", [])) for s in sections)
                work_record = build_work_json_record(canonical_work_record_by_id.get(wid, {"work_id": wid}))
                content_html: Optional[str] = None
                if source_prose_path is not None and source_prose_path.exists():
                    content_html = render_markdown_with_jekyll(source_prose_path)
                payload_version = compute_payload_version(compact_json_object({"work": work_record, "sections": sections, "content_html": content_html}))

                payload = compact_json_object({
                    "header": {
                        "schema": "work_record_v3",
                        "version": payload_version,
                        "generated_at_utc": generated_at_utc,
                        "work_id": wid,
                        "count": details_total,
                    },
                    "work": work_record,
                    "sections": sections,
                    "content_html": content_html,
                })
                out_json_path = works_json_dir / f"{wid}.json"
                exists = out_json_path.exists()
                existing_version = extract_existing_header_scalar(out_json_path, "version") if exists else None
                payload_version = payload["header"]["version"]

                if (existing_version is not None) and (existing_version == payload_version) and (not args.force):
                    wj_skipped += 1
                    continue

                if args.write:
                    out_json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
                    print(f"{prefix_wj}WRITE: {display_path(out_json_path)}")
                    wj_written += 1
                else:
                    print(f"{prefix_wj}DRY-RUN: would write {display_path(out_json_path)} (overwrite={exists})")
                    wj_written += 1

            print(
                f"Work JSON done. {'Would write' if not args.write else 'Wrote'}: {wj_written}. Skipped: {wj_skipped}."
            )
        else:
            print("Work detail JSON skipped: not selected by --only.")

    works_payload: Dict[str, Dict[str, Any]] = {}
    for wr in works_rows[1:]:
        wid_raw = cell(wr, works_hi, "work_id")
        if is_empty(wid_raw):
            continue
        status = normalize_status(cell(wr, works_hi, "status"))
        if status not in {"draft", "published"}:
            continue
        wid = slug_id(wid_raw)
        record = canonical_work_record_by_id.get(wid)
        if record is None:
            continue
        works_payload[wid] = build_work_index_record(record)

    work_storage_payload: Dict[str, Dict[str, Any]] = {}
    for wid, record in works_payload.items():
        storage_record = build_work_storage_index_record(canonical_work_record_by_id.get(wid, {}))
        if storage_record is not None:
            work_storage_payload[wid] = storage_record

    detail_records_by_work: Dict[str, List[Dict[str, Any]]] = {}
    if work_details_rows and len(work_details_rows) > 1:
        for dr, dr_cells in zip(work_details_rows[1:], work_details_ws.iter_rows(min_row=2), strict=False):
            wid_raw = cell(dr, work_details_hi, "work_id")
            did_raw = cell(dr, work_details_hi, "detail_id")
            if is_empty(wid_raw) or is_empty(did_raw):
                continue
            wid = slug_id(wid_raw)
            if wid not in works_payload:
                continue
            status = normalize_status(dr_cells[work_details_hi["status"]].value if work_details_hi["status"] < len(dr_cells) else cell(dr, work_details_hi, "status"))
            if status not in {"draft", "published"}:
                continue
            did = slug_id(did_raw, width=3)
            detail_records_by_work.setdefault(wid, []).append(
                build_canonical_detail_record(
                    wid=wid,
                    did=did,
                    title=coerce_string(read_cell_value(dr, dr_cells, work_details_hi, "title")),
                    project_subfolder=coerce_string(read_cell_value(dr, dr_cells, work_details_hi, "project_subfolder")),
                    width_px=coerce_int(dr_cells[work_details_hi["width_px"]].value) if "width_px" in work_details_hi and work_details_hi["width_px"] < len(dr_cells) else None,
                    height_px=coerce_int(dr_cells[work_details_hi["height_px"]].value) if "height_px" in work_details_hi and work_details_hi["height_px"] < len(dr_cells) else None,
                )
            )

    version_payload = compact_json_object({
        "schema": "works_index_v4",
        "works": works_payload,
    })
    version = compute_payload_version(version_payload)
    payload = compact_json_object({
        "header": {
            "schema": "works_index_v4",
            "version": version,
            "generated_at_utc": utc_timestamp_now(),
            "count": len(works_payload),
        },
        "works": works_payload,
    })
    payload_version = payload["header"]["version"]
    exists = works_index_json_path.exists()
    existing_version = extract_existing_header_scalar(works_index_json_path, "version") if exists else None
    if (existing_version is not None) and (existing_version == payload_version) and (not args.force):
        print("Works index JSON done. Wrote: 0. Skipped: 1.")
    else:
        if args.write:
            works_index_json_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            print(f"Works index JSON done. Wrote: 1. Skipped: 0. Path: {display_path(works_index_json_path)}")
        else:
            print(
                "Works index JSON done. Would write: 1. Skipped: 0. "
                f"Path: {display_path(works_index_json_path)} (overwrite={exists})"
            )

    recent_entries_existing = load_recent_entries(recent_index_json_path)
    current_work_ids = set(works_payload.keys())
    current_series_ids = set(series_payload.keys())
    retained_recent_entries = [
        entry for entry in recent_entries_existing
        if (
            (entry.get("kind") == "work" and str(entry.get("target_id") or "") in current_work_ids)
            or (entry.get("kind") == "series" and str(entry.get("target_id") or "") in current_series_ids)
        )
    ]

    recorded_at_utc = utc_timestamp_now()
    recent_entries_new: List[Dict[str, Any]] = []
    newly_published_series_ids = {
        str(entry.get("series_id") or "")
        for entry in series_publish_transitions
        if str(entry.get("series_id") or "")
    }

    session_order = 0
    for entry in series_publish_transitions:
        series_id = str(entry.get("series_id") or "")
        if not series_id:
            continue
        session_order += 1
        normalized = normalize_recent_entry({
            "kind": "series",
            "target_id": series_id,
            "title": coerce_string(entry.get("title")) or series_id,
            "caption": f"{int(entry.get('work_count') or 0)} work{'s' if int(entry.get('work_count') or 0) != 1 else ''}",
            "published_date": entry.get("published_date"),
            "thumb_id": coerce_string(entry.get("primary_work_id")),
            "recorded_at_utc": recorded_at_utc,
            "session_order": session_order,
        })
        if normalized is not None:
            recent_entries_new.append(normalized)

    grouped_work_transitions: Dict[str, List[Dict[str, Any]]] = {}
    for entry in work_publish_transitions:
        primary_series_id = str(entry.get("primary_series_id") or "")
        if not primary_series_id or primary_series_id in newly_published_series_ids:
            continue
        grouped_work_transitions.setdefault(primary_series_id, []).append(entry)

    for series_id, entries in sorted(grouped_work_transitions.items()):
        entries_sorted = sorted(
            entries,
            key=lambda item: (
                str(series_sort_by_series_id.get(series_id, {}).get(str(item.get("work_id") or ""), str(item.get("work_id") or ""))),
                str(item.get("work_id") or ""),
            ),
        )
        first_entry = entries_sorted[0] if entries_sorted else {}
        if not first_entry:
            continue
        new_work_count = len(entries_sorted)
        series_title = coerce_string(first_entry.get("series_title")) or series_title_by_id.get(series_id) or series_id
        caption = series_title if new_work_count == 1 else f"{new_work_count} new works in {series_title}"
        session_order += 1
        normalized = normalize_recent_entry({
            "kind": "work",
            "target_id": str(first_entry.get("work_id") or ""),
            "title": coerce_string(first_entry.get("title")) or str(first_entry.get("work_id") or ""),
            "caption": caption,
            "published_date": first_entry.get("published_date"),
            "thumb_id": str(first_entry.get("work_id") or ""),
            "recorded_at_utc": recorded_at_utc,
            "session_order": session_order,
        })
        if normalized is not None:
            recent_entries_new.append(normalized)

    existing_by_id = {
        str(entry.get("id") or f"{entry.get('kind')}:{entry.get('target_id')}"): entry
        for entry in retained_recent_entries
    }
    for entry in recent_entries_new:
        existing_by_id[str(entry.get("id") or "")] = entry
    recent_index_payload = build_recent_index_payload(list(existing_by_id.values()))
    recent_exists = recent_index_json_path.exists()
    recent_existing_version = extract_existing_header_scalar(recent_index_json_path, "version") if recent_exists else None
    recent_payload_version = recent_index_payload["header"]["version"]
    if (recent_existing_version is not None) and (recent_existing_version == recent_payload_version) and (not args.force):
        print("Recent index JSON done. Wrote: 0. Skipped: 1.")
    else:
        if args.write:
            recent_index_json_path.write_text(
                json.dumps(recent_index_payload, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            print(f"Recent index JSON done. Wrote: 1. Skipped: 0. Path: {display_path(recent_index_json_path)}")
        else:
            print(
                "Recent index JSON done. Would write: 1. Skipped: 0. "
                f"Path: {display_path(recent_index_json_path)} (overwrite={recent_exists})"
            )

    work_storage_version_payload = compact_json_object({
        "schema": "work_storage_index_v1",
        "works": work_storage_payload,
    })
    work_storage_version = compute_payload_version(work_storage_version_payload)
    work_storage_payload_out = compact_json_object({
        "header": {
            "schema": "work_storage_index_v1",
            "version": work_storage_version,
            "generated_at_utc": utc_timestamp_now(),
            "count": len(work_storage_payload),
        },
        "works": work_storage_payload,
    })
    work_storage_payload_version = work_storage_payload_out["header"]["version"]
    work_storage_exists = work_storage_index_json_path.exists()
    existing_work_storage_version = extract_existing_header_scalar(work_storage_index_json_path, "version") if work_storage_exists else None
    if (existing_work_storage_version is not None) and (existing_work_storage_version == work_storage_payload_version) and (not args.force):
        print("Work storage index JSON done. Wrote: 0. Skipped: 1.")
    else:
        if args.write:
            work_storage_index_json_path.write_text(
                json.dumps(work_storage_payload_out, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            print(f"Work storage index JSON done. Wrote: 1. Skipped: 0. Path: {display_path(work_storage_index_json_path)}")
        else:
            print(
                "Work storage index JSON done. Would write: 1. Skipped: 0. "
                f"Path: {display_path(work_storage_index_json_path)} (overwrite={work_storage_exists})"
            )

    # ----------------------------
    # Moment page + JSON generation (Moments)
    # ----------------------------
    # Required columns:
    # - moment_id, title, status, published_date, date, date_display,
    #   width_px, height_px
    # Optional columns:
    # - moment_id / slug / id (preferred filename stem)
    # - date, date_display
    # - image_file/image_filename/project_filename
    # - image_alt
    # - width_px, height_px
    # - project_folder, project_subfolder, project_filename, work_id (for source image resolution)
    moment_source_records: List[Dict[str, Any]] = []
    projects_base_dir = Path(args.projects_base_dir).expanduser()
    moments_root = projects_base_dir / source_moments_root_subdir(PIPELINE_CONFIG)
    moments_images_root = projects_base_dir / source_moments_images_subdir(PIPELINE_CONFIG)
    source_moment_index = scan_moment_source_files(moments_root)
    if not run_moments and not run_moments_index_json:
        if selected_artifacts is not None and not run_moments_artifact and not run_moments_index_json:
            print("Moment pages/JSON skipped: not selected by --only.")
        else:
            print("Moment pages/JSON skipped: scoped run without --moment-ids/--moment-ids-file.")
    elif not moment_sources_manifest_records and not source_moment_index:
        print("No moment artifacts to generate (no moment source files found).")
    else:
        def collect_moment_source_records() -> List[Dict[str, Any]]:
            records: List[Dict[str, Any]] = []
            if moment_sources_manifest_records:
                for moment_id in sorted(moment_sources_manifest_records.keys()):
                    manifest_record = moment_sources_manifest_records[moment_id]
                    records.append({
                        "moment_id": moment_id,
                        "title": coerce_string(manifest_record.get("title")) or moment_id,
                        "status": normalize_status(manifest_record.get("status")),
                        "published_date": parse_date(manifest_record.get("published_date")),
                        "date": parse_date(manifest_record.get("date")),
                        "date_display": coerce_string(manifest_record.get("date_display")),
                        "width_px": coerce_int(manifest_record.get("width_px")),
                        "height_px": coerce_int(manifest_record.get("height_px")),
                        "image_file": coerce_string(manifest_record.get("image_file")) or f"{moment_id}.jpg",
                        "source_image_file": coerce_string(manifest_record.get("source_image_file")) or f"{moment_id}.jpg",
                        "image_alt": coerce_string(manifest_record.get("image_alt")),
                        "source_prose_path": Path(coerce_string(manifest_record.get("source_prose_path")) or (moments_root / f"{moment_id}.md")),
                        "source_image_path": Path(coerce_string(manifest_record.get("source_image_path")) or (moments_images_root / (coerce_string(manifest_record.get('source_image_file')) or f"{moment_id}.jpg"))),
                    })
                return records

            for moment_id in sorted(source_moment_index.keys()):
                source_record = source_moment_index[moment_id]
                records.append({
                    "moment_id": moment_id,
                    "title": coerce_string(source_record.get("title")) or moment_id,
                    "status": normalize_status(source_record.get("status")),
                    "published_date": parse_date(source_record.get("published_date")),
                    "date": parse_date(source_record.get("date")),
                    "date_display": coerce_string(source_record.get("date_display")),
                    "width_px": None,
                    "height_px": None,
                    "image_file": f"{moment_id}.jpg",
                    "source_image_file": coerce_string(source_record.get("source_image_file")) or f"{moment_id}.jpg",
                    "image_alt": coerce_string(source_record.get("image_alt")),
                    "source_prose_path": Path(coerce_string(source_record.get("source_prose_path")) or (moments_root / f"{moment_id}.md")),
                    "source_image_path": moments_images_root / ((coerce_string(source_record.get("source_image_file")) or f"{moment_id}.jpg")),
                })
            return records

        moment_source_records = collect_moment_source_records()

        def is_actionable_moment_status(status_value: str) -> bool:
            if status_value == "draft":
                return True
            if status_value == "published" and args.force:
                return True
            return False

        moments_pages_total = 0
        moments_json_total = 0
        for moment_entry in moment_source_records:
            mid = str(moment_entry.get("moment_id") or "").strip().lower()
            if not mid:
                continue
            if selected_moment_ids is not None and mid not in selected_moment_ids:
                continue
            status = normalize_status(moment_entry.get("status"))
            if run_moments and is_actionable_moment_status(status):
                moments_pages_total += 1
                moments_json_total += 1

        moments_pages_written = 0
        moments_pages_skipped = 0
        moments_json_written = 0
        moments_json_skipped = 0
        moments_source_metadata_updated = 0
        moment_json_generated_at_utc = utc_timestamp_now()
        moments_processed = 0

        for moment_entry in moment_source_records:
            moment_id = str(moment_entry.get("moment_id") or "").strip().lower()
            if not moment_id:
                if run_moments:
                    moments_pages_skipped += 1
                    moments_json_skipped += 1
                continue
            mid = moment_id
            if selected_moment_ids is not None and mid not in selected_moment_ids:
                if run_moments:
                    moments_pages_skipped += 1
                    moments_json_skipped += 1
                continue
            status = normalize_status(moment_entry.get("status"))
            moment_actionable = run_moments and is_actionable_moment_status(status)
            if not moment_actionable and run_moments:
                moments_pages_skipped += 1
                moments_json_skipped += 1
            if not moment_actionable:
                continue

            if not is_slug_safe(moment_id):
                raise SystemExit(f"Moment source filename must be slug-safe; got: {moment_id!r}")

            moments_processed += 1
            prefix_m = f"[moment {moment_id}] "
            print(f"[moment {moments_processed}/{moments_pages_total}] {moment_id}", flush=True)

            title = coerce_string(moment_entry.get("title")) or moment_id
            date_value = parse_date(moment_entry.get("date"))
            date_display = coerce_string(moment_entry.get("date_display"))
            width_px = coerce_int(moment_entry.get("width_px"))
            height_px = coerce_int(moment_entry.get("height_px"))

            default_moment_filename = f"{moment_id}.jpg"
            image_file = coerce_string(moment_entry.get("image_file")) or default_moment_filename
            source_image_file = coerce_string(moment_entry.get("source_image_file")) or default_moment_filename
            image_alt = coerce_string(moment_entry.get("image_alt"))

            # Resolve source image for dimensions when possible.
            src_path: Optional[Path] = None
            source_filename = source_image_file
            if source_filename:
                source_image_path_value = coerce_string(moment_entry.get("source_image_path"))
                src_path = Path(source_image_path_value) if source_image_path_value else (moments_images_root / source_filename)

            source_image_exists = bool(src_path is not None and src_path.exists())

            if src_path is not None:
                src_w, src_h = read_image_dims_px(src_path)
                if src_w is not None and src_h is not None:
                    width_px = src_w
                    height_px = src_h

            images_list: List[Dict[str, Any]] = []
            if image_file is not None and image_alt is None:
                image_alt = title or moment_id
            if image_file is not None and source_image_exists:
                images_list.append(
                    {
                        "file": image_file,
                        "alt": image_alt,
                    }
                )

            moment_record: Dict[str, Any] = {
                "moment_id": moment_id,
                "title": title or moment_id,
                "date": date_value,
                "date_display": date_display,
                "images": images_list,
                "width_px": width_px,
                "height_px": height_px,
                "layout": "moment",
            }
            moment_page_fm: Dict[str, Any] = {
                "moment_id": moment_id,
                "title": title or moment_id,
                "layout": "moment",
            }
            m_checksum = compute_work_checksum(moment_record)
            moment_page_fm["checksum"] = m_checksum

            source_prose_path_value = coerce_string(moment_entry.get("source_prose_path"))
            source_prose_path = Path(source_prose_path_value) if source_prose_path_value else (moments_root / f"{moment_id}.md")
            if not source_prose_path.exists():
                print(f"{prefix_m}WARNING: missing source prose {display_projects_path(source_prose_path)}; skipping moment.")
                if moment_actionable:
                    moments_pages_skipped += 1
                    moments_json_skipped += 1
                continue

            if moment_actionable:
                # Canonical moment metadata lives in JSON artifacts; keep `_moments`
                # pages minimal and reserve front matter for routing + title fallback.
                m_content = build_front_matter(moment_page_fm) + "\n"
                m_path = moments_out_dir / f"{moment_id}.md"
                m_exists = m_path.exists()
                existing_checksum = extract_existing_checksum(m_path) if m_exists else None

                if (existing_checksum is not None) and (existing_checksum == m_checksum) and (not args.force):
                    existing_content: Optional[str] = None
                    try:
                        existing_content = m_path.read_text(encoding="utf-8")
                    except Exception:
                        existing_content = None
                    if existing_content == m_content:
                        moments_pages_skipped += 1
                    else:
                        if args.write:
                            m_path.write_text(m_content, encoding="utf-8")
                            print(f"{prefix_m}WRITE: {display_path(m_path)}")
                            moments_pages_written += 1
                        else:
                            print(f"{prefix_m}DRY-RUN: would write {display_path(m_path)} (overwrite={m_exists})")
                            moments_pages_written += 1
                else:
                    if args.write:
                        m_path.write_text(m_content, encoding="utf-8")
                        print(f"{prefix_m}WRITE: {display_path(m_path)}")
                        moments_pages_written += 1
                    else:
                        print(f"{prefix_m}DRY-RUN: would write {display_path(m_path)} (overwrite={m_exists})")
                        moments_pages_written += 1

                if args.write:
                    image_override = source_image_file if source_image_file != default_moment_filename else None
                    if update_moment_source_front_matter(
                        source_prose_path,
                        title=title,
                        status="published",
                        published_date=today.isoformat(),
                        image_file=image_override,
                        preserve_existing_published_date=True,
                    ):
                        moments_source_metadata_updated += 1

                content_html = render_markdown_with_jekyll(source_prose_path)
                moment_json_record = build_moment_json_record(moment_record)
                payload_version = compute_payload_version(compact_json_object({"moment": moment_json_record, "content_html": content_html}))
                payload = compact_json_object({
                    "header": {
                        "schema": "moment_record_v1",
                        "version": payload_version,
                        "generated_at_utc": moment_json_generated_at_utc,
                        "moment_id": moment_id,
                    },
                    "moment": moment_json_record,
                    "content_html": content_html,
                })
                out_json_path = moments_json_dir / f"{moment_id}.json"
                exists = out_json_path.exists()
                existing_version = extract_existing_header_scalar(out_json_path, "version") if exists else None
                payload_version = payload["header"]["version"]

                if (existing_version is not None) and (existing_version == payload_version) and (not args.force):
                    moments_json_skipped += 1
                else:
                    if args.write:
                        out_json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
                        print(f"{prefix_m}WRITE moment JSON: {display_path(out_json_path)}")
                        moments_json_written += 1
                    else:
                        print(f"{prefix_m}DRY-RUN: would write moment JSON {display_path(out_json_path)} (overwrite={exists})")
                        moments_json_written += 1

        if args.write and moments_source_metadata_updated > 0:
            print(f"Updated moment source front matter for {moments_source_metadata_updated} file(s).")

        if run_moments:
            print(
                f"Moment pages done. {'Would write' if not args.write else 'Wrote'}: {moments_pages_written}. Skipped: {moments_pages_skipped}."
            )
            print(
                f"Moment JSON done. {'Would write' if not args.write else 'Wrote'}: {moments_json_written}. Skipped: {moments_json_skipped}."
            )

    moments_payload: Dict[str, Dict[str, Any]] = {}
    projects_base_dir = Path(args.projects_base_dir).expanduser()
    moments_root = projects_base_dir / source_moments_root_subdir(PIPELINE_CONFIG)
    moments_images_root = projects_base_dir / source_moments_images_subdir(PIPELINE_CONFIG)

    for moment_entry in moment_source_records:
        moment_id = str(moment_entry.get("moment_id") or "").strip().lower()
        if not moment_id:
            continue

        status = normalize_status(moment_entry.get("status"))
        if status not in {"draft", "published"}:
            continue

        if not is_slug_safe(moment_id):
            raise SystemExit(f"Moment source filename must be slug-safe; got: {moment_id!r}")

        title = coerce_string(moment_entry.get("title")) or moment_id
        date_value = parse_date(moment_entry.get("date"))
        date_display = coerce_string(moment_entry.get("date_display"))
        width_px = coerce_int(moment_entry.get("width_px"))
        height_px = coerce_int(moment_entry.get("height_px"))
        image_file = coerce_string(moment_entry.get("image_file")) or f"{moment_id}.jpg"
        source_image_file = coerce_string(moment_entry.get("source_image_file")) or f"{moment_id}.jpg"
        image_alt = coerce_string(moment_entry.get("image_alt"))

        src_path: Optional[Path] = None
        source_filename = source_image_file
        if source_filename:
            source_image_path_value = coerce_string(moment_entry.get("source_image_path"))
            src_path = Path(source_image_path_value) if source_image_path_value else (moments_images_root / source_filename)

        source_image_exists = bool(src_path is not None and src_path.exists())

        if src_path is not None:
            src_w, src_h = read_image_dims_px(src_path)
            if src_w is not None and src_h is not None:
                width_px = src_w
                height_px = src_h

        images_list: List[Dict[str, Any]] = []
        if image_file is not None and image_alt is None:
            image_alt = title or moment_id
        if image_file is not None and source_image_exists:
            images_list.append(
                {
                    "file": image_file,
                    "alt": image_alt,
                }
            )

        source_prose_path_value = coerce_string(moment_entry.get("source_prose_path"))
        source_prose_path = Path(source_prose_path_value) if source_prose_path_value else (moments_root / f"{moment_id}.md")
        if not source_prose_path.exists():
            print(f"[moment {moment_id}] WARNING: missing source prose {display_projects_path(source_prose_path)}; skipping moments index.")
            continue

        moment_record = {
            "moment_id": moment_id,
            "title": title,
            "date": date_value,
            "date_display": date_display,
            "images": images_list,
            "width_px": width_px,
            "height_px": height_px,
        }
        moments_payload[moment_id] = build_moment_index_record(moment_record)

    version_payload = compact_json_object({
        "schema": "moments_index_v1",
        "moments": moments_payload,
    })
    version = compute_payload_version(version_payload)
    payload = compact_json_object({
        "header": {
            "schema": "moments_index_v1",
            "version": version,
            "generated_at_utc": utc_timestamp_now(),
            "count": len(moments_payload),
        },
        "moments": moments_payload,
    })
    payload_version = payload["header"]["version"]
    exists = moments_index_json_path.exists()
    existing_version = extract_existing_header_scalar(moments_index_json_path, "version") if exists else None
    if (existing_version is not None) and (existing_version == payload_version) and (not args.force):
        print("Moments index JSON done. Wrote: 0. Skipped: 1.")
    else:
        if args.write:
            moments_index_json_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            print(f"Moments index JSON done. Wrote: 1. Skipped: 0. Path: {display_path(moments_index_json_path)}")
        else:
            print(
                "Moments index JSON done. Would write: 1. Skipped: 0. "
                f"Path: {display_path(moments_index_json_path)} (overwrite={exists})"
            )

    log_event(
        "generate_complete",
        {
            "write": bool(args.write),
            "force": bool(args.force),
        },
    )

if __name__ == "__main__":
    try:
        main()
    except SystemExit as exc:
        code = exc.code if isinstance(exc.code, int) else (0 if exc.code is None else 1)
        log_event("generate_exit", {"status": "system_exit", "code": code})
        raise
    except Exception as exc:  # noqa: BLE001
        log_event("generate_exit", {"status": "error", "error": str(exc)})
        raise
