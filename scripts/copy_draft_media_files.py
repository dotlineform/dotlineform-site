#!/usr/bin/env python3
"""
Copy and rename draft media files from data/works.xlsx.

Modes:
- work:        projects/[project_folder]/[project_filename] -> works/make_srcset_images/[work_id][.ext]
- work_details: projects/[work.project_folder]/[project_subfolder]/[project_filename] -> work_details/make_srcset_images/[work_id]-[detail_id][.ext]
- moment:      moments/[moment_folder]/[moment_filename] -> moments/make_srcset_images/[moment_id][.ext]

Use --ids-file to limit processing to selected IDs (one per line):
- work: work_id
- work_details: detail_uid (work_id-detail_id)
- moment: moment_id

Defaults to dry-run. Pass --write to actually copy files.

Flag behavior:
- --mode: selects worksheet mapping and source/target paths.
- --ids-file: optional manifest to restrict rows processed in this run.
- --copied-ids-file: optional output manifest consumed by follow-up steps
  (typically srcset generation).
- --keep-ext / --no-ext: controls whether copied filenames keep source extension.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

import openpyxl


# Source roots:
# - projects/ for works + work_details
# - moments/ for moments
PROJECTS_BASE_DIR = Path("/Users/dlf/Library/CloudStorage/OneDrive-Personal/dotlineform")
# Destination root where make_srcset_images.sh picks up copied source files.
WORKS_BASE_DIR = Path("/Users/dlf/Library/Mobile Documents/com~apple~CloudDocs/dotlineform")
WORKBOOK_PATH = Path("data/works.xlsx")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("'") and len(s) > 1:
        s = s[1:]
    return s


def normalize_id(value: Any, width: int) -> str:
    s = normalize_text(value)
    if s.endswith(".0"):
        s = s[:-2]
    return s.zfill(width) if s.isdigit() else s


def work_id(value: Any) -> str:
    return normalize_id(value, 5)


def detail_id(value: Any) -> str:
    return normalize_id(value, 3)


def detail_uid(raw_work_id: Any, raw_detail_id: Any) -> str:
    return f"{work_id(raw_work_id)}-{detail_id(raw_detail_id)}"


def moment_id(value: Any) -> str:
    return normalize_text(value).lower()


def header_map(ws) -> Dict[str, int]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return {str(h).strip(): i for i, h in enumerate(headers) if h is not None}


def read_selected_ids(path: str, mode: str) -> Optional[Set[str]]:
    """Read ID filter file and normalize each ID to the mode-specific format."""
    if not path:
        return None
    ids_path = Path(path).expanduser()
    if not ids_path.exists():
        raise SystemExit(f"ids file not found: {ids_path}")
    out: Set[str] = set()
    for line in ids_path.read_text(encoding="utf-8").splitlines():
        raw = line.strip()
        if not raw:
            continue
        if mode == "work":
            v = work_id(raw)
        elif mode == "work_details":
            parts = raw.split("-", 1)
            if len(parts) == 2:
                v = f"{work_id(parts[0])}-{detail_id(parts[1])}"
            else:
                v = raw
        else:
            v = moment_id(raw)
        if v:
            out.add(v)
    return out


def copy_work(ws, cols, selected_ids: Optional[Set[str]], keep_ext: bool, write: bool, copied_ids: List[str]) -> tuple[int, int, int]:
    """Copy base work source images to works/make_srcset_images."""
    required = ["work_id", "project_folder", "project_filename"]
    missing = [c for c in required if c not in cols]
    if missing:
        raise SystemExit(f"Missing required columns in Works sheet: {', '.join(missing)}")

    dest_dir = WORKS_BASE_DIR / "works/make_srcset_images"
    if write:
        dest_dir.mkdir(parents=True, exist_ok=True)

    total = copied = missing_src = 0
    for row_cells in ws.iter_rows(min_row=2):
        row = [cell.value for cell in row_cells]
        wid = work_id(row[cols["work_id"]])
        folder = row[cols["project_folder"]]
        filename = row[cols["project_filename"]]
        if not (wid and folder and filename):
            continue
        if selected_ids is not None and wid not in selected_ids:
            continue

        src = PROJECTS_BASE_DIR / "projects" / str(folder).strip() / str(filename).strip()
        ext = Path(str(filename).strip()).suffix
        dest = (dest_dir / f"{wid}{ext}") if keep_ext else (dest_dir / wid)

        total += 1
        if not src.exists():
            print(f"Missing source: {src}")
            missing_src += 1
            continue

        copied_ids.append(wid)
        if write:
            from shutil import copy2

            copy2(src, dest)
            print(f"Copied: {src} -> {dest}")
            copied += 1
        else:
            print(f"Dry-run: {src} -> {dest}")
    return total, copied, missing_src


def copy_work_details(wb, selected_ids: Optional[Set[str]], keep_ext: bool, write: bool, copied_ids: List[str]) -> tuple[int, int, int]:
    """Copy work detail source images to work_details/make_srcset_images."""
    if "Works" not in wb.sheetnames or "WorkDetails" not in wb.sheetnames:
        raise SystemExit("Works and WorkDetails worksheets are required for mode=work_details")

    works_ws = wb["Works"]
    details_ws = wb["WorkDetails"]
    works_cols = header_map(works_ws)
    details_cols = header_map(details_ws)

    missing_works = [c for c in ["work_id", "project_folder"] if c not in works_cols]
    if missing_works:
        raise SystemExit(f"Missing required columns in Works sheet: {', '.join(missing_works)}")
    missing_details = [c for c in ["work_id", "detail_id", "project_subfolder", "project_filename"] if c not in details_cols]
    if missing_details:
        raise SystemExit(f"Missing required columns in WorkDetails sheet: {', '.join(missing_details)}")

    work_folder_by_id: Dict[str, str] = {}
    for row_cells in works_ws.iter_rows(min_row=2):
        row = [cell.value for cell in row_cells]
        wid = work_id(row[works_cols["work_id"]])
        folder = row[works_cols["project_folder"]]
        if wid and folder:
            work_folder_by_id[wid] = str(folder).strip()

    dest_dir = WORKS_BASE_DIR / "work_details/make_srcset_images"
    if write:
        dest_dir.mkdir(parents=True, exist_ok=True)

    total = copied = missing_src = 0
    for row_cells in details_ws.iter_rows(min_row=2):
        row = [cell.value for cell in row_cells]
        raw_wid = row[details_cols["work_id"]]
        raw_did = row[details_cols["detail_id"]]
        subfolder = row[details_cols["project_subfolder"]]
        filename = row[details_cols["project_filename"]]
        if not (raw_wid and raw_did and filename is not None):
            continue

        uid = detail_uid(raw_wid, raw_did)
        if selected_ids is not None and uid not in selected_ids:
            continue

        wid = work_id(raw_wid)
        folder = work_folder_by_id.get(wid)
        if not folder:
            print(f"Warning: missing Works.project_folder for work_id={wid}; skipping detail {uid}")
            continue

        subfolder_str = str(subfolder).strip() if subfolder else ""
        filename_str = str(filename).strip()
        src = PROJECTS_BASE_DIR / "projects" / folder
        if subfolder_str:
            src = src / subfolder_str
        src = src / filename_str

        ext = Path(filename_str).suffix
        dest = (dest_dir / f"{uid}{ext}") if keep_ext else (dest_dir / uid)
        total += 1
        if not src.exists():
            print(f"Missing source: {src}")
            missing_src += 1
            continue

        copied_ids.append(uid)
        if write:
            from shutil import copy2

            copy2(src, dest)
            print(f"Copied: {src} -> {dest}")
            copied += 1
        else:
            print(f"Dry-run: {src} -> {dest}")
    return total, copied, missing_src


def copy_moments(ws, cols, selected_ids: Optional[Set[str]], keep_ext: bool, write: bool, copied_ids: List[str]) -> tuple[int, int, int]:
    """Copy moment source images to moments/make_srcset_images."""
    required = ["moment_id", "moment_folder", "moment_filename"]
    missing = [c for c in required if c not in cols]
    if missing:
        raise SystemExit(f"Missing required columns in Moments sheet: {', '.join(missing)}")

    dest_dir = WORKS_BASE_DIR / "moments/make_srcset_images"
    if write:
        dest_dir.mkdir(parents=True, exist_ok=True)

    total = copied = missing_src = 0
    for row_cells in ws.iter_rows(min_row=2):
        row = [cell.value for cell in row_cells]
        mid = moment_id(row[cols["moment_id"]])
        folder = row[cols["moment_folder"]]
        filename = row[cols["moment_filename"]]
        if not (mid and folder and filename):
            continue
        if selected_ids is not None and mid not in selected_ids:
            continue

        src = PROJECTS_BASE_DIR / "moments" / str(folder).strip() / str(filename).strip()
        ext = Path(str(filename).strip()).suffix
        dest = (dest_dir / f"{mid}{ext}") if keep_ext else (dest_dir / mid)
        total += 1
        if not src.exists():
            print(f"Missing source: {src}")
            missing_src += 1
            continue

        copied_ids.append(mid)
        if write:
            from shutil import copy2

            copy2(src, dest)
            print(f"Copied: {src} -> {dest}")
            copied += 1
        else:
            print(f"Dry-run: {src} -> {dest}")
    return total, copied, missing_src


def main() -> int:
    parser = argparse.ArgumentParser(description="Copy draft media files from works.xlsx")
    parser.add_argument(
        "--mode",
        choices=["work", "work_details", "moment"],
        required=True,
        help="Copy mode: work|work_details|moment",
    )
    parser.add_argument("--write", action="store_true", help="Actually copy files (default: dry-run)")
    parser.add_argument(
        "--keep-ext",
        action="store_true",
        help="Keep source extension in target filename (default behavior).",
    )
    parser.add_argument(
        "--no-ext",
        action="store_true",
        help="Drop extension in target filename (normally not needed).",
    )
    parser.add_argument(
        "--ids-file",
        default="",
        help="Optional ID filter file (one ID per line). Format depends on --mode.",
    )
    parser.add_argument(
        "--copied-ids-file",
        default="",
        help="Optional output manifest of copied IDs (one ID per line).",
    )
    args = parser.parse_args()

    keep_ext = not args.no_ext
    if args.keep_ext:
        keep_ext = True

    wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=False, data_only=True)
    selected_ids = read_selected_ids(args.ids_file, args.mode)
    copied_ids: List[str] = []

    if args.mode == "work":
        if "Works" not in wb.sheetnames:
            raise SystemExit("Worksheet not found: 'Works'")
        ws = wb["Works"]
        total, copied, missing_src = copy_work(ws, header_map(ws), selected_ids, keep_ext, args.write, copied_ids)
        label = "Work rows"
    elif args.mode == "work_details":
        total, copied, missing_src = copy_work_details(wb, selected_ids, keep_ext, args.write, copied_ids)
        label = "Detail rows"
    else:
        if "Moments" not in wb.sheetnames:
            raise SystemExit("Worksheet not found: 'Moments'")
        ws = wb["Moments"]
        total, copied, missing_src = copy_moments(ws, header_map(ws), selected_ids, keep_ext, args.write, copied_ids)
        label = "Moment rows"

    if args.copied_ids_file:
        ids_path = Path(args.copied_ids_file).expanduser()
        ids_path.parent.mkdir(parents=True, exist_ok=True)
        ids_path.write_text("\n".join(copied_ids) + ("\n" if copied_ids else ""), encoding="utf-8")
        print(f"Wrote copied IDs manifest: {ids_path} ({len(copied_ids)} ids)")

    print(f"{label}: {total}, copied: {copied}, missing source: {missing_src}")
    if not args.write:
        print("Dry-run only. Re-run with --write to copy files.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
