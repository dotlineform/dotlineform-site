#!/usr/bin/env python3
"""
Backfill assets/data/recent_index.json from git history of data/works.xlsx.

This script reconstructs genuine draft -> published transitions for Works and
Series rows by comparing historical workbook snapshots across commits that
changed the workbook. It writes the same recent_index_v1 payload used by the
public /recent/ page.

Rules mirrored from the live generator:
- record only first-time draft -> published transitions
- series entries suppress work entries from the same newly published series
- multiple newly published works in the same existing series collapse to one
  entry anchored to the first work row from that commit
- prune entries whose current work/series target no longer exists

The history inference is intentionally strict. Rows added directly as
"published" are ignored because there is no provable draft -> published
transition in git history.
"""

from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import openpyxl

try:
    from generate_work_pages import (
        build_recent_index_payload,
        coerce_string,
        normalize_recent_entry,
        normalize_status,
        normalize_text,
        parse_date,
        slug_id,
    )
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.generate_work_pages import (
        build_recent_index_payload,
        coerce_string,
        normalize_recent_entry,
        normalize_status,
        normalize_text,
        parse_date,
        slug_id,
    )

try:
    from series_ids import normalize_series_id, parse_series_ids
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.series_ids import normalize_series_id, parse_series_ids


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX_PATH = Path("data/works.xlsx")
DEFAULT_OUTPUT_PATH = Path("assets/data/recent_index.json")


@dataclass(frozen=True)
class CommitRef:
    sha: str
    committed_at_utc: str


@dataclass
class WorkSnapshot:
    work_id: str
    status: str
    title: Optional[str]
    published_date: Optional[str]
    series_ids: List[str]
    row_order: int


@dataclass
class SeriesSnapshot:
    series_id: str
    status: str
    title: Optional[str]
    published_date: Optional[str]
    primary_work_id: Optional[str]
    work_ids: List[str]


@dataclass
class CatalogueSnapshot:
    works: Dict[str, WorkSnapshot]
    series: Dict[str, SeriesSnapshot]


def header_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())):
        key = normalize_text(value).lower()
        if key and key not in headers:
            headers[key] = idx
    return headers


def row_value(row: Sequence[Any], headers: Dict[str, int], key: str) -> Any:
    idx = headers.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parse_workbook_snapshot_from_bytes(blob: bytes) -> CatalogueSnapshot:
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    if "Works" not in wb.sheetnames or "Series" not in wb.sheetnames:
        raise ValueError("Workbook snapshot missing Works or Series sheet")

    works_ws = wb["Works"]
    series_ws = wb["Series"]
    works_hi = header_map(works_ws)
    series_hi = header_map(series_ws)

    if "work_id" not in works_hi or "status" not in works_hi:
        raise ValueError("Works sheet missing required columns")
    if "series_id" not in series_hi or "status" not in series_hi:
        raise ValueError("Series sheet missing required columns")

    works: Dict[str, WorkSnapshot] = {}
    work_ids_by_series: Dict[str, List[str]] = {}
    for row_index, row in enumerate(works_ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_work_id = row_value(row, works_hi, "work_id")
        if normalize_text(raw_work_id) == "":
            continue
        try:
            work_id = slug_id(raw_work_id)
        except ValueError:
            continue
        series_ids_raw = row_value(row, works_hi, "series_ids")
        try:
            series_ids = parse_series_ids(series_ids_raw)
        except ValueError:
            series_ids = []
        for series_id in series_ids:
            work_ids_by_series.setdefault(series_id, []).append(work_id)
        works[work_id] = WorkSnapshot(
            work_id=work_id,
            status=normalize_status(row_value(row, works_hi, "status")),
            title=coerce_string(row_value(row, works_hi, "title")),
            published_date=parse_date(row_value(row, works_hi, "published_date")),
            series_ids=series_ids,
            row_order=row_index,
        )

    series: Dict[str, SeriesSnapshot] = {}
    for row in series_ws.iter_rows(min_row=2, values_only=True):
        raw_series_id = row_value(row, series_hi, "series_id")
        if normalize_text(raw_series_id) == "":
            continue
        try:
            series_id = normalize_series_id(raw_series_id)
        except ValueError:
            continue
        try:
            primary_work_id = slug_id(row_value(row, series_hi, "primary_work_id"))
        except ValueError:
            primary_work_id = None
        series[series_id] = SeriesSnapshot(
            series_id=series_id,
            status=normalize_status(row_value(row, series_hi, "status")),
            title=coerce_string(row_value(row, series_hi, "title")),
            published_date=parse_date(row_value(row, series_hi, "published_date")),
            primary_work_id=primary_work_id,
            work_ids=work_ids_by_series.get(series_id, []),
        )

    return CatalogueSnapshot(works=works, series=series)


def load_current_snapshot(xlsx_path: Path) -> CatalogueSnapshot:
    return parse_workbook_snapshot_from_bytes(xlsx_path.read_bytes())


def load_snapshot_from_commit(repo_root: Path, commit_sha: str, xlsx_path: Path) -> CatalogueSnapshot:
    result = subprocess.run(
        ["git", "-C", str(repo_root), "show", f"{commit_sha}:{xlsx_path.as_posix()}"],
        check=True,
        capture_output=True,
    )
    return parse_workbook_snapshot_from_bytes(result.stdout)


def iter_workbook_commits(repo_root: Path, xlsx_path: Path) -> List[CommitRef]:
    result = subprocess.run(
        [
            "git",
            "-C",
            str(repo_root),
            "log",
            "--reverse",
            "--format=%H%x09%cI",
            "--",
            xlsx_path.as_posix(),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    commits: List[CommitRef] = []
    for line in result.stdout.splitlines():
        sha, _, committed_at_utc = line.partition("\t")
        sha = sha.strip()
        committed_at_utc = committed_at_utc.strip()
        if sha and committed_at_utc:
            commits.append(CommitRef(sha=sha, committed_at_utc=committed_at_utc))
    return commits


def series_title_key(value: Optional[str]) -> str:
    return normalize_text(value).casefold()


def map_historical_series_id(
    historical: SeriesSnapshot,
    current_series_by_id: Dict[str, SeriesSnapshot],
) -> Optional[str]:
    if historical.series_id in current_series_by_id:
        return historical.series_id

    historical_work_ids = set(historical.work_ids)
    historical_title_key = series_title_key(historical.title)
    candidates: List[Tuple[int, str]] = []

    for current_id, current in current_series_by_id.items():
        current_work_ids = set(current.work_ids)
        title_matches = historical_title_key != "" and historical_title_key == series_title_key(current.title)
        primary_matches = (
            historical.primary_work_id is not None
            and current.primary_work_id is not None
            and historical.primary_work_id == current.primary_work_id
        )
        exact_work_match = bool(historical_work_ids) and historical_work_ids == current_work_ids
        subset_work_match = bool(historical_work_ids) and historical_work_ids.issubset(current_work_ids)

        qualifies = (
            exact_work_match
            or (primary_matches and (title_matches or subset_work_match or not historical_work_ids))
            or (title_matches and subset_work_match)
        )
        if not qualifies:
            continue

        score = 0
        if primary_matches:
            score += 10
        if title_matches:
            score += 6
        if exact_work_match:
            score += 8
        elif subset_work_match:
            score += 3
        candidates.append((score, current_id))

    if not candidates:
        return None

    candidates.sort(key=lambda item: (-item[0], item[1]))
    best_score, best_id = candidates[0]
    if len(candidates) > 1 and candidates[1][0] == best_score:
        return None
    return best_id


def first_publish_date(row_date: Optional[str], commit_date: str) -> str:
    return row_date or commit_date


def normalize_commit_timestamp(value: str) -> str:
    dt_value = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    return dt_value.astimezone(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def build_backfilled_entries(
    *,
    commits: Sequence[CommitRef],
    repo_root: Path,
    xlsx_path: Path,
    current_snapshot: CatalogueSnapshot,
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    entries: List[Dict[str, Any]] = []
    seen_entry_ids: Set[str] = set()
    stats = {
        "commits_scanned": 0,
        "commits_skipped_invalid": 0,
        "series_events": 0,
        "work_events": 0,
        "series_unmapped": 0,
        "work_pruned_missing": 0,
        "series_pruned_missing": 0,
    }

    previous_snapshot: Optional[CatalogueSnapshot] = None
    for commit in commits:
        stats["commits_scanned"] += 1
        try:
            current = load_snapshot_from_commit(repo_root, commit.sha, xlsx_path)
        except Exception:
            stats["commits_skipped_invalid"] += 1
            continue
        if previous_snapshot is None:
            previous_snapshot = current
            continue

        recorded_at_utc = normalize_commit_timestamp(commit.committed_at_utc)
        commit_date = recorded_at_utc[:10]
        session_order = 0
        newly_published_series_ids: Set[str] = set()

        for series_id in sorted(current.series.keys()):
            current_series = current.series[series_id]
            previous_series = previous_snapshot.series.get(series_id)
            if previous_series is None:
                continue
            if previous_series.status != "draft" or current_series.status != "published":
                continue

            target_series_id = map_historical_series_id(current_series, current_snapshot.series)
            if target_series_id is None or target_series_id not in current_snapshot.series:
                stats["series_unmapped"] += 1
                newly_published_series_ids.add(series_id)
                continue

            session_order += 1
            entry = normalize_recent_entry({
                "kind": "series",
                "target_id": target_series_id,
                "title": coerce_string(current_series.title) or target_series_id,
                "caption": f"{len(current_series.work_ids)} work{'s' if len(current_series.work_ids) != 1 else ''}",
                "published_date": first_publish_date(current_series.published_date, commit_date),
                "thumb_id": current_series.primary_work_id,
                "recorded_at_utc": recorded_at_utc,
                "session_order": session_order,
            })
            newly_published_series_ids.add(series_id)
            if entry is None:
                continue
            entry_id = str(entry.get("id") or "")
            if entry_id in seen_entry_ids:
                continue
            seen_entry_ids.add(entry_id)
            entries.append(entry)
            stats["series_events"] += 1
            session_order += 1

        grouped_work_transitions: Dict[str, List[WorkSnapshot]] = {}
        for work_id in sorted(current.works.keys()):
            current_work = current.works[work_id]
            previous_work = previous_snapshot.works.get(work_id)
            if previous_work is None:
                continue
            if previous_work.status != "draft" or current_work.status != "published":
                continue
            primary_series_id = current_work.series_ids[0] if current_work.series_ids else ""
            if not primary_series_id or primary_series_id in newly_published_series_ids:
                continue
            grouped_work_transitions.setdefault(primary_series_id, []).append(current_work)

        for primary_series_id in sorted(grouped_work_transitions.keys()):
            group = grouped_work_transitions[primary_series_id]
            if not group:
                continue
            group.sort(key=lambda item: (item.row_order, item.work_id))
            first_work = group[0]
            if first_work.work_id not in current_snapshot.works:
                stats["work_pruned_missing"] += 1
                continue
            new_work_count = len(group)
            historical_series = current.series.get(primary_series_id)
            series_title = (
                historical_series.title
                if historical_series is not None and historical_series.title
                else primary_series_id
            )
            caption = (
                series_title
                if new_work_count == 1
                else f"{new_work_count} new works in {series_title}"
            )
            session_order += 1
            entry = normalize_recent_entry({
                "kind": "work",
                "target_id": first_work.work_id,
                "title": coerce_string(first_work.title) or first_work.work_id,
                "caption": caption,
                "published_date": first_publish_date(first_work.published_date, commit_date),
                "thumb_id": first_work.work_id,
                "recorded_at_utc": recorded_at_utc,
                "session_order": session_order,
            })
            if entry is None:
                continue
            entry_id = str(entry.get("id") or "")
            if entry_id in seen_entry_ids:
                continue
            seen_entry_ids.add(entry_id)
            entries.append(entry)
            stats["work_events"] += 1
            session_order += 1

        previous_snapshot = current

    retained_entries: List[Dict[str, Any]] = []
    for entry in entries:
        kind = str(entry.get("kind") or "")
        target_id = str(entry.get("target_id") or "")
        if kind == "work":
            if target_id not in current_snapshot.works:
                stats["work_pruned_missing"] += 1
                continue
        elif kind == "series":
            if target_id not in current_snapshot.series:
                stats["series_pruned_missing"] += 1
                continue
        retained_entries.append(entry)

    return retained_entries, stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Backfill recent_index.json from git history of data/works.xlsx.")
    parser.add_argument("--repo-root", default=str(REPO_ROOT), help="Repository root")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX_PATH), help="Workbook path relative to repo root")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_PATH), help="Recent index output path relative to repo root")
    parser.add_argument("--write", action="store_true", help="Write the generated payload")
    parser.add_argument("--limit", type=int, default=10, help="Preview count to print")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(args.repo_root).expanduser().resolve()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = (repo_root / xlsx_path).resolve()
    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = (repo_root / output_path).resolve()

    current_snapshot = load_current_snapshot(xlsx_path)
    commits = iter_workbook_commits(repo_root, xlsx_path.relative_to(repo_root))
    entries, stats = build_backfilled_entries(
        commits=commits,
        repo_root=repo_root,
        xlsx_path=xlsx_path.relative_to(repo_root),
        current_snapshot=current_snapshot,
    )
    payload = build_recent_index_payload(entries)

    print(
        json.dumps(
            {
                "commits_scanned": stats["commits_scanned"],
                "commits_skipped_invalid": stats["commits_skipped_invalid"],
                "series_events": stats["series_events"],
                "work_events": stats["work_events"],
                "series_unmapped": stats["series_unmapped"],
                "series_pruned_missing": stats["series_pruned_missing"],
                "work_pruned_missing": stats["work_pruned_missing"],
                "final_count": payload.get("header", {}).get("count"),
                "output": str(output_path.relative_to(repo_root)),
            },
            ensure_ascii=False,
            indent=2,
        )
    )

    preview_entries = payload.get("entries", [])[: max(args.limit, 0)]
    if preview_entries:
        print("\nPreview:")
        for entry in preview_entries:
            print(
                "- "
                + " | ".join(
                    [
                        str(entry.get("published_date") or ""),
                        str(entry.get("kind") or ""),
                        str(entry.get("target_id") or ""),
                        str(entry.get("title") or ""),
                        str(entry.get("caption") or ""),
                    ]
                )
            )
    else:
        print("\nPreview: no entries")

    if args.write:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"\nWrote {output_path.relative_to(repo_root)}")
    else:
        print(f"\nDry run only. Would write {output_path.relative_to(repo_root)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
