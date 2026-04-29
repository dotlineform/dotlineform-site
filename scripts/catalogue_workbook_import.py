from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping

try:
    from pipeline_config import bulk_import_workbook_path, load_pipeline_config
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.pipeline_config import bulk_import_workbook_path, load_pipeline_config

try:
    from catalogue_source import (
        CatalogueSourceRecords,
        DETAIL_FIELDS,
        DETAIL_TEXT_FIELDS,
        WORK_FIELDS,
        WORK_TEXT_FIELDS,
        cell,
        header_map,
        normalize_json_value,
        normalize_scalar_text,
        normalize_source_record,
        normalize_status,
        parse_series_ids,
        records_from_json_source,
        slug_id,
        sort_record_map,
        validate_source_records,
    )
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.catalogue_source import (
        CatalogueSourceRecords,
        DETAIL_FIELDS,
        DETAIL_TEXT_FIELDS,
        WORK_FIELDS,
        WORK_TEXT_FIELDS,
        cell,
        header_map,
        normalize_json_value,
        normalize_scalar_text,
        normalize_source_record,
        normalize_status,
        parse_series_ids,
        records_from_json_source,
        slug_id,
        sort_record_map,
        validate_source_records,
    )

try:
    from series_ids import normalize_series_id
except ModuleNotFoundError:  # pragma: no cover - package import fallback
    from scripts.series_ids import normalize_series_id


PIPELINE_CONFIG = load_pipeline_config(Path(__file__))
DEFAULT_IMPORT_WORKBOOK_PATH = bulk_import_workbook_path(PIPELINE_CONFIG)
IMPORT_MODE_WORKS = "works"
IMPORT_MODE_WORK_DETAILS = "work_details"
VALID_IMPORT_MODES = {IMPORT_MODE_WORKS, IMPORT_MODE_WORK_DETAILS}
PREVIEW_SAMPLE_LIMIT = 20
BLOCKED_SAMPLE_LIMIT = 40


@dataclass(frozen=True)
class WorkbookImportPlan:
    mode: str
    workbook_path: Path
    total_candidate_rows: int
    importable_records: Dict[str, Dict[str, Any]]
    duplicate_ids: list[str]
    blocked_rows: list[Dict[str, str]]
    blocked_reason_counts: Dict[str, int]
    validation_errors: list[str]

    @property
    def target_kind(self) -> str:
        return "works" if self.mode == IMPORT_MODE_WORKS else "work_details"

    @property
    def importable_count(self) -> int:
        return len(self.importable_records)

    @property
    def duplicate_count(self) -> int:
        return len(self.duplicate_ids)

    @property
    def blocked_count(self) -> int:
        return len(self.blocked_rows) + len(self.validation_errors)


def normalize_import_mode(raw_mode: Any) -> str:
    mode = str(raw_mode or "").strip().lower()
    if mode not in VALID_IMPORT_MODES:
        raise ValueError("import mode must be one of: works, work_details")
    return mode


def build_workbook_import_plan(source_dir: Path, workbook_path: Path, mode: str) -> WorkbookImportPlan:
    normalized_mode = normalize_import_mode(mode)
    source_records = records_from_json_source(source_dir)
    workbook = _load_workbook(workbook_path)

    if normalized_mode == IMPORT_MODE_WORKS:
        return _build_work_import_plan(source_records, workbook, workbook_path)
    return _build_work_detail_import_plan(source_records, workbook, workbook_path)


def apply_workbook_import_plan(source_dir: Path, plan: WorkbookImportPlan) -> CatalogueSourceRecords:
    if plan.blocked_count > 0:
        raise ValueError("import preview contains blocked rows; fix the workbook before apply")

    source_records = records_from_json_source(source_dir)
    if plan.target_kind == "works":
        merged = dict(source_records.works)
        merged.update(plan.importable_records)
        return CatalogueSourceRecords(
            works=sort_record_map(merged),
            work_details=source_records.work_details,
            series=source_records.series,
            work_files=source_records.work_files,
            work_links=source_records.work_links,
        )

    merged = dict(source_records.work_details)
    merged.update(plan.importable_records)
    return CatalogueSourceRecords(
        works=source_records.works,
        work_details=sort_record_map(merged),
        series=source_records.series,
        work_files=source_records.work_files,
        work_links=source_records.work_links,
    )


def plan_to_response(plan: WorkbookImportPlan, repo_root: Path | None = None) -> Dict[str, Any]:
    workbook_display = str(plan.workbook_path)
    if repo_root is not None:
        try:
            workbook_display = str(plan.workbook_path.resolve().relative_to(repo_root.resolve()))
        except ValueError:
            workbook_display = str(plan.workbook_path.resolve())

    return {
        "mode": plan.mode,
        "workbook_path": workbook_display,
        "target_kind": plan.target_kind,
        "summary": {
            "candidate_rows": plan.total_candidate_rows,
            "importable_count": plan.importable_count,
            "duplicate_count": plan.duplicate_count,
            "blocked_count": plan.blocked_count,
        },
        "importable_ids": _sample_ids(plan.importable_records.keys()),
        "duplicates": {
            "count": plan.duplicate_count,
            "sample_ids": _sample_ids(plan.duplicate_ids),
            "truncated": max(0, plan.duplicate_count - PREVIEW_SAMPLE_LIMIT),
        },
        "blocked": {
            "count": plan.blocked_count,
            "reason_counts": dict(sorted(plan.blocked_reason_counts.items())),
            "rows": plan.blocked_rows[:BLOCKED_SAMPLE_LIMIT],
            "validation_errors": plan.validation_errors[:BLOCKED_SAMPLE_LIMIT],
            "truncated": max(0, len(plan.blocked_rows) - BLOCKED_SAMPLE_LIMIT) + max(0, len(plan.validation_errors) - BLOCKED_SAMPLE_LIMIT),
        },
        "ready_to_apply": plan.blocked_count == 0 and plan.importable_count > 0,
    }


def _load_workbook(workbook_path: Path):
    import openpyxl

    resolved_path = workbook_path.expanduser().resolve()
    if not resolved_path.exists():
        raise ValueError(f"Workbook not found: {resolved_path}")
    return openpyxl.load_workbook(resolved_path, read_only=True, data_only=True)


def _require_sheet(wb, sheet_name: str) -> tuple[list[tuple[Any, ...]], Dict[str, int]]:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found in workbook: {sheet_name}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet is empty: {sheet_name}")
    return rows, header_map(rows[0])


def _sample_ids(ids: Iterable[str]) -> list[str]:
    out: list[str] = []
    for value in ids:
        text = str(value).strip()
        if not text:
            continue
        out.append(text)
        if len(out) >= PREVIEW_SAMPLE_LIMIT:
            break
    return out


def _append_blocked(
    blocked_rows: list[Dict[str, str]],
    reason_counts: Dict[str, int],
    *,
    row_number: int,
    record_id: str,
    reason: str,
    message: str,
) -> None:
    blocked_rows.append(
        {
            "row_number": str(row_number),
            "id": record_id,
            "reason": reason,
            "message": message,
        }
    )
    reason_counts[reason] = reason_counts.get(reason, 0) + 1


def _build_work_import_plan(source_records: CatalogueSourceRecords, workbook, workbook_path: Path) -> WorkbookImportPlan:
    rows, headers = _require_sheet(workbook, "Works")
    _require_headers(headers, ["work_id", "series_ids", "title"], sheet_name="Works")

    importable: Dict[str, Dict[str, Any]] = {}
    duplicate_ids: list[str] = []
    blocked_rows: list[Dict[str, str]] = []
    blocked_reason_counts: Dict[str, int] = {}
    seen_work_ids: set[str] = set()
    total_candidate_rows = 0
    known_series_ids = set(source_records.series.keys())

    for row_number, row in enumerate(rows[1:], start=2):
        if _row_has_no_value(row):
            continue
        raw_work_id = cell(row, headers, "work_id")
        if raw_work_id in {None, ""}:
            continue
        total_candidate_rows += 1

        try:
            work_id = slug_id(raw_work_id)
        except ValueError as exc:
            _append_blocked(blocked_rows, blocked_reason_counts, row_number=row_number, record_id=str(raw_work_id or ""), reason="invalid_work_id", message=str(exc))
            continue

        if work_id in seen_work_ids:
            _append_blocked(blocked_rows, blocked_reason_counts, row_number=row_number, record_id=work_id, reason="duplicate_workbook_id", message="work_id is duplicated in the workbook")
            continue
        seen_work_ids.add(work_id)

        if work_id in source_records.works:
            duplicate_ids.append(work_id)
            continue

        title = normalize_scalar_text(cell(row, headers, "title"))
        if not title:
            _append_blocked(blocked_rows, blocked_reason_counts, row_number=row_number, record_id=work_id, reason="missing_title", message="title is required")
            continue

        try:
            series_ids = parse_series_ids(cell(row, headers, "series_ids"))
        except ValueError as exc:
            _append_blocked(
                blocked_rows,
                blocked_reason_counts,
                row_number=row_number,
                record_id=work_id,
                reason="invalid_series_ids",
                message=str(exc),
            )
            continue
        if not series_ids:
            _append_blocked(blocked_rows, blocked_reason_counts, row_number=row_number, record_id=work_id, reason="missing_series_ids", message="series_ids is required")
            continue
        missing_series_ids = [series_id for series_id in series_ids if normalize_series_id(series_id) not in known_series_ids]
        if missing_series_ids:
            _append_blocked(
                blocked_rows,
                blocked_reason_counts,
                row_number=row_number,
                record_id=work_id,
                reason="unknown_series",
                message=f"unknown series ids: {', '.join(sorted(set(missing_series_ids)))}",
            )
            continue

        record = {
            "work_id": work_id,
            "status": "draft",
            "published_date": None,
            "series_ids": [normalize_series_id(series_id) for series_id in series_ids],
        }
        for field in WORK_FIELDS:
            if field in record or field in {"status", "published_date"}:
                continue
            record[field] = normalize_json_value(cell(row, headers, field))
        normalized_record = normalize_source_record(record, WORK_FIELDS, text_fields=WORK_TEXT_FIELDS)
        importable[work_id] = normalized_record

    validation_errors = _validate_imported_records(source_records, importable, mode=IMPORT_MODE_WORKS)
    return WorkbookImportPlan(
        mode=IMPORT_MODE_WORKS,
        workbook_path=workbook_path.expanduser().resolve(),
        total_candidate_rows=total_candidate_rows,
        importable_records=sort_record_map(importable),
        duplicate_ids=sorted(duplicate_ids),
        blocked_rows=blocked_rows,
        blocked_reason_counts=blocked_reason_counts,
        validation_errors=validation_errors,
    )


def _build_work_detail_import_plan(source_records: CatalogueSourceRecords, workbook, workbook_path: Path) -> WorkbookImportPlan:
    rows, headers = _require_sheet(workbook, "WorkDetails")
    _require_headers(headers, ["work_id", "detail_id", "title"], sheet_name="WorkDetails")

    importable: Dict[str, Dict[str, Any]] = {}
    duplicate_ids: list[str] = []
    blocked_rows: list[Dict[str, str]] = []
    blocked_reason_counts: Dict[str, int] = {}
    seen_detail_uids: set[str] = set()
    total_candidate_rows = 0
    for row_number, row in enumerate(rows[1:], start=2):
        if _row_has_no_value(row):
            continue
        raw_work_id = cell(row, headers, "work_id")
        raw_detail_id = cell(row, headers, "detail_id")
        if raw_work_id in {None, ""} and raw_detail_id in {None, ""}:
            continue
        total_candidate_rows += 1

        try:
            work_id = slug_id(raw_work_id)
            detail_id = slug_id(raw_detail_id, width=3)
        except ValueError as exc:
            _append_blocked(blocked_rows, blocked_reason_counts, row_number=row_number, record_id=f"{raw_work_id or ''}-{raw_detail_id or ''}".strip("-"), reason="invalid_detail_uid", message=str(exc))
            continue

        detail_uid = f"{work_id}-{detail_id}"
        if detail_uid in seen_detail_uids:
            _append_blocked(blocked_rows, blocked_reason_counts, row_number=row_number, record_id=detail_uid, reason="duplicate_workbook_id", message="detail_uid is duplicated in the workbook")
            continue
        seen_detail_uids.add(detail_uid)

        if detail_uid in source_records.work_details:
            duplicate_ids.append(detail_uid)
            continue

        parent_work = source_records.works.get(work_id)
        if not isinstance(parent_work, dict):
            _append_blocked(blocked_rows, blocked_reason_counts, row_number=row_number, record_id=detail_uid, reason="unknown_work", message=f"parent work_id not found: {work_id}")
            continue
        if normalize_status(parent_work.get("status")) != "published":
            _append_blocked(blocked_rows, blocked_reason_counts, row_number=row_number, record_id=detail_uid, reason="parent_work_unpublished", message=f"parent work {work_id} must be published before adding work details")
            continue

        title = normalize_scalar_text(cell(row, headers, "title"))
        if not title:
            _append_blocked(blocked_rows, blocked_reason_counts, row_number=row_number, record_id=detail_uid, reason="missing_title", message="title is required")
            continue

        record = {
            "detail_uid": detail_uid,
            "work_id": work_id,
            "detail_id": detail_id,
            "status": "draft",
            "published_date": None,
        }
        for field in DETAIL_FIELDS:
            if field in record or field in {"status", "published_date"}:
                continue
            record[field] = normalize_json_value(cell(row, headers, field))
        normalized_record = normalize_source_record(record, DETAIL_FIELDS, text_fields=DETAIL_TEXT_FIELDS)
        importable[detail_uid] = normalized_record

    validation_errors = _validate_imported_records(source_records, importable, mode=IMPORT_MODE_WORK_DETAILS)
    return WorkbookImportPlan(
        mode=IMPORT_MODE_WORK_DETAILS,
        workbook_path=workbook_path.expanduser().resolve(),
        total_candidate_rows=total_candidate_rows,
        importable_records=sort_record_map(importable),
        duplicate_ids=sorted(duplicate_ids),
        blocked_rows=blocked_rows,
        blocked_reason_counts=blocked_reason_counts,
        validation_errors=validation_errors,
    )


def _validate_imported_records(
    source_records: CatalogueSourceRecords,
    importable_records: Mapping[str, Dict[str, Any]],
    *,
    mode: str,
) -> list[str]:
    if not importable_records:
        return []

    if mode == IMPORT_MODE_WORKS:
        merged = dict(source_records.works)
        merged.update(importable_records)
        candidate_records = CatalogueSourceRecords(
            works=sort_record_map(merged),
            work_details=source_records.work_details,
            series=source_records.series,
            work_files=source_records.work_files,
            work_links=source_records.work_links,
        )
    else:
        merged = dict(source_records.work_details)
        merged.update(importable_records)
        candidate_records = CatalogueSourceRecords(
            works=source_records.works,
            work_details=sort_record_map(merged),
            series=source_records.series,
            work_files=source_records.work_files,
            work_links=source_records.work_links,
        )

    return validate_source_records(candidate_records)


def _require_headers(headers: Mapping[str, int], names: Iterable[str], *, sheet_name: str) -> None:
    missing = [name for name in names if name not in headers]
    if missing:
        raise ValueError(f"{sheet_name} sheet missing required columns: {', '.join(missing)}")


def _row_has_no_value(row: Iterable[Any]) -> bool:
    for value in row:
        if value not in {None, ""}:
            return False
    return True
